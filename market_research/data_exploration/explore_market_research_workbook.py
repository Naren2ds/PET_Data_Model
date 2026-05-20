from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE_XLSX = REPO_ROOT / "MR Data" / "PET Resin Total landed cost calculator.xlsx"
DEFAULT_OUTPUT_JSON = (
    REPO_ROOT
    / "market_research"
    / "data_exploration"
    / "outputs"
    / "market_research_workbook_profile.json"
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalized_path(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def profile_workbook(source_xlsx: Path) -> dict[str, Any]:
    formula_workbook = load_workbook(source_xlsx, data_only=False, read_only=False)
    value_workbook = load_workbook(source_xlsx, data_only=True, read_only=False)
    try:
        sheets: list[dict[str, Any]] = []
        for worksheet in formula_workbook.worksheets:
            value_sheet = value_workbook[worksheet.title]
            non_empty_cells = 0
            formula_cells = 0
            sample_non_empty_rows: list[list[str]] = []

            for row in worksheet.iter_rows():
                row_values: list[str] = []
                has_value = False
                for cell in row:
                    value = cell.value
                    cached_value = value_sheet[cell.coordinate].value
                    if value is not None or cached_value is not None:
                        non_empty_cells += 1
                        has_value = True
                    if isinstance(value, str) and value.startswith("="):
                        formula_cells += 1
                    row_values.append(clean_text(value if value is not None else cached_value))

                if has_value and len(sample_non_empty_rows) < 8:
                    sample_non_empty_rows.append(row_values)

            sheets.append(
                {
                    "sheet_name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "non_empty_cells": non_empty_cells,
                    "formula_cells": formula_cells,
                    "merged_ranges": len(worksheet.merged_cells.ranges),
                    "sample_non_empty_rows": sample_non_empty_rows,
                }
            )

        selected = formula_workbook["Total landed cost calculation"]
        return {
            "source_xlsx": normalized_path(source_xlsx),
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "selected_destination": clean_text(selected["B6"].value),
            "selected_month": clean_text(selected["C5"].value),
            "sheets": sheets,
        }
    finally:
        formula_workbook.close()
        value_workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Profile the market research PET resin TLC workbook.")
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    profile = profile_workbook(args.source_xlsx)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps(profile, indent=2, default=str), encoding="utf-8")
    print(json.dumps(profile, indent=2, default=str))


if __name__ == "__main__":
    main()
