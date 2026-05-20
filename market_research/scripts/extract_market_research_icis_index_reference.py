from __future__ import annotations

import argparse
import json
import re
import subprocess
import tempfile
import warnings
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE_FILE = (
    REPO_ROOT
    / "Index Fprecast"
    / "Market Research Index"
    / "ICIS Dashboard Price History 2026-05-13 202724.xls"
)
DEFAULT_SHEET = "ICIS Price History"
DEFAULT_RANGE = "B18:L34"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "Index Fprecast" / "Market Research Index"
DEFAULT_EXCEL_OUTPUT = DEFAULT_OUTPUT_DIR / "market_research_icis_resin_index_reference_table.xlsx"
DEFAULT_CSV_OUTPUT = DEFAULT_OUTPUT_DIR / "market_research_icis_resin_index_reference_table.csv"
DEFAULT_WIDE_CSV_OUTPUT = DEFAULT_OUTPUT_DIR / "market_research_icis_resin_index_reference_wide.csv"
FORECAST_YEAR = 2026
FORECAST_START_MONTH = 5
FORECAST_END_MONTH = 12
BACKTEST_MONTHS = [1, 2, 3, 4]
MONTHLY_GUARDRAIL_CAP = 0.08
CAP_BACKTEST_CANDIDATES = [0.08, 0.10, 0.12, 0.15, 0.20, None]

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def normalized_path(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def excel_serial_to_date(value: float) -> date:
    return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def parse_period(text: Any, value2: Any) -> date | None:
    if isinstance(value2, (int, float)) and value2 > 20000:
        return excel_serial_to_date(float(value2)).replace(day=1)

    cleaned = clean_text(text)
    if cleaned is None:
        return None

    for fmt in ("%b-%Y", "%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(cleaned, fmt).date().replace(day=1)
        except ValueError:
            continue
    return None


def parse_number(text: Any, value2: Any) -> float | None:
    if isinstance(value2, bool) or value2 is None:
        return None
    if isinstance(value2, (int, float)):
        return float(value2)

    cleaned = clean_text(text)
    if cleaned is None:
        return None
    try:
        return float(cleaned.replace(",", ""))
    except ValueError:
        return None


def normalize_formula(value: Any) -> str | None:
    return clean_text(value)


def infer_sourcing_country(series: Any) -> str | None:
    text = clean_text(series)
    if not text:
        return None

    country_patterns = [
        (r"\bArgentina\b", "Argentina"),
        (r"\bBrazil\b", "Brazil"),
        (r"\bChina\b", "China"),
        (r"\bIndia\b", "India"),
        (r"\bMexico\b", "Mexico"),
        (r"\bSouth Korea\b", "South Korea"),
        (r"\bTaiwan\b", "Taiwan"),
        (r"\bAsia SE\b", "Asia SE"),
        (r"\bAsia NE\b", "Asia NE"),
    ]
    for pattern, country in country_patterns:
        if re.search(pattern, text, flags=re.IGNORECASE):
            return country
    return None


def infer_resin_index_level(series: Any) -> str | None:
    text = clean_text(series)
    if not text:
        return None
    lowered = text.lower()
    if "(low)" in lowered:
        return "Low"
    if "(mid)" in lowered:
        return "Mid"
    if "(high)" in lowered:
        return "High"
    return None


def infer_resin_index_type(series: Any) -> str | None:
    text = clean_text(series)
    if not text:
        return None

    country = infer_sourcing_country(text)
    level = infer_resin_index_level(text)
    if not country and not level:
        return re.sub(r"\s*:\s*USD/tonne\s*$", "", text, flags=re.IGNORECASE)

    basis_parts: list[str] = []
    if re.search(r"\bDEL\b", text, flags=re.IGNORECASE):
        basis_parts.append("DEL")
    elif re.search(r"\bFOB\b", text, flags=re.IGNORECASE):
        basis_parts.append("FOB")

    if re.search(r"\bDomestic\b", text, flags=re.IGNORECASE):
        basis_parts.append("Domestic")
    if re.search(r"\bExport\b", text, flags=re.IGNORECASE):
        basis_parts.append("Export")
    if "Fibre Grade" in text:
        basis_parts.append("Fibre")

    label_parts = ["ICIS"]
    if country:
        label_parts.append(country)
    label_parts.extend(basis_parts)
    if level:
        label_parts.append(level)
    return " ".join(label_parts)


def read_xls_range_with_excel_com(
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    powershell = r"""
param(
  [Parameter(Mandatory=$true)][string]$Path,
  [Parameter(Mandatory=$true)][string]$SheetName,
  [Parameter(Mandatory=$true)][string]$RangeRef
)

$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$workbook = $null
$worksheet = $null

try {
  $workbook = $excel.Workbooks.Open($Path, 0, $true)
  $worksheet = $workbook.Worksheets.Item($SheetName)
  $range = $worksheet.Range($RangeRef)
  $rows = @()

  foreach ($cell in $range.Cells) {
    $rows += [pscustomobject]@{
      Row = [int]$cell.Row
      Column = [int]$cell.Column
      Address = [string]$cell.Address($false, $false)
      Text = [string]$cell.Text
      Value2 = $cell.Value2
      Formula = [string]$cell.Formula
    }
  }

  $rows | ConvertTo-Json -Depth 4 -Compress
  $workbook.Close($false)
}
finally {
  if ($workbook) {
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
  }
  if ($worksheet) {
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($worksheet) | Out-Null
  }
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}
"""
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as script:
        script.write(powershell)
        script_path = Path(script.name)

    try:
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-Path",
                str(source_file.resolve()),
                "-SheetName",
                sheet_name,
                "-RangeRef",
                range_ref,
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    finally:
        script_path.unlink(missing_ok=True)

    if not result.stdout.strip():
        return []

    payload = json.loads(result.stdout)
    if isinstance(payload, dict):
        return [payload]
    return payload


def build_raw_dataframe(
    cells: list[dict[str, Any]],
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> pd.DataFrame:
    rows = []
    for cell in cells:
        rows.append(
            {
                "source_file": normalized_path(source_file),
                "source_sheet": sheet_name,
                "source_range": range_ref,
                "source_row": cell["Row"],
                "source_column": get_column_letter(cell["Column"]),
                "source_cell": cell["Address"],
                "text": clean_text(cell.get("Text")),
                "value": cell.get("Value2"),
                "formula": normalize_formula(cell.get("Formula")),
            }
        )
    return pd.DataFrame(rows)


def build_long_dataframe(
    raw_df: pd.DataFrame,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> pd.DataFrame:
    header_row = 18
    header_df = raw_df[raw_df["source_row"] == header_row].copy()
    data_df = raw_df[raw_df["source_row"] > header_row].copy()

    header_by_column = {
        row["source_column"]: row["text"]
        for _, row in header_df.iterrows()
        if row["source_column"] != "B"
    }

    date_by_row: dict[int, dict[str, Any]] = {}
    for _, row in data_df[data_df["source_column"] == "B"].iterrows():
        period = parse_period(row["text"], row["value"])
        date_by_row[int(row["source_row"])] = {
            "period": period,
            "date_range": row["text"],
            "date_source_cell": row["source_cell"],
        }

    value_columns = sorted(column for column in header_by_column if column != "B")
    value_rows = []
    for _, row in data_df[data_df["source_column"].isin(value_columns)].iterrows():
        value = parse_number(row["text"], row["value"])
        if value is None:
            continue

        row_number = int(row["source_row"])
        column = row["source_column"]
        row_date = date_by_row.get(row_number, {})
        period = row_date.get("period")
        series = header_by_column.get(column)
        value_rows.append(
            {
                "source_file": normalized_path(source_file),
                "data_type": "Actual",
                "source_sheet": sheet_name,
                "source_range": range_ref,
                "source_row": row_number,
                "source_cell": row["source_cell"],
                "date_source_cell": row_date.get("date_source_cell"),
                "date_range": row_date.get("date_range"),
                "forecast_date": period.isoformat() if period else None,
                "time_period": f"{MONTH_LABELS[period.month]} {period.year}" if period else None,
                "time_period_year": period.year if period else None,
                "time_period_month": period.month if period else None,
                "forecast_series": series,
                "resin_index_type": infer_resin_index_type(series),
                "resin_index_level": infer_resin_index_level(series),
                "sourcing_country": infer_sourcing_country(series),
                "value_source_column": column,
                "value_source_header_cell": f"{column}{header_row}",
                "value": value,
                "raw_value_text": row["text"],
                "formula": row["formula"],
                "forecast_horizon_month": None,
                "forecast_growth_factor": None,
                "forecast_uncapped_value": None,
                "forecast_guardrail_cap_pct": None,
                "forecast_formula": None,
            }
        )

    return pd.DataFrame(value_rows)


def clamp(value: float, lower: float, upper: float) -> float:
    return min(max(value, lower), upper)


def period_label(period: date) -> str:
    return f"{MONTH_LABELS[period.month]} {period.year}"


def month_key_from_timestamp(value: Any) -> tuple[int, int]:
    timestamp = pd.Timestamp(value)
    return int(timestamp.year), int(timestamp.month)


def forecast_formula_text(month_name: str) -> str:
    cap_pct = MONTHLY_GUARDRAIL_CAP * 100
    return (
        f"Forecast {month_name} {FORECAST_YEAR} = MIN(MAX("
        f"Actual {month_name} 2025 * Growth Factor, Prior Value * (1 - {cap_pct:.0f}%)), "
        f"Prior Value * (1 + {cap_pct:.0f}%)); "
        "Growth Factor = AVERAGE(Actual Jan-Apr 2026 / Actual Jan-Apr 2025)."
    )


def build_series_lookup(group: pd.DataFrame) -> dict[tuple[int, int], float]:
    lookup: dict[tuple[int, int], float] = {}
    for _, row in group.iterrows():
        lookup[month_key_from_timestamp(row["forecast_date"])] = float(row["value"])
    return lookup


def build_forecast_dataframes(
    actual_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    working = actual_df.copy()
    working["forecast_date"] = pd.to_datetime(working["forecast_date"])

    group_columns = [
        "forecast_series",
        "resin_index_type",
        "resin_index_level",
        "sourcing_country",
        "value_source_column",
        "value_source_header_cell",
    ]

    forecast_rows: list[dict[str, Any]] = []
    backtest_rows: list[dict[str, Any]] = []
    accuracy_rows: list[dict[str, Any]] = []
    formula_rows: list[dict[str, Any]] = []
    cap_candidate_rows: list[dict[str, Any]] = []

    for group_key, group in working.groupby(group_columns, dropna=False):
        group_data = dict(zip(group_columns, group_key))
        series_values = build_series_lookup(group)

        required_keys = [(2025, month) for month in range(1, 13)] + [
            (FORECAST_YEAR, month) for month in BACKTEST_MONTHS
        ]
        if any(key not in series_values for key in required_keys):
            continue

        growth_ratios = [
            series_values[(FORECAST_YEAR, month)] / series_values[(2025, month)]
            for month in BACKTEST_MONTHS
            if series_values[(2025, month)] != 0
        ]
        if not growth_ratios:
            continue
        growth_factor = sum(growth_ratios) / len(growth_ratios)

        formula_rows.append(
            {
                **group_data,
                "forecast_method": "Seasonal same-month 2025 value scaled by 2026 YTD growth with fixed MoM guardrail",
                "growth_factor_formula": "AVERAGE(Actual Jan-Apr 2026 / Actual Jan-Apr 2025)",
                "growth_factor": growth_factor,
                "forecast_formula": forecast_formula_text("target month"),
                "monthly_guardrail_cap_pct": MONTHLY_GUARDRAIL_CAP,
                "forecast_months": f"{FORECAST_START_MONTH}-{FORECAST_END_MONTH} {FORECAST_YEAR}",
            }
        )

        for candidate_cap in CAP_BACKTEST_CANDIDATES:
            previous_value = series_values[(2025, 12)]
            apes: list[float] = []
            for month in BACKTEST_MONTHS:
                uncapped = series_values[(2025, month)] * growth_factor
                if candidate_cap is None:
                    backcast = uncapped
                else:
                    backcast = clamp(
                        uncapped,
                        previous_value * (1 - candidate_cap),
                        previous_value * (1 + candidate_cap),
                    )
                actual = series_values[(FORECAST_YEAR, month)]
                apes.append(abs(backcast - actual) / actual if actual else 0.0)
                previous_value = backcast
            cap_candidate_rows.append(
                {
                    **group_data,
                    "candidate_guardrail_cap_pct": "uncapped" if candidate_cap is None else candidate_cap,
                    "backtest_mape": sum(apes) / len(apes),
                    "backtest_max_ape": max(apes),
                }
            )

        previous_value = series_values[(2025, 12)]
        series_backtest_apes: list[float] = []
        for month in BACKTEST_MONTHS:
            month_name = MONTH_LABELS[month]
            uncapped = series_values[(2025, month)] * growth_factor
            lower_bound = previous_value * (1 - MONTHLY_GUARDRAIL_CAP)
            upper_bound = previous_value * (1 + MONTHLY_GUARDRAIL_CAP)
            backcast = clamp(uncapped, lower_bound, upper_bound)
            actual = series_values[(FORECAST_YEAR, month)]
            absolute_error = backcast - actual
            ape = abs(absolute_error) / actual if actual else 0.0
            series_backtest_apes.append(ape)
            backtest_rows.append(
                {
                    **group_data,
                    "backtest_month": month_name,
                    "backtest_date": date(FORECAST_YEAR, month, 1).isoformat(),
                    "actual_2025_same_month": series_values[(2025, month)],
                    "actual_2026": actual,
                    "previous_value_used": previous_value,
                    "growth_factor": growth_factor,
                    "uncapped_backcast": uncapped,
                    "lower_guardrail": lower_bound,
                    "upper_guardrail": upper_bound,
                    "backcast_value": backcast,
                    "absolute_error": absolute_error,
                    "absolute_percentage_error": ape,
                    "forecast_formula": forecast_formula_text(month_name),
                }
            )
            previous_value = backcast

        accuracy_rows.append(
            {
                **group_data,
                "backtest_months": "Jan-Apr 2026",
                "backtest_mape": sum(series_backtest_apes) / len(series_backtest_apes),
                "backtest_max_ape": max(series_backtest_apes),
                "monthly_guardrail_cap_pct": MONTHLY_GUARDRAIL_CAP,
                "growth_factor": growth_factor,
            }
        )

        previous_value = series_values[(FORECAST_YEAR, 4)]
        for horizon, month in enumerate(range(FORECAST_START_MONTH, FORECAST_END_MONTH + 1), start=1):
            forecast_period = date(FORECAST_YEAR, month, 1)
            month_name = MONTH_LABELS[month]
            uncapped = series_values[(2025, month)] * growth_factor
            lower_bound = previous_value * (1 - MONTHLY_GUARDRAIL_CAP)
            upper_bound = previous_value * (1 + MONTHLY_GUARDRAIL_CAP)
            forecast_value = clamp(uncapped, lower_bound, upper_bound)
            forecast_rows.append(
                {
                    "source_file": group["source_file"].iloc[0],
                    "data_type": "Forecast",
                    "source_sheet": group["source_sheet"].iloc[0],
                    "source_range": f"{group['source_range'].iloc[0]} forecast",
                    "source_row": None,
                    "source_cell": None,
                    "date_source_cell": None,
                    "date_range": forecast_period.strftime("%b-%Y"),
                    "forecast_date": forecast_period.isoformat(),
                    "time_period": period_label(forecast_period),
                    "time_period_year": forecast_period.year,
                    "time_period_month": forecast_period.month,
                    **group_data,
                    "value": forecast_value,
                    "raw_value_text": None,
                    "formula": forecast_formula_text(month_name),
                    "forecast_horizon_month": horizon,
                    "forecast_growth_factor": growth_factor,
                    "forecast_uncapped_value": uncapped,
                    "forecast_guardrail_cap_pct": MONTHLY_GUARDRAIL_CAP,
                    "forecast_formula": forecast_formula_text(month_name),
                }
            )
            previous_value = forecast_value

    return (
        pd.DataFrame(forecast_rows),
        pd.DataFrame(backtest_rows),
        pd.DataFrame(accuracy_rows),
        pd.DataFrame(formula_rows),
        pd.DataFrame(cap_candidate_rows),
    )


def build_wide_dataframe(long_df: pd.DataFrame) -> pd.DataFrame:
    if long_df.empty:
        return pd.DataFrame()

    key_columns = [
        "data_type",
        "time_period",
        "time_period_year",
        "time_period_month",
        "forecast_date",
        "date_range",
    ]
    wide = (
        long_df.pivot_table(
            index=key_columns,
            columns="resin_index_type",
            values="value",
            aggfunc="first",
        )
        .reset_index()
        .sort_values(["time_period_year", "time_period_month"])
    )
    wide.columns.name = None
    return wide


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            if worksheet.max_row == 0:
                continue
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for col_idx in range(1, worksheet.max_column + 1):
                header = worksheet.cell(1, col_idx).value
                width = len(str(header or ""))
                for row_idx in range(2, min(worksheet.max_row, 250) + 1):
                    value = worksheet.cell(row_idx, col_idx).value
                    if value is not None:
                        width = max(width, len(str(value)))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(width + 2, 12),
                    70,
                )
        workbook.save(path)
    finally:
        workbook.close()


def write_outputs(
    excel_output: Path,
    csv_output: Path,
    wide_csv_output: Path,
    long_df: pd.DataFrame,
    wide_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    forecast_df: pd.DataFrame,
    backtest_df: pd.DataFrame,
    accuracy_df: pd.DataFrame,
    formula_df: pd.DataFrame,
    cap_candidate_df: pd.DataFrame,
    metadata: dict[str, Any],
) -> None:
    excel_output.parent.mkdir(parents=True, exist_ok=True)
    csv_output.parent.mkdir(parents=True, exist_ok=True)
    wide_csv_output.parent.mkdir(parents=True, exist_ok=True)

    long_df.to_csv(csv_output, index=False, encoding="utf-8-sig")
    wide_df.to_csv(wide_csv_output, index=False, encoding="utf-8-sig")
    metadata_df = pd.DataFrame([{"field": key, "value": value} for key, value in metadata.items()])
    with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
        long_df.to_excel(writer, sheet_name="icis_resin_reference", index=False)
        wide_df.to_excel(writer, sheet_name="monthly_wide_reference", index=False)
        forecast_df.to_excel(writer, sheet_name="forecast_may_dec_2026", index=False)
        backtest_df.to_excel(writer, sheet_name="forecast_backtest", index=False)
        accuracy_df.to_excel(writer, sheet_name="forecast_accuracy", index=False)
        formula_df.to_excel(writer, sheet_name="forecast_formulae", index=False)
        cap_candidate_df.to_excel(writer, sheet_name="guardrail_backtest", index=False)
        raw_df.to_excel(writer, sheet_name="raw_B18_L34", index=False)
        metadata_df.to_excel(writer, sheet_name="run_metadata", index=False)
    style_workbook(excel_output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract Market Research ICIS resin index history from B18:L34."
    )
    parser.add_argument("--file", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--range", default=DEFAULT_RANGE)
    parser.add_argument("--excel-output", type=Path, default=DEFAULT_EXCEL_OUTPUT)
    parser.add_argument("--csv-output", type=Path, default=DEFAULT_CSV_OUTPUT)
    parser.add_argument("--wide-csv-output", type=Path, default=DEFAULT_WIDE_CSV_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    cells = read_xls_range_with_excel_com(args.file, args.sheet, args.range)
    raw_df = build_raw_dataframe(cells, args.file, args.sheet, args.range)
    actual_df = build_long_dataframe(raw_df, args.file, args.sheet, args.range)
    forecast_df, backtest_df, accuracy_df, formula_df, cap_candidate_df = build_forecast_dataframes(
        actual_df
    )
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.*",
            category=FutureWarning,
        )
        long_df = (
            pd.concat([actual_df, forecast_df], ignore_index=True)
            .sort_values(["resin_index_type", "time_period_year", "time_period_month", "data_type"])
            .reset_index(drop=True)
        )
    wide_df = build_wide_dataframe(long_df)
    overall_backtest_mape = (
        float(accuracy_df["backtest_mape"].mean()) if not accuracy_df.empty else None
    )
    overall_backtest_max_ape = (
        float(accuracy_df["backtest_max_ape"].max()) if not accuracy_df.empty else None
    )

    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "source_file": normalized_path(args.file),
        "source_sheet": args.sheet,
        "source_range": args.range,
        "header_row": 18,
        "data_rows": "19:34",
        "raw_cells": len(raw_df),
        "reference_rows": len(long_df),
        "actual_rows": len(actual_df),
        "forecast_rows": len(forecast_df),
        "wide_rows": len(wide_df),
        "forecast_method": (
            "Actual 2025 same-month value * average Jan-Apr 2026/2025 growth factor, "
            "capped to +/-8% from the prior actual/forecast value."
        ),
        "forecast_monthly_guardrail_cap_pct": MONTHLY_GUARDRAIL_CAP,
        "forecast_months": f"May-Dec {FORECAST_YEAR}",
        "backtest_months": "Jan-Apr 2026",
        "backtest_overall_mape": overall_backtest_mape,
        "backtest_overall_max_ape": overall_backtest_max_ape,
        "resin_index_types": ", ".join(
            sorted(str(value) for value in long_df["resin_index_type"].dropna().unique())
        ),
        "sourcing_countries": ", ".join(
            sorted(str(value) for value in long_df["sourcing_country"].dropna().unique())
        ),
        "output_excel": normalized_path(args.excel_output),
        "output_csv": normalized_path(args.csv_output),
        "output_wide_csv": normalized_path(args.wide_csv_output),
    }
    write_outputs(
        args.excel_output,
        args.csv_output,
        args.wide_csv_output,
        long_df,
        wide_df,
        raw_df,
        forecast_df,
        backtest_df,
        accuracy_df,
        formula_df,
        cap_candidate_df,
        metadata,
    )

    print(json.dumps(metadata, indent=2, default=str))


if __name__ == "__main__":
    main()
