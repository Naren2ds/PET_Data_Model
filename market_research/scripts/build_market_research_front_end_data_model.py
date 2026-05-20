from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE_XLSX = REPO_ROOT / "MR Data" / "PET Resin Total landed cost calculator.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "market_research" / "standardization" / "outputs"
DEFAULT_OUTPUT_XLSX = DEFAULT_OUTPUT_DIR / "Data_Standardized_MR_Front_End_Data_Model.xlsx"
DEFAULT_OUTPUT_CSV = DEFAULT_OUTPUT_DIR / "Data_Standardized_MR_Front_End_Data_Model.csv"
DEFAULT_SUMMARY_JSON = DEFAULT_OUTPUT_DIR / "market_research_data_model_run_summary.json"
DEFAULT_MAPPING_XLSX = REPO_ROOT / "mapping files" / "Mapping_Columns.xlsx"
DEFAULT_MAPPING_SHEET = "MR PET Resin TLC"
DEFAULT_ICIS_INDEX_CSV = (
    REPO_ROOT
    / "Index Fprecast"
    / "Market Research Index"
    / "market_research_icis_resin_index_reference_table.csv"
)

ACTUAL_DATA_TYPE = "Actual"
FORECAST_DATA_TYPE = "Forecast"

CANONICAL_COLUMNS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]

MONTH_NUMBERS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

FORECAST_INDEX_MAPPINGS = {
    "Argentina": {
        "index_country": "Argentina",
        "resin_index_type": "ICIS Argentina DEL Domestic Mid",
        "basis": "Direct source-country ICIS series",
    },
    "Brazil": {
        "index_country": "Brazil",
        "resin_index_type": "ICIS Brazil DEL Domestic Mid",
        "basis": "Direct source-country ICIS series",
    },
    "China": {
        "index_country": "China",
        "resin_index_type": "ICIS China FOB Mid",
        "basis": "Direct source-country ICIS series",
    },
    "India": {
        "index_country": "India",
        "resin_index_type": "ICIS India FOB Mid",
        "basis": "Direct source-country ICIS series",
    },
    "Mexico": {
        "index_country": "Mexico",
        "resin_index_type": "ICIS Mexico FOB Export Mid",
        "basis": "Mexico export FOB ICIS series used for landed-cost FOB resin",
    },
    "South Korea": {
        "index_country": "South Korea",
        "resin_index_type": "ICIS South Korea FOB Mid",
        "basis": "Direct source-country ICIS series",
    },
    "Taiwan": {
        "index_country": "Taiwan",
        "resin_index_type": "ICIS Taiwan FOB Mid",
        "basis": "Direct source-country ICIS series",
    },
    "Indonesia": {
        "index_country": "Asia SE",
        "resin_index_type": "ICIS Asia SE FOB Mid",
        "basis": "Southeast Asia regional ICIS proxy",
    },
    "Thailand": {
        "index_country": "Asia SE",
        "resin_index_type": "ICIS Asia SE FOB Mid",
        "basis": "Southeast Asia regional ICIS proxy",
    },
    "Vietnam": {
        "index_country": "Asia SE",
        "resin_index_type": "ICIS Asia SE FOB Mid",
        "basis": "Southeast Asia regional ICIS proxy",
    },
    "USA": {
        "index_country": "Mexico",
        "resin_index_type": "ICIS Mexico DEL Domestic Mid",
        "basis": "North America ICIS proxy because a USA series is not present in the MR ICIS file",
    },
}

TLC_FORMULA = (
    "Total landed cost (PET resin) = PET resin cost (FOB) + Freight cost + "
    "Insurance + Import duty + Anti-dumping duty + IPI + PIS + COFINS + "
    "Statistical fee + Tasa consular + Customs service fee + Customs insurance + "
    "IGV/IPM + Percepcion IGV + FODINFA + Destination port transportation. "
    "Additional VAT, income tax perception, IBB, IRAE, and VAT are retained as "
    "non-required reference rows only."
)


@dataclass(frozen=True)
class Component:
    key: str
    raw_label: str
    mapping_column: str
    required: str


DEFAULT_COMPONENTS = [
    Component("resin_fob", "PET resin cost (FOB)", "Resin Index vPET", "Yes"),
    Component("freight", "Freight cost", "Freight", "Yes"),
    Component("insurance", "Insurance", "Insurance", "Yes"),
    Component("import_duty", "Import duty", "Tax", "Yes"),
    Component("anti_dumping", "Anti-dumping duty", "Tax", "Yes"),
    Component("ipi", "IPI", "Tax", "Yes"),
    Component("pis", "PIS", "Tax", "Yes"),
    Component("cofins", "Confins", "Tax", "Yes"),
    Component("statistical_fee", "Statistical fee", "Tax", "Yes"),
    Component("additional_vat", "Additional VAT", "Tax", "No"),
    Component("income_tax_perception", "Income tax perception", "Tax", "No"),
    Component("ibb", "IBB", "Tax", "No"),
    Component("tasa_consular", "Tasa consular", "Tax", "Yes"),
    Component("customs_service_fee", "Customs service fee", "Customs clearance", "Yes"),
    Component("irae", "IRAE", "Tax", "No"),
    Component("customs_insurance", "Customs insurance", "Insurance", "Yes"),
    Component("igv_ipm", "Impuesto General a las Ventas (IGV & IPM)", "Tax", "Yes"),
    Component("percepcion_igv", "Percepción IGV", "Tax", "Yes"),
    Component("fodinfa", "FODINFA", "Tax", "Yes"),
    Component("vat_import", "Taxes (VAT/Import)", "Tax", "No"),
    Component(
        "local_transport",
        "Destination port to supplier location transportation",
        "Freight",
        "Yes",
    ),
    Component("tlc", "Total landed cost (PET resin)", "Total Landing Cost", "Yes"),
]

INCLUDED_COMPONENT_KEYS = [
    "resin_fob",
    "freight",
    "insurance",
    "import_duty",
    "anti_dumping",
    "ipi",
    "pis",
    "cofins",
    "statistical_fee",
    "tasa_consular",
    "customs_service_fee",
    "customs_insurance",
    "igv_ipm",
    "percepcion_igv",
    "fodinfa",
    "local_transport",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalized_path(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def parse_number(value: Any, *, blank_as_zero: bool = False) -> float | None:
    if value is None:
        return 0.0 if blank_as_zero else None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return 0.0 if blank_as_zero else None
    if text.casefold() in {"n/a", "na", "#n/a", "none", "nan"}:
        return None

    number_text = text.replace(",", "")
    if number_text.endswith("%"):
        try:
            return float(number_text[:-1]) / 100
        except ValueError:
            return None

    try:
        return float(number_text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    if number is None:
        return None
    return int(number)


def period_label(month_number: int, year: int) -> str:
    return f"{MONTH_LABELS.get(month_number, str(month_number))} {year}"


def safe_add(*values: float | None) -> float | None:
    if any(value is None for value in values):
        return None
    return sum(value for value in values if value is not None)


def safe_multiply(left: float | None, right: float | None) -> float | None:
    if left is None or right is None:
        return None
    return left * right


def infer_model_period(workbook: Any) -> tuple[str, int, int]:
    worksheet = workbook["Total landed cost calculation"]
    month_name = clean_text(worksheet["C5"].value) or "February"
    month_number = MONTH_NUMBERS.get(month_name.casefold(), 2)

    lookup = workbook["Lookups & references"]
    years: list[int] = []
    for row in range(3, lookup.max_row + 1):
        for col in (7, 13):
            value = lookup.cell(row, col).value
            if isinstance(value, datetime):
                years.append(value.year)

    year = max(years) if years else datetime.now().year
    return month_name, year, month_number


def extract_fob_prices(worksheet: Any) -> tuple[dict[str, float], dict[str, dict[str, Any]], list[str]]:
    prices: dict[str, float] = {}
    metadata: dict[str, dict[str, Any]] = {}
    source_order: list[str] = []
    for row in range(3, worksheet.max_row + 1):
        country = clean_text(worksheet.cell(row, 4).value)
        price = parse_number(worksheet.cell(row, 5).value)
        if not country or price is None:
            continue

        prices[country] = price
        source_order.append(country)
        metadata[country] = {
            "resin_index_type": clean_text(worksheet.cell(row, 6).value),
            "source_information_last_updated": worksheet.cell(row, 7).value,
            "comments": clean_text(worksheet.cell(row, 8).value),
        }
    return prices, metadata, source_order


def extract_freight_rates(worksheet: Any) -> dict[tuple[str, str], dict[str, Any]]:
    rates: dict[tuple[str, str], dict[str, Any]] = {}
    for row in range(3, worksheet.max_row + 1):
        route = clean_text(worksheet.cell(row, 10).value)
        if "|" not in route:
            continue

        source, destination = [clean_text(part) for part in route.split("|", 1)]
        rates[(source, destination)] = {
            "freight_rate": parse_number(worksheet.cell(row, 11).value),
            "source": clean_text(worksheet.cell(row, 12).value),
            "source_information_last_updated": worksheet.cell(row, 13).value,
            "source_port": clean_text(worksheet.cell(row, 14).value),
            "destination_port": clean_text(worksheet.cell(row, 15).value),
            "total_container_rate": parse_number(worksheet.cell(row, 16).value),
        }
    return rates


def extract_local_transport_rates(worksheet: Any) -> dict[str, float]:
    rates: dict[str, float] = {}
    for row in range(3, worksheet.max_row + 1):
        country = clean_text(worksheet.cell(row, 1).value)
        if not country or country.startswith("<"):
            continue
        rates[country] = parse_number(worksheet.cell(row, 2).value, blank_as_zero=True) or 0.0
    return rates


def extract_tax_references(worksheet: Any) -> tuple[dict[str, int], dict[str, int]]:
    source_columns: dict[str, int] = {}
    for col in range(5, 16):
        source = clean_text(worksheet.cell(3, col).value)
        if source:
            source_columns[source] = col

    destination_start_rows: dict[str, int] = {}
    for row in range(4, worksheet.max_row + 1):
        destination = clean_text(worksheet.cell(row, 3).value)
        if destination:
            destination_start_rows[destination] = row

    return source_columns, destination_start_rows


def tax_rate(
    worksheet: Any,
    destination_start_rows: dict[str, int],
    source_columns: dict[str, int],
    destination: str,
    source: str,
    offset: int,
    *,
    invalid_as_zero: bool = False,
) -> float | None:
    start_row = destination_start_rows[destination]
    col = source_columns[source]
    value = worksheet.cell(start_row + offset, col).value
    number = parse_number(value, blank_as_zero=True)
    if number is None and invalid_as_zero:
        return 0.0
    return number


def calculate_components(
    tax_sheet: Any,
    destination_start_rows: dict[str, int],
    source_columns: dict[str, int],
    fob_prices: dict[str, float],
    freight_rates: dict[tuple[str, str], dict[str, Any]],
    local_transport_rates: dict[str, float],
    destination: str,
    source: str,
    resin_fob_override: float | None = None,
) -> dict[str, float | None]:
    resin_fob = resin_fob_override if resin_fob_override is not None else fob_prices.get(source)
    freight = freight_rates.get((source, destination), {}).get("freight_rate")
    insurance = safe_multiply(safe_add(resin_fob, freight), 0.004)

    def rate(offset: int, *, invalid_as_zero: bool = False) -> float | None:
        return tax_rate(
            tax_sheet,
            destination_start_rows,
            source_columns,
            destination,
            source,
            offset,
            invalid_as_zero=invalid_as_zero,
        )

    import_duty = safe_multiply(rate(0), safe_add(resin_fob, freight, insurance))
    anti_dumping = rate(1, invalid_as_zero=True)
    ipi = safe_multiply(rate(2), safe_add(resin_fob, import_duty, freight, insurance))
    pis = safe_multiply(rate(3), safe_add(resin_fob, freight, insurance))
    cofins = safe_multiply(rate(4), safe_add(resin_fob, freight, insurance))
    statistical_fee = safe_multiply(rate(5), safe_add(resin_fob, freight, insurance))
    additional_vat = safe_multiply(
        rate(6),
        safe_add(resin_fob, freight, import_duty, statistical_fee, insurance),
    )
    income_tax_perception = safe_multiply(
        rate(7),
        safe_add(resin_fob, freight, import_duty, statistical_fee, insurance),
    )
    ibb = safe_multiply(
        rate(8),
        safe_add(resin_fob, freight, import_duty, statistical_fee, insurance),
    )
    tasa_consular = safe_multiply(rate(9), safe_add(resin_fob, freight, insurance))
    customs_service_base = (
        1.0
        if destination in {"El Salvador", "Honduras"}
        else safe_add(resin_fob, freight, insurance)
    )
    customs_service_fee = safe_multiply(rate(10), customs_service_base)
    irae = safe_multiply(rate(11), safe_add(resin_fob, freight, import_duty, insurance))
    customs_insurance = safe_multiply(rate(12), safe_add(resin_fob, freight, insurance))
    igv_ipm = safe_multiply(
        rate(13),
        safe_add(resin_fob, freight, import_duty, customs_insurance, insurance),
    )
    fodinfa = safe_multiply(rate(15), safe_add(resin_fob, freight, insurance))

    vat_rate = rate(16)
    vat_base = safe_add(
        freight,
        import_duty,
        resin_fob,
        ipi,
        pis,
        cofins,
        statistical_fee,
        customs_service_fee,
        customs_insurance,
        fodinfa,
        insurance,
    )
    if vat_rate is None or vat_base is None:
        vat_import = None
    elif destination == "Brazil":
        vat_import = vat_rate * (vat_base / (1 - vat_rate)) if vat_rate != 1 else None
    else:
        vat_import = vat_rate * vat_base

    percepcion_igv = safe_multiply(
        rate(14),
        safe_add(resin_fob, freight, customs_insurance, import_duty, vat_import, insurance),
    )
    local_transport = local_transport_rates.get(destination, 0.0)

    components: dict[str, float | None] = {
        "resin_fob": resin_fob,
        "freight": freight,
        "insurance": insurance,
        "import_duty": import_duty,
        "anti_dumping": anti_dumping,
        "ipi": ipi,
        "pis": pis,
        "cofins": cofins,
        "statistical_fee": statistical_fee,
        "additional_vat": additional_vat,
        "income_tax_perception": income_tax_perception,
        "ibb": ibb,
        "tasa_consular": tasa_consular,
        "customs_service_fee": customs_service_fee,
        "irae": irae,
        "customs_insurance": customs_insurance,
        "igv_ipm": igv_ipm,
        "percepcion_igv": percepcion_igv,
        "fodinfa": fodinfa,
        "vat_import": vat_import,
        "local_transport": local_transport,
    }
    components["tlc"] = safe_add(*(components[key] for key in INCLUDED_COMPONENT_KEYS))
    return components


def find_header(headers: list[str], candidates: set[str]) -> int | None:
    normalized_candidates = {candidate.casefold() for candidate in candidates}
    for index, header in enumerate(headers):
        if clean_text(header).casefold() in normalized_candidates:
            return index
    return None


def load_component_mappings(
    mapping_xlsx: Path,
    sheet_name: str,
) -> dict[str, tuple[str, str]]:
    if not mapping_xlsx.exists():
        return {}

    workbook = load_workbook(mapping_xlsx, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return {}

        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return {}

        headers = [clean_text(value) for value in rows[0]]
        metric_col = find_header(
            headers,
            {
                "Raw Cost Breakdown",
                "metric_name",
                "metric_label_english",
                "metric_local",
                "metric_label_original",
            },
        )
        mapping_col = find_header(headers, {"Mapping Columns", "Mapping Columns "})
        required_col = find_header(
            headers,
            {"Column Required for Calculation", "Column Required for Calculation "},
        )
        if metric_col is None or mapping_col is None:
            return {}

        mappings: dict[str, tuple[str, str]] = {}
        for row in rows[1:]:
            metric = clean_text(row[metric_col] if metric_col < len(row) else None)
            mapping = clean_text(row[mapping_col] if mapping_col < len(row) else None)
            required = (
                clean_text(row[required_col] if required_col < len(row) else None)
                if required_col is not None
                else ""
            )
            if metric and mapping:
                mappings[metric] = (mapping, required or "Yes")
        return mappings
    finally:
        workbook.close()


def apply_component_mappings(
    default_components: list[Component],
    mappings: dict[str, tuple[str, str]],
) -> list[Component]:
    mapped_components: list[Component] = []
    for component in default_components:
        mapping_column, required = mappings.get(
            component.raw_label,
            (component.mapping_column, component.required),
        )
        mapped_components.append(
            Component(
                component.key,
                component.raw_label,
                mapping_column,
                required or component.required,
            )
        )
    return mapped_components


def missing_reason(
    calculated_components: dict[str, float | None],
    mapped_components: list[Component],
) -> str | None:
    if calculated_components.get("tlc") is not None:
        return None

    missing = [
        component.raw_label
        for component in mapped_components
        if component.required == "Yes"
        and component.key != "tlc"
        and calculated_components.get(component.key) is None
    ]
    return "Unable to calculate required component(s): " + "; ".join(missing)


def load_icis_index_rows(index_csv: Path) -> dict[tuple[str, str, str, int, int], dict[str, Any]]:
    if not index_csv.exists():
        return {}

    rows: dict[tuple[str, str, str, int, int], dict[str, Any]] = {}
    with index_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            country = clean_text(row.get("sourcing_country"))
            resin_index_type = clean_text(row.get("resin_index_type"))
            data_type = clean_text(row.get("data_type"))
            year = parse_int(row.get("time_period_year"))
            month = parse_int(row.get("time_period_month"))
            value = parse_number(row.get("value"))
            if not country or not resin_index_type or not data_type or year is None or month is None:
                continue
            if value is None:
                continue

            rows[(country, resin_index_type, data_type, year, month)] = {
                "source_file": clean_text(row.get("source_file")),
                "data_type": data_type,
                "time_period": clean_text(row.get("time_period")) or period_label(month, year),
                "time_period_year": year,
                "time_period_month": month,
                "forecast_date": clean_text(row.get("forecast_date")),
                "sourcing_country": country,
                "resin_index_type": resin_index_type,
                "value": value,
                "formula": clean_text(row.get("formula")),
                "forecast_formula": clean_text(row.get("forecast_formula")),
            }
    return rows


def append_model_component_rows(
    model_rows: list[dict[str, Any]],
    mapped_components: list[Component],
    components: dict[str, float | None],
    *,
    data_type: str,
    source_file: str,
    source: str,
    destination: str,
    month_label: str,
    year: int,
    month_number: int,
    resin_index_type: str,
    forecast_resin_index_type: str | None,
) -> None:
    for component in mapped_components:
        model_rows.append(
            {
                "Data Type": data_type,
                "Source File ": source_file,
                "Supplier Name": source,
                "Destination Country": destination,
                "Time_Period ": month_label,
                "Time Period Year": year,
                "Time Period Month": month_number,
                "Location": source,
                "Raw Cost Breakdown": component.raw_label,
                "Resin Index Type": resin_index_type,
                "Forecast Resin Index Type": forecast_resin_index_type,
                "Mapping Columns": component.mapping_column,
                "Column Required for Calculation": component.required,
                "Value ": components.get(component.key),
                "TLC Formula": TLC_FORMULA,
            }
        )


def build_tlc_validation_row(
    *,
    data_type: str,
    source: str,
    destination: str,
    month_label: str,
    year: int,
    month_number: int,
    components: dict[str, float | None],
    forecast_resin_index_type: str | None = None,
) -> dict[str, Any]:
    included_values = [components.get(key) for key in INCLUDED_COMPONENT_KEYS]
    back_calculated_tlc = safe_add(*included_values)
    output_tlc = components.get("tlc")
    difference = None
    status = "validated"
    if back_calculated_tlc is None or output_tlc is None:
        status = "missing_component"
    else:
        difference = output_tlc - back_calculated_tlc
        if abs(difference) > 0.000001:
            status = "variance"

    return {
        "data_type": data_type,
        "source_country": source,
        "destination_country": destination,
        "time_period": month_label,
        "time_period_year": year,
        "time_period_month": month_number,
        "resin_fob": components.get("resin_fob"),
        "freight": components.get("freight"),
        "insurance": components.get("insurance"),
        "output_tlc_value": output_tlc,
        "back_calculated_tlc": back_calculated_tlc,
        "difference": difference,
        "validation_status": status,
        "forecast_resin_index_type": forecast_resin_index_type,
        "included_component_keys": "; ".join(INCLUDED_COMPONENT_KEYS),
    }


def build_rows(
    source_xlsx: Path,
    mapping_xlsx: Path,
    mapping_sheet: str,
    icis_index_csv: Path,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    dict[str, Any],
    list[Component],
]:
    mapped_components = apply_component_mappings(
        DEFAULT_COMPONENTS,
        load_component_mappings(mapping_xlsx, mapping_sheet),
    )
    icis_index_rows = load_icis_index_rows(icis_index_csv)

    workbook = load_workbook(source_xlsx, read_only=True, data_only=True)
    try:
        lookup_sheet = workbook["Lookups & references"]
        tax_sheet = workbook["Tax references"]

        period_month, period_year, period_month_number = infer_model_period(workbook)
        fob_prices, source_metadata, source_order = extract_fob_prices(lookup_sheet)
        freight_rates = extract_freight_rates(lookup_sheet)
        local_transport_rates = extract_local_transport_rates(lookup_sheet)
        source_columns, destination_start_rows = extract_tax_references(tax_sheet)

        destinations = list(destination_start_rows)
        sources = [source for source in source_order if source in source_columns]
        source_file = normalized_path(source_xlsx)
        actual_period_label = f"{period_month} {period_year}"
        forecast_source_file = f"{source_file}; {normalized_path(icis_index_csv)}"

        model_rows: list[dict[str, Any]] = []
        audit_rows: list[dict[str, Any]] = []
        forecast_audit_rows: list[dict[str, Any]] = []
        forecast_index_mapping_rows: list[dict[str, Any]] = []
        validation_rows: list[dict[str, Any]] = []
        calculated_by_destination: dict[str, list[tuple[str, float]]] = {
            destination: [] for destination in destinations
        }

        calculated_components: dict[tuple[str, str], dict[str, float | None]] = {}
        for destination in destinations:
            for source in sources:
                components = calculate_components(
                    tax_sheet,
                    destination_start_rows,
                    source_columns,
                    fob_prices,
                    freight_rates,
                    local_transport_rates,
                    destination,
                    source,
                )
                calculated_components[(destination, source)] = components
                if components["tlc"] is not None:
                    calculated_by_destination[destination].append((source, components["tlc"]))

        ranks: dict[tuple[str, str], int] = {}
        for destination, entries in calculated_by_destination.items():
            for rank, (source, _tlc) in enumerate(sorted(entries, key=lambda item: item[1]), start=1):
                ranks[(destination, source)] = rank

        for destination in destinations:
            for source in sources:
                components = calculated_components[(destination, source)]
                reason = missing_reason(components, mapped_components)
                freight_meta = freight_rates.get((source, destination), {})
                audit_rows.append(
                    {
                        "data_type": ACTUAL_DATA_TYPE,
                        "time_period": actual_period_label,
                        "time_period_year": period_year,
                        "time_period_month": period_month_number,
                        "destination_country": destination,
                        "source_country": source,
                        "calculation_status": "calculated" if reason is None else "skipped",
                        "rank": ranks.get((destination, source)),
                        "tlc_value": components.get("tlc"),
                        "missing_reason": reason,
                        "fob_price": components.get("resin_fob"),
                        "freight_rate": components.get("freight"),
                        "source_port": freight_meta.get("source_port"),
                        "destination_port": freight_meta.get("destination_port"),
                    }
                )
                if reason is not None:
                    continue

                resin_index_type = source_metadata.get(source, {}).get("resin_index_type", "")
                append_model_component_rows(
                    model_rows,
                    mapped_components,
                    components,
                    data_type=ACTUAL_DATA_TYPE,
                    source_file=source_file,
                    source=source,
                    destination=destination,
                    month_label=actual_period_label,
                    year=period_year,
                    month_number=period_month_number,
                    resin_index_type=resin_index_type,
                    forecast_resin_index_type=None,
                )
                validation_rows.append(
                    build_tlc_validation_row(
                        data_type=ACTUAL_DATA_TYPE,
                        source=source,
                        destination=destination,
                        month_label=actual_period_label,
                        year=period_year,
                        month_number=period_month_number,
                        components=components,
                    )
                )

        forecast_points_by_source: dict[str, dict[tuple[int, int], dict[str, Any]]] = {}
        for source in sources:
            mapping = FORECAST_INDEX_MAPPINGS.get(source)
            mapping_status = "mapped"
            missing_reason_text = None
            forecast_points: dict[tuple[int, int], dict[str, Any]] = {}
            actual_index_point = None

            if mapping is None:
                mapping_status = "skipped"
                missing_reason_text = "No forecast index mapping configured for source country."
            else:
                index_country = mapping["index_country"]
                forecast_resin_index_type = mapping["resin_index_type"]
                actual_index_point = icis_index_rows.get(
                    (
                        index_country,
                        forecast_resin_index_type,
                        ACTUAL_DATA_TYPE,
                        period_year,
                        period_month_number,
                    )
                )
                forecast_points = {
                    (row["time_period_year"], row["time_period_month"]): row
                    for key, row in icis_index_rows.items()
                    if key[0] == index_country
                    and key[1] == forecast_resin_index_type
                    and key[2] in {ACTUAL_DATA_TYPE, FORECAST_DATA_TYPE}
                    and row["time_period_year"] == period_year
                    and row["time_period_month"] > period_month_number
                    and row["time_period_month"] <= 12
                }
                if not forecast_points:
                    mapping_status = "skipped"
                    missing_reason_text = (
                        "No March-December actual/forecast rows found in ICIS index table."
                    )

            forecast_points_by_source[source] = forecast_points
            forecast_index_mapping_rows.append(
                {
                    "source_country": source,
                    "source_actual_resin_index_type": source_metadata.get(source, {}).get(
                        "resin_index_type", ""
                    ),
                    "source_actual_fob": fob_prices.get(source),
                    "forecast_index_country": mapping.get("index_country") if mapping else None,
                    "forecast_resin_index_type": mapping.get("resin_index_type") if mapping else None,
                    "mapping_basis": mapping.get("basis") if mapping else None,
                    "mapping_status": mapping_status,
                    "missing_reason": missing_reason_text,
                    "actual_index_value_for_model_month": (
                        actual_index_point.get("value") if actual_index_point else None
                    ),
                    "actual_index_model_month": (
                        actual_index_point.get("time_period") if actual_index_point else None
                    ),
                    "forecast_month_count": len(forecast_points),
                    "forecast_months": "; ".join(
                        period_label(month, year) for year, month in sorted(forecast_points)
                    ),
                    "index_data_types_used": "; ".join(
                        sorted({row["data_type"] for row in forecast_points.values()})
                    ),
                }
            )

        forecast_components: dict[tuple[int, int, str, str], dict[str, float | None]] = {}
        forecast_by_destination_period: dict[tuple[int, int, str], list[tuple[str, float]]] = {}
        for source, forecast_points in forecast_points_by_source.items():
            for (year, month_number), index_point in forecast_points.items():
                for destination in destinations:
                    components = calculate_components(
                        tax_sheet,
                        destination_start_rows,
                        source_columns,
                        fob_prices,
                        freight_rates,
                        local_transport_rates,
                        destination,
                        source,
                        resin_fob_override=index_point["value"],
                    )
                    forecast_components[(year, month_number, destination, source)] = components
                    if components["tlc"] is not None:
                        forecast_by_destination_period.setdefault(
                            (year, month_number, destination), []
                        ).append((source, components["tlc"]))

        forecast_ranks: dict[tuple[int, int, str, str], int] = {}
        for (year, month_number, destination), entries in forecast_by_destination_period.items():
            for rank, (source, _tlc) in enumerate(sorted(entries, key=lambda item: item[1]), start=1):
                forecast_ranks[(year, month_number, destination, source)] = rank

        for (year, month_number, destination, source), components in sorted(forecast_components.items()):
            point = forecast_points_by_source[source][(year, month_number)]
            reason = missing_reason(components, mapped_components)
            freight_meta = freight_rates.get((source, destination), {})
            month_label = point.get("time_period") or period_label(month_number, year)
            forecast_audit_rows.append(
                {
                    "data_type": FORECAST_DATA_TYPE,
                    "time_period": month_label,
                    "time_period_year": year,
                    "time_period_month": month_number,
                    "destination_country": destination,
                    "source_country": source,
                    "calculation_status": "calculated" if reason is None else "skipped",
                    "rank": forecast_ranks.get((year, month_number, destination, source)),
                    "tlc_value": components.get("tlc"),
                    "missing_reason": reason,
                    "fob_price": components.get("resin_fob"),
                    "freight_rate": components.get("freight"),
                    "source_port": freight_meta.get("source_port"),
                    "destination_port": freight_meta.get("destination_port"),
                    "forecast_index_country": point.get("sourcing_country"),
                    "forecast_index_data_type": point.get("data_type"),
                    "forecast_resin_index_type": point.get("resin_index_type"),
                    "forecast_index_value": point.get("value"),
                    "forecast_index_formula": point.get("forecast_formula"),
                }
            )
            if reason is not None:
                continue

            resin_index_type = source_metadata.get(source, {}).get("resin_index_type", "")
            append_model_component_rows(
                model_rows,
                mapped_components,
                components,
                data_type=FORECAST_DATA_TYPE,
                source_file=forecast_source_file,
                source=source,
                destination=destination,
                month_label=month_label,
                year=year,
                month_number=month_number,
                resin_index_type=resin_index_type,
                forecast_resin_index_type=point.get("resin_index_type"),
            )
            validation_rows.append(
                build_tlc_validation_row(
                    data_type=FORECAST_DATA_TYPE,
                    source=source,
                    destination=destination,
                    month_label=month_label,
                    year=year,
                    month_number=month_number,
                    components=components,
                    forecast_resin_index_type=point.get("resin_index_type"),
                )
            )

        forecast_periods = sorted(
            {
                (row["time_period_year"], row["time_period_month"], row["time_period"])
                for row in forecast_audit_rows
                if row["calculation_status"] == "calculated"
            }
        )
        validation_status_counts: dict[str, int] = {}
        for row in validation_rows:
            status = row["validation_status"]
            validation_status_counts[status] = validation_status_counts.get(status, 0) + 1

        summary = {
            "source_xlsx": source_file,
            "icis_index_csv": normalized_path(icis_index_csv),
            "mapping_xlsx": normalized_path(mapping_xlsx),
            "mapping_sheet": mapping_sheet,
            "data_types": f"{ACTUAL_DATA_TYPE}; {FORECAST_DATA_TYPE}",
            "actual_period": actual_period_label,
            "forecast_periods": "; ".join(label for _year, _month, label in forecast_periods),
            "destination_count": len(destinations),
            "source_country_count": len(sources),
            "source_destination_combinations": len(destinations) * len(sources),
            "calculated_combinations": sum(
                1 for row in audit_rows if row["calculation_status"] == "calculated"
            ),
            "skipped_combinations": sum(
                1 for row in audit_rows if row["calculation_status"] == "skipped"
            ),
            "forecast_calculated_combinations": sum(
                1 for row in forecast_audit_rows if row["calculation_status"] == "calculated"
            ),
            "forecast_skipped_combinations": sum(
                1 for row in forecast_audit_rows if row["calculation_status"] == "skipped"
            ),
            "actual_model_rows": sum(1 for row in model_rows if row["Data Type"] == ACTUAL_DATA_TYPE),
            "forecast_model_rows": sum(
                1 for row in model_rows if row["Data Type"] == FORECAST_DATA_TYPE
            ),
            "model_rows": len(model_rows),
            "validation_rows": len(validation_rows),
            "validation_statuses": "; ".join(
                f"{status}={count}"
                for status, count in sorted(validation_status_counts.items())
            ),
            "columns": "; ".join(CANONICAL_COLUMNS),
        }
        return (
            model_rows,
            audit_rows,
            forecast_audit_rows,
            forecast_index_mapping_rows,
            validation_rows,
            summary,
            mapped_components,
        )
    finally:
        workbook.close()


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 300) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name[:31])
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    style_sheet(worksheet)


def write_outputs(
    model_rows: list[dict[str, Any]],
    audit_rows: list[dict[str, Any]],
    forecast_audit_rows: list[dict[str, Any]],
    forecast_index_mapping_rows: list[dict[str, Any]],
    validation_rows: list[dict[str, Any]],
    summary: dict[str, Any],
    mapped_components: list[Component],
    output_xlsx: Path,
    output_csv: Path,
    summary_json: Path,
) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CANONICAL_COLUMNS)
        writer.writeheader()
        writer.writerows([{column: row.get(column) for column in CANONICAL_COLUMNS} for row in model_rows])

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(workbook, "front_end_data_model", CANONICAL_COLUMNS, model_rows)

    audit_headers = [
        "data_type",
        "time_period",
        "time_period_year",
        "time_period_month",
        "destination_country",
        "source_country",
        "calculation_status",
        "rank",
        "tlc_value",
        "missing_reason",
        "fob_price",
        "freight_rate",
        "source_port",
        "destination_port",
    ]
    write_sheet(workbook, "calculation_audit", audit_headers, audit_rows)

    forecast_audit_headers = [
        *audit_headers,
        "forecast_index_country",
        "forecast_index_data_type",
        "forecast_resin_index_type",
        "forecast_index_value",
        "forecast_index_formula",
    ]
    write_sheet(workbook, "forecast_calculation_audit", forecast_audit_headers, forecast_audit_rows)

    forecast_mapping_headers = [
        "source_country",
        "source_actual_resin_index_type",
        "source_actual_fob",
        "forecast_index_country",
        "forecast_resin_index_type",
        "mapping_basis",
        "mapping_status",
        "missing_reason",
        "actual_index_value_for_model_month",
        "actual_index_model_month",
        "forecast_month_count",
        "forecast_months",
        "index_data_types_used",
    ]
    write_sheet(workbook, "forecast_index_mapping", forecast_mapping_headers, forecast_index_mapping_rows)

    validation_headers = [
        "data_type",
        "source_country",
        "destination_country",
        "time_period",
        "time_period_year",
        "time_period_month",
        "resin_fob",
        "freight",
        "insurance",
        "output_tlc_value",
        "back_calculated_tlc",
        "difference",
        "validation_status",
        "forecast_resin_index_type",
        "included_component_keys",
    ]
    write_sheet(workbook, "tlc_validation", validation_headers, validation_rows)

    formula_rows = [
        {"field": "tlc_formula", "value": TLC_FORMULA},
        {
            "field": "forecast_tlc_formula",
            "value": (
                "For March-December 2026, replace PET resin cost (FOB) with the mapped MR ICIS "
                "monthly index value, keep freight, tax rates, local transport, and other country "
                "inputs unchanged, then recalculate every dependent component and TLC using the "
                "same TLC formula. March-April use MR ICIS rows marked Actual; May-December use "
                "MR ICIS rows marked Forecast."
            ),
        },
        {
            "field": "forecast_index_source",
            "value": summary.get("icis_index_csv", normalized_path(DEFAULT_ICIS_INDEX_CSV)),
        },
        {
            "field": "included_components",
            "value": "; ".join(
                component.raw_label
                for component in mapped_components
                if component.key in INCLUDED_COMPONENT_KEYS
            ),
        },
        {
            "field": "non_required_reference_components",
            "value": "; ".join(
                component.raw_label for component in mapped_components if component.required == "No"
            ),
        },
        {
            "field": "source_country_storage",
            "value": "Source country is stored in Supplier Name and Location for MR rows.",
        },
        {
            "field": "forecast_mapping_visibility",
            "value": "See forecast_index_mapping for direct, regional, and proxy index mappings.",
        },
        {
            "field": "tlc_back_calculation_visibility",
            "value": "See tlc_validation for Actual and Forecast TLC = sum of included components validation.",
        },
    ]
    write_sheet(workbook, "formula_notes", ["field", "value"], formula_rows)

    metadata = {
        **summary,
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "output_xlsx": normalized_path(output_xlsx),
        "output_csv": normalized_path(output_csv),
    }
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(output_xlsx)

    summary_json.write_text(json.dumps(metadata, indent=2, default=str), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the market research front-end data model with actual and forecast TLC rows."
    )
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--icis-index-csv", type=Path, default=DEFAULT_ICIS_INDEX_CSV)
    parser.add_argument("--mapping-xlsx", type=Path, default=DEFAULT_MAPPING_XLSX)
    parser.add_argument("--mapping-sheet", default=DEFAULT_MAPPING_SHEET)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--summary-json", type=Path, default=DEFAULT_SUMMARY_JSON)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    (
        model_rows,
        audit_rows,
        forecast_audit_rows,
        forecast_index_mapping_rows,
        validation_rows,
        summary,
        mapped_components,
    ) = build_rows(
        args.source_xlsx,
        args.mapping_xlsx,
        args.mapping_sheet,
        args.icis_index_csv,
    )
    write_outputs(
        model_rows,
        audit_rows,
        forecast_audit_rows,
        forecast_index_mapping_rows,
        validation_rows,
        summary,
        mapped_components,
        args.output_xlsx,
        args.output_csv,
        args.summary_json,
    )
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
