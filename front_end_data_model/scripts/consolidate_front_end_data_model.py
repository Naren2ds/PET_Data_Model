from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_DIR = REPO_ROOT / "front_end_data_model" / "outputs"
DEFAULT_OUTPUT_XLSX = DEFAULT_OUTPUT_DIR / "Data_Standardized_Front_End_Data_Model.xlsx"
DEFAULT_OUTPUT_CSV = DEFAULT_OUTPUT_DIR / "Data_Standardized_Front_End_Data_Model.csv"
DEFAULT_SUMMARY_JSON = DEFAULT_OUTPUT_DIR / "consolidation_run_summary.json"

SUPPLIER_FRONT_END_PATTERN = (
    "supplier_pipelines/*/artifacts/front_end/*actual_forecast_front_end_standardized.xlsx"
)
BAVARIA_FRONT_END_FILE = (
    REPO_ROOT / "data standardization" / "Bavaria_actual_forecast_final_standardized.xlsx"
)
BAVARIA_FRONT_END_CSV_FALLBACK = (
    REPO_ROOT
    / "Front End"
    / "PET-Backend"
    / "PET-Backend"
    / "standardized_data"
    / "bavaria_actual_forecast_final_standardized.csv"
)

CANONICAL_COLUMNS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]


def clean_column_name(value: Any) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def normalized_path(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def discover_input_files(include_bavaria: bool = True) -> list[Path]:
    paths = sorted(REPO_ROOT.glob(SUPPLIER_FRONT_END_PATTERN))
    if include_bavaria and BAVARIA_FRONT_END_FILE.exists():
        paths.append(BAVARIA_FRONT_END_FILE)
    return paths


def read_excel_table(path: Path) -> tuple[pd.DataFrame, str]:
    excel_file = pd.ExcelFile(path)
    if "front_end_standardized" in excel_file.sheet_names:
        sheet_name = "front_end_standardized"
    elif "final_standardized" in excel_file.sheet_names:
        sheet_name = "final_standardized"
    else:
        sheet_name = excel_file.sheet_names[0]
    return pd.read_excel(path, sheet_name=sheet_name), sheet_name


def read_csv_table(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, encoding="utf-8-sig")


def csv_fallback_for(path: Path) -> Path | None:
    same_name_csv = path.with_suffix(".csv")
    if same_name_csv.exists():
        return same_name_csv
    if path.resolve() == BAVARIA_FRONT_END_FILE.resolve() and BAVARIA_FRONT_END_CSV_FALLBACK.exists():
        return BAVARIA_FRONT_END_CSV_FALLBACK
    return None


def read_source_table(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    audit = {
        "source_file": normalized_path(path),
        "read_status": "success",
        "read_mode": "excel",
        "sheet_name": None,
        "fallback_file": None,
        "error": None,
    }
    try:
        frame, sheet_name = read_excel_table(path)
        audit["sheet_name"] = sheet_name
        return frame, audit
    except PermissionError as exc:
        fallback = csv_fallback_for(path)
        if not fallback:
            audit["read_status"] = "skipped"
            audit["error"] = f"{type(exc).__name__}: {exc}"
            return pd.DataFrame(), audit
        frame = read_csv_table(fallback)
        audit["read_mode"] = "csv_fallback"
        audit["fallback_file"] = normalized_path(fallback)
        audit["error"] = f"Excel locked; read CSV fallback. {type(exc).__name__}: {exc}"
        return frame, audit
    except OSError as exc:
        fallback = csv_fallback_for(path)
        if not fallback:
            audit["read_status"] = "skipped"
            audit["error"] = f"{type(exc).__name__}: {exc}"
            return pd.DataFrame(), audit
        frame = read_csv_table(fallback)
        audit["read_mode"] = "csv_fallback"
        audit["fallback_file"] = normalized_path(fallback)
        audit["error"] = f"Excel unavailable; read CSV fallback. {type(exc).__name__}: {exc}"
        return frame, audit


def apply_location_policy(frame: pd.DataFrame) -> pd.DataFrame:
    if "Location" not in frame.columns:
        frame["Location"] = pd.NA

    supplier = frame.get("Supplier Name", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip()
    destination = frame.get("Destination Country", pd.Series("", index=frame.index))
    location = frame["Location"]
    has_location = location.notna() & location.astype(str).str.strip().ne("")
    keep_source_location = supplier.str.casefold().eq("amcor") & has_location
    frame["Location"] = location.where(keep_source_location, destination)
    return frame


def standardize_frame(frame: pd.DataFrame, source_path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    original_columns = list(frame.columns)
    rename_map = {column: clean_column_name(column) for column in frame.columns}
    frame = frame.rename(columns=rename_map)

    # Restore the two intentionally spaced canonical labels used by the front-end model.
    spacing_map = {
        "Source File": "Source File ",
        "Time_Period": "Time_Period ",
        "Value": "Value ",
    }
    frame = frame.rename(columns={column: spacing_map.get(column, column) for column in frame.columns})
    frame = apply_location_policy(frame)

    added_columns: list[str] = []
    for column in CANONICAL_COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
            added_columns.append(column)

    extra_columns = [column for column in frame.columns if column not in CANONICAL_COLUMNS]
    standardized = frame[CANONICAL_COLUMNS].copy()
    standardized["__source_front_end_file"] = normalized_path(source_path)

    audit = {
        "input_columns": "; ".join(str(column) for column in original_columns),
        "added_columns": "; ".join(added_columns),
        "extra_columns_ignored": "; ".join(extra_columns),
        "rows": len(standardized),
        "location_policy": (
            "Preserved source Location for Supplier Name = Amcor when present; "
            "otherwise Location = Destination Country."
        ),
    }
    return standardized, audit


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 300) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name[:31])
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    style_sheet(worksheet)


def write_outputs(
    consolidated: pd.DataFrame,
    audit_rows: list[dict[str, Any]],
    output_xlsx: Path,
    output_csv: Path,
    summary_json: Path,
) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    data_headers = CANONICAL_COLUMNS
    write_sheet(
        workbook,
        "front_end_data_model",
        data_headers,
        consolidated[data_headers].where(pd.notna(consolidated[data_headers]), None).to_dict("records"),
    )

    audit_headers = [
        "source_file",
        "read_status",
        "read_mode",
        "sheet_name",
        "fallback_file",
        "rows",
        "added_columns",
        "extra_columns_ignored",
        "location_policy",
        "error",
    ]
    write_sheet(workbook, "source_file_audit", audit_headers, audit_rows)

    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "output_xlsx": normalized_path(output_xlsx),
        "output_csv": normalized_path(output_csv),
        "source_file_count": len(audit_rows),
        "loaded_source_file_count": sum(1 for row in audit_rows if row.get("read_status") == "success"),
        "consolidated_rows": len(consolidated),
        "columns": "; ".join(CANONICAL_COLUMNS),
    }
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(output_xlsx)

    consolidated[data_headers].to_csv(output_csv, index=False, encoding="utf-8-sig")
    summary_json.write_text(
        json.dumps({**metadata, "sources": audit_rows}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def consolidate(include_bavaria: bool = True) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    frames: list[pd.DataFrame] = []
    audit_rows: list[dict[str, Any]] = []
    for path in discover_input_files(include_bavaria=include_bavaria):
        frame, read_audit = read_source_table(path)
        if frame.empty:
            read_audit.update(
                {
                    "rows": 0,
                    "added_columns": "",
                    "extra_columns_ignored": "",
                    "location_policy": "",
                }
            )
            audit_rows.append(read_audit)
            continue

        standardized, transform_audit = standardize_frame(frame, path)
        frames.append(standardized)
        audit_rows.append({**read_audit, **transform_audit})

    if not frames:
        return pd.DataFrame(columns=CANONICAL_COLUMNS + ["__source_front_end_file"]), audit_rows
    return pd.concat(frames, ignore_index=True), audit_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Consolidate supplier front-end actual+forecast files into one standardized data model."
    )
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--summary-json", type=Path, default=DEFAULT_SUMMARY_JSON)
    parser.add_argument(
        "--exclude-bavaria",
        action="store_true",
        help="Exclude the legacy Bavaria actual+forecast standardized file.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    consolidated, audit_rows = consolidate(include_bavaria=not args.exclude_bavaria)
    write_outputs(consolidated, audit_rows, args.output_xlsx, args.output_csv, args.summary_json)

    print(f"source_file_count={len(audit_rows)}")
    print(f"loaded_source_file_count={sum(1 for row in audit_rows if row.get('read_status') == 'success')}")
    print(f"consolidated_rows={len(consolidated)}")
    print(f"output_xlsx={args.output_xlsx}")
    print(f"output_csv={args.output_csv}")
    print(f"summary_json={args.summary_json}")
    for row in audit_rows:
        print(f"{row.get('read_status')} | {row.get('rows')} rows | {row.get('source_file')}")


if __name__ == "__main__":
    main()
