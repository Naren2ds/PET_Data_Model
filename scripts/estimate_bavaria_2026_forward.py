from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_FORECAST_CSV = Path("Estimation") / "resin_index_forecast_table.csv"
DEFAULT_OUTPUT_FILE = Path("Estimation") / "bavaria_2026_forward_estimation.xlsx"
DEFAULT_FORECAST_YEAR = 2026
DEFAULT_START_MONTH = 4
DEFAULT_END_MONTH = 12

MONTHLY_FORECAST_SERIES_KEYWORD = "Forecast Apr-2026 Latest"
SOURCE_RESIN_INDEX_TYPE = "ICIS China MID (n-1)"
FORECAST_RESIN_INDEX_TYPE = "PET Bottle Grade FOB China Spot"

DEFAULT_ASSUMPTIONS = {
    "freight_regular": 97.286177169,
    "freight_incremental_override": 6.304731921909095,
    "finance_factor": 0.03496712876712329,
    "duty_rate": 0.05,
    "landed_factor_rate": 0.08,
    "zf_legislation_change": 18.95837133417279,
    "surcharge_alpek_br": 0.0,
}

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def clean_number(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    return float(text)


def load_resin_forecast(
    forecast_csv: Path,
    forecast_year: int,
    start_month: int,
    end_month: int,
) -> pd.DataFrame:
    forecast = pd.read_csv(forecast_csv)
    forecast["time_period_year"] = pd.to_numeric(
        forecast["time_period_year"],
        errors="coerce",
    ).astype("Int64")
    forecast["time_period_month"] = pd.to_numeric(
        forecast["time_period_month"],
        errors="coerce",
    ).astype("Int64")
    forecast["value"] = pd.to_numeric(forecast["value"], errors="coerce")

    year_rows = forecast[
        (forecast["time_period_year"] == forecast_year)
        & (forecast["time_period_month"] >= start_month)
        & (forecast["time_period_month"] <= end_month)
    ].copy()
    if "resin_index_type" in year_rows.columns:
        typed_rows = year_rows[
            year_rows["resin_index_type"].astype(str).str.casefold()
            == FORECAST_RESIN_INDEX_TYPE.casefold()
        ].copy()
        if not typed_rows.empty:
            year_rows = typed_rows

    monthly_rows = year_rows[
        year_rows["forecast_series"].str.contains(
            MONTHLY_FORECAST_SERIES_KEYWORD,
            case=False,
            na=False,
        )
    ].copy()

    selected_rows = []
    for month in range(start_month, end_month + 1):
        month_monthly = monthly_rows[monthly_rows["time_period_month"] == month]
        if not month_monthly.empty:
            selected_rows.append(month_monthly.sort_values("forecast_date").iloc[-1])
            continue

        month_rows = year_rows[year_rows["time_period_month"] == month]
        if month_rows.empty:
            continue

        fallback = month_rows.sort_values("forecast_date").iloc[-1].copy()
        fallback["value"] = month_rows["value"].mean()
        fallback["forecast_series"] = (
            f"Monthly average fallback from {len(month_rows)} forecast rows"
        )
        selected_rows.append(fallback)

    if not selected_rows:
        return pd.DataFrame()

    return pd.DataFrame(selected_rows).reset_index(drop=True)


def estimate_rows(resin_forecast: pd.DataFrame, assumptions: dict[str, float]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _, forecast_row in resin_forecast.iterrows():
        month = int(forecast_row["time_period_month"])
        year = int(forecast_row["time_period_year"])
        resin_price_index = clean_number(forecast_row["value"])
        if resin_price_index is None:
            continue

        freight_regular = assumptions["freight_regular"]
        freight_incremental_override = assumptions["freight_incremental_override"]
        finance_factor = assumptions["finance_factor"]
        duty_rate = assumptions["duty_rate"]
        landed_factor_rate = assumptions["landed_factor_rate"]
        zf_legislation_change = assumptions["zf_legislation_change"]
        surcharge_alpek_br = assumptions["surcharge_alpek_br"]

        freight_incremental_used = freight_incremental_override
        finance_used = (
            resin_price_index
            + freight_regular
            + freight_incremental_used
        ) * finance_factor
        sub_total_incremental = (
            resin_price_index
            + finance_used
            + freight_regular
            + freight_incremental_used
        )
        sub_total_regular = resin_price_index + finance_used + freight_regular
        duty = sub_total_incremental * duty_rate
        landed_factor = sub_total_regular * landed_factor_rate
        total_landing_cost_forecast = (
            sub_total_incremental
            + duty
            + landed_factor
            + zf_legislation_change
            + surcharge_alpek_br
        )

        rows.append(
            {
                "time_period": f"{MONTH_LABELS[month]} {year}",
                "time_period_year": year,
                "time_period_month": month,
                "resin_price_index": resin_price_index,
                "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
                "forecast_resin_index_type": forecast_row.get(
                    "resin_index_type",
                    FORECAST_RESIN_INDEX_TYPE,
                ),
                "resin_forecast_date": forecast_row.get("forecast_date"),
                "resin_forecast_series": forecast_row.get("forecast_series"),
                "freight_regular": freight_regular,
                "freight_incremental_override": freight_incremental_override,
                "freight_incremental_used": freight_incremental_used,
                "finance_factor": finance_factor,
                "finance_used": finance_used,
                "sub_total_incremental": sub_total_incremental,
                "sub_total_regular": sub_total_regular,
                "duty_rate": duty_rate,
                "duty": duty,
                "landed_factor_rate": landed_factor_rate,
                "landed_factor": landed_factor,
                "zf_legislation_change": zf_legislation_change,
                "surcharge_alpek_br": surcharge_alpek_br,
                "total_landing_cost_forecast": total_landing_cost_forecast,
                "formula": (
                    "TLC = sub_total_incremental + duty + landed_factor + "
                    "zf_legislation_change + surcharge_alpek_br"
                ),
            }
        )

    return pd.DataFrame(rows)


def standardized_component_rows(estimation: pd.DataFrame) -> pd.DataFrame:
    component_specs = [
        ("Resin Price Index", "Resin Index vPET", "resin_price_index"),
        ("Freight China-Buenaventura (Regular)", "Freight", "freight_regular"),
        ("Freight China-Buenaventura (Incremental)", "Freight", "freight_incremental_used"),
        ("Finance", "Insurance", "finance_used"),
        ("Sub Total (with Incremental Freight)", "Sub Total (CIF)", "sub_total_incremental"),
        ("Sub Total (with Regular Freight)", "Sub Total (CIF)", "sub_total_regular"),
        ("Duty 5% (Change According to Regulation)", "Tax", "duty"),
        ("Landed Factor 8%", "Tax", "landed_factor"),
        ("ZF Legislation Change", "Tax", "zf_legislation_change"),
        ("Sur Charge Alpek Br", "Tax", "surcharge_alpek_br"),
        ("Total Resin Price ABI VIRGIN Formula", "Total Landing Cost", "total_landing_cost_forecast"),
    ]

    rows: list[dict[str, Any]] = []
    for _, estimate_row in estimation.iterrows():
        for raw_cost_breakdown, mapping_column, value_column in component_specs:
            rows.append(
                {
                    "Source File ": "resin_index_forecast_table.csv",
                    "Supplier Name": "Amcor",
                    "Destination Country": "Colombia",
                    "Time_Period ": estimate_row["time_period"],
                    "Raw Cost Breakdown": raw_cost_breakdown,
                    "Resin Index Type": estimate_row.get("source_resin_index_type"),
                    "Forecast Resin Index Type": estimate_row.get("forecast_resin_index_type"),
                    "Mapping Columns": mapping_column,
                    "Value ": estimate_row[value_column],
                    "forecast_source": estimate_row.get("resin_forecast_series"),
                }
            )
    return pd.DataFrame(rows)


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for col_idx in range(1, worksheet.max_column + 1):
                header = worksheet.cell(1, col_idx).value
                width = len(str(header or ""))
                for row_idx in range(2, min(worksheet.max_row, 250) + 1):
                    value = worksheet.cell(row_idx, col_idx).value
                    if value is not None:
                        width = max(width, len(str(value)))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(width + 2, 12),
                    70,
                )
        workbook.save(path)
    finally:
        workbook.close()


def write_output(
    output_file: Path,
    estimation: pd.DataFrame,
    standardized_rows: pd.DataFrame,
    assumptions: dict[str, float],
    metadata: dict[str, Any],
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    assumptions_df = pd.DataFrame(
        [{"assumption": key, "value": value} for key, value in assumptions.items()]
    )
    metadata_df = pd.DataFrame(
        [{"field": key, "value": value} for key, value in metadata.items()]
    )
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        estimation.to_excel(writer, sheet_name="bavaria_2026_estimate", index=False)
        standardized_rows.to_excel(writer, sheet_name="standardized_rows", index=False)
        assumptions_df.to_excel(writer, sheet_name="assumptions", index=False)
        metadata_df.to_excel(writer, sheet_name="run_metadata", index=False)
    style_workbook(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Estimate Bavaria forward 2026 TLC using the Resin Index forecast table."
    )
    parser.add_argument("--forecast-csv", type=Path, default=DEFAULT_FORECAST_CSV)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--forecast-year", type=int, default=DEFAULT_FORECAST_YEAR)
    parser.add_argument("--start-month", type=int, default=DEFAULT_START_MONTH)
    parser.add_argument("--end-month", type=int, default=DEFAULT_END_MONTH)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    resin_forecast = load_resin_forecast(
        args.forecast_csv,
        args.forecast_year,
        args.start_month,
        args.end_month,
    )
    estimation = estimate_rows(resin_forecast, DEFAULT_ASSUMPTIONS)
    standardized_rows = standardized_component_rows(estimation)
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "forecast_csv": str(args.forecast_csv),
        "output_file": str(args.output_file),
        "forecast_year": args.forecast_year,
        "start_month": args.start_month,
        "end_month": args.end_month,
        "months_estimated": len(estimation),
        "resin_forecast_selection": (
            f"Forecast resin index type must match {FORECAST_RESIN_INDEX_TYPE!r}. "
            f"Prefer rows whose forecast_series contains {MONTHLY_FORECAST_SERIES_KEYWORD!r}; "
            "fallback to monthly average if unavailable."
        ),
        "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
        "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
        "freight_incremental_override_policy": (
            "Carried forward from latest Bavaria historical month, March 2026."
        ),
    }
    write_output(args.output_file, estimation, standardized_rows, DEFAULT_ASSUMPTIONS, metadata)

    print(f"forecast_csv={args.forecast_csv}")
    print(f"output_file={args.output_file}")
    print(f"months_estimated={len(estimation)}")
    if not estimation.empty:
        print(f"first_month={estimation.iloc[0]['time_period']}")
        print(f"last_month={estimation.iloc[-1]['time_period']}")


if __name__ == "__main__":
    main()
