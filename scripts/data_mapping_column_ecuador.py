from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT_FILE = (
    Path("output") / "3 Precios AB Inbev_Abril 26 ECUADOR_formula_virgen_final.xlsx"
)
DEFAULT_MAPPING_FILE = Path("mapping files") / "Mapping_Columns.xlsx"
DEFAULT_MAPPING_SHEET = "3 Precios AB Inbev_Abril 26 ECU"
DEFAULT_FINAL_SHEET = "final_data"
LOOKUP_COLUMN = "metric_en"
MAPPING_VALUE_COLUMN = "Mapping Columns"


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or ""))
    text = text.replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def safe_sheet_title(title: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", title)[:31]


def load_mapping(mapping_file: Path, mapping_sheet: str) -> dict[str, Any]:
    workbook = load_workbook(mapping_file, data_only=True, read_only=True)
    try:
        if mapping_sheet not in workbook.sheetnames:
            raise ValueError(
                f"Mapping sheet {mapping_sheet!r} not found in {mapping_file}."
            )
        worksheet = workbook[mapping_sheet]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        lookup_idx = headers.index(LOOKUP_COLUMN)
        mapping_idx = headers.index(MAPPING_VALUE_COLUMN)

        mapping: dict[str, Any] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            lookup_value = row[lookup_idx]
            mapping_value = row[mapping_idx]
            if lookup_value is None:
                continue
            mapping[normalize_key(lookup_value)] = mapping_value
        return mapping
    finally:
        workbook.close()


def style_header(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def adjust_widths(worksheet: Any) -> None:
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            55,
        )


def write_rows_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    headers: list[str],
) -> None:
    title = safe_sheet_title(sheet_name)
    if title in workbook.sheetnames:
        del workbook[title]
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    adjust_widths(worksheet)


def apply_mapping(
    output_file: Path,
    mapping_file: Path,
    mapping_sheet: str,
    final_sheet: str,
) -> dict[str, Any]:
    mapping = load_mapping(mapping_file, mapping_sheet)
    workbook = load_workbook(output_file, data_only=False)
    try:
        if final_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {final_sheet!r} not found in {output_file}.")

        worksheet = workbook[final_sheet]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        if LOOKUP_COLUMN not in headers:
            raise ValueError(f"Column {LOOKUP_COLUMN!r} not found in {final_sheet}.")

        lookup_col = headers.index(LOOKUP_COLUMN) + 1
        if MAPPING_VALUE_COLUMN in headers:
            mapping_col = headers.index(MAPPING_VALUE_COLUMN) + 1
        else:
            mapping_col = lookup_col + 1
            worksheet.insert_cols(mapping_col)
            worksheet.cell(1, mapping_col).value = MAPPING_VALUE_COLUMN

        unmatched: dict[str, int] = {}
        updated = 0
        matched = 0
        blank_lookup = 0
        for row_idx in range(2, worksheet.max_row + 1):
            lookup_value = worksheet.cell(row_idx, lookup_col).value
            normalized = normalize_key(lookup_value)
            mapping_value = mapping.get(normalized)
            worksheet.cell(row_idx, mapping_col).value = mapping_value
            updated += 1

            if lookup_value is None or str(lookup_value).strip() == "":
                blank_lookup += 1
            elif mapping_value is None:
                unmatched[str(lookup_value)] = unmatched.get(str(lookup_value), 0) + 1
            else:
                matched += 1

        style_header(worksheet)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        adjust_widths(worksheet)

        mapping_rows = [
            {LOOKUP_COLUMN: key, MAPPING_VALUE_COLUMN: value}
            for key, value in sorted(mapping.items())
        ]
        unmatched_rows = [
            {
                LOOKUP_COLUMN: key,
                "row_count": count,
                "normalized_key": normalize_key(key),
            }
            for key, count in sorted(unmatched.items())
        ]
        metadata_rows = [
            {"field": "output_file", "value": str(output_file)},
            {"field": "final_sheet", "value": final_sheet},
            {"field": "mapping_file", "value": str(mapping_file)},
            {"field": "mapping_sheet", "value": mapping_sheet},
            {"field": "lookup_column", "value": LOOKUP_COLUMN},
            {"field": "mapping_value_column", "value": MAPPING_VALUE_COLUMN},
            {"field": "updated_rows", "value": updated},
            {"field": "matched_rows", "value": matched},
            {"field": "blank_lookup_rows", "value": blank_lookup},
            {"field": "unmatched_nonblank_metric_count", "value": len(unmatched)},
            {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
        ]
        write_rows_to_sheet(
            workbook,
            "mapping_source",
            mapping_rows,
            [LOOKUP_COLUMN, MAPPING_VALUE_COLUMN],
        )
        write_rows_to_sheet(
            workbook,
            "unmatched_mapping",
            unmatched_rows,
            [LOOKUP_COLUMN, "row_count", "normalized_key"],
        )
        write_rows_to_sheet(
            workbook,
            "mapping_run_metadata",
            metadata_rows,
            ["field", "value"],
        )
        workbook.save(output_file)

        return {
            "updated_rows": updated,
            "matched_rows": matched,
            "blank_lookup_rows": blank_lookup,
            "unmatched_nonblank_metric_count": len(unmatched),
            "output_file": str(output_file),
        }
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add Mapping Columns to the Ecuador final_data sheet by metric_en."
    )
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--mapping-file", type=Path, default=DEFAULT_MAPPING_FILE)
    parser.add_argument("--mapping-sheet", default=DEFAULT_MAPPING_SHEET)
    parser.add_argument("--final-sheet", default=DEFAULT_FINAL_SHEET)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = apply_mapping(
        output_file=args.output_file,
        mapping_file=args.mapping_file,
        mapping_sheet=args.mapping_sheet,
        final_sheet=args.final_sheet,
    )
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
