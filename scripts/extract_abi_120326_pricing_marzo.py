from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_FILE_PATTERN = "ABI*Pricing*Marzo*2026*.xlsx"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_SHEET = "CSD Resina Ush"
DEFAULT_HEADER_RANGE = "B8:U8"
DEFAULT_DATA_RANGE = "B56:U61"
DEFAULT_METRIC_LABEL_ROW = 2
DEFAULT_SECTION_ROW = 55
DEFAULT_RAW_CSV_NAME = "abi_120326_pricing_marzo_csd_resina_ush_raw.csv"
DEFAULT_LONG_CSV_NAME = "abi_120326_pricing_marzo_csd_resina_ush_long.csv"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_sheet_name(value: str) -> str:
    value = strip_accents(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def format_header_value(value: Any) -> Any:
    value = clean_value(value)
    if isinstance(value, datetime):
        return f"{MONTH_LABELS[value.month]} {value.year}"
    if isinstance(value, date):
        return f"{MONTH_LABELS[value.month]} {value.year}"
    return clean_text(value) if isinstance(value, str) else value


def period_fields(period_value: Any) -> dict[str, Any]:
    if isinstance(period_value, datetime):
        period_date = period_value.date()
    elif isinstance(period_value, date):
        period_date = period_value
    else:
        return {
            "time_period": None,
            "time_period_year": None,
            "time_period_month": None,
        }

    return {
        "time_period": f"{MONTH_LABELS[period_date.month]} {period_date.year}",
        "time_period_year": period_date.year,
        "time_period_month": period_date.month,
    }


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def find_sheet(workbook: Any, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name

    normalized_target = normalize_sheet_name(sheet_name)
    matches = [
        candidate
        for candidate in workbook.sheetnames
        if normalize_sheet_name(candidate) == normalized_target
    ]
    if len(matches) == 1:
        return matches[0]

    raise ValueError(f"Sheet not found: {sheet_name}")


def extract_raw_rows(worksheet: Any, range_refs: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for range_ref in range_refs:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        for row_idx in range(min_row, max_row + 1):
            row = {
                "source_range": range_ref,
                "source_row": row_idx,
            }
            for col_idx in range(min_col, max_col + 1):
                column_letter = get_column_letter(col_idx)
                row[column_letter] = clean_value(worksheet.cell(row_idx, col_idx).value)
            rows.append(row)
    return rows


def extract_final_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    header_range: str,
    data_range: str,
    metric_label_row: int,
    section_row: int,
) -> list[dict[str, Any]]:
    header_min_col, header_row, header_max_col, _ = range_boundaries(header_range)
    data_min_col, data_min_row, data_max_col, data_max_row = range_boundaries(data_range)
    if (header_min_col, header_max_col) != (data_min_col, data_max_col):
        raise ValueError("Header range and data range must cover the same columns.")

    period_col = data_min_col
    metric_cols = range(data_min_col + 1, data_max_col + 1)
    section_name = clean_value(value_worksheet.cell(section_row, period_col).value)
    supplier = clean_value(value_worksheet.cell(1, period_col).value)
    rows: list[dict[str, Any]] = []

    for row_idx in range(data_min_row, data_max_row + 1):
        price_period = clean_value(value_worksheet.cell(row_idx, period_col).value)
        period_source_cell = f"{get_column_letter(period_col)}{row_idx}"

        for col_idx in metric_cols:
            value = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            if value is None:
                continue

            column_letter = get_column_letter(col_idx)
            metric_label = clean_text(value_worksheet.cell(metric_label_row, col_idx).value)
            requested_header_value = format_header_value(
                value_worksheet.cell(header_row, col_idx).value
            )
            rows.append(
                {
                    "source_file": str(source_file),
                    "source_sheet": sheet_name,
                    "source_range": data_range,
                    "source_cell": f"{column_letter}{row_idx}",
                    "source_row": row_idx,
                    "section_name": section_name,
                    "supplier": supplier,
                    "metric_name": metric_label or f"field_{column_letter}",
                    "metric_label_original": metric_label,
                    "metric_label_source_cell": f"{column_letter}{metric_label_row}",
                    "requested_header_range": header_range,
                    "requested_header_source_cell": f"{column_letter}{header_row}",
                    "requested_header_value": requested_header_value,
                    "requested_header_formula": formula_worksheet.cell(header_row, col_idx).value,
                    "period_source_cell": period_source_cell,
                    "price_period": price_period,
                    **period_fields(price_period),
                    "value": value,
                    "formula": formula_worksheet.cell(row_idx, col_idx).value,
                }
            )

    return rows


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=ordered_headers(rows))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_csd_resina_ush_final.xlsx"


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        return

    headers = ordered_headers(rows)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(row.get(header) for row in rows[:100])
        width = min(
            max(len(str(value)) for value in values if value is not None) + 2,
            55,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for key, value in metadata.items():
        worksheet.append([key, value])

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 80


def write_excel_output(
    path: Path,
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, "raw_B8_U8_B56_U61", raw_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the ABI 120326 Pricing Marzo CSD Resina Ush extraction pipeline."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Specific workbook to process. If omitted, the latest ABI Pricing Marzo workbook in data is used.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--header-range", default=DEFAULT_HEADER_RANGE)
    parser.add_argument("--data-range", default=DEFAULT_DATA_RANGE)
    parser.add_argument("--metric-label-row", type=int, default=DEFAULT_METRIC_LABEL_ROW)
    parser.add_argument("--section-row", type=int, default=DEFAULT_SECTION_ROW)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=None,
        help="Optional explicit Excel output path.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files into the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)
    excel_output = args.excel_output or default_excel_output_path(
        source_file,
        args.output_dir,
    )

    values_workbook = load_workbook(source_file, read_only=True, data_only=True)
    formulas_workbook = load_workbook(source_file, read_only=True, data_only=False)

    try:
        sheet_name = find_sheet(values_workbook, args.sheet)
        value_worksheet = values_workbook[sheet_name]
        formula_worksheet = formulas_workbook[sheet_name]

        raw_rows = extract_raw_rows(
            value_worksheet,
            [args.header_range, args.data_range],
        )
        final_rows = extract_final_rows(
            value_worksheet=value_worksheet,
            formula_worksheet=formula_worksheet,
            source_file=source_file,
            sheet_name=sheet_name,
            header_range=args.header_range,
            data_range=args.data_range,
            metric_label_row=args.metric_label_row,
            section_row=args.section_row,
        )

        metadata = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(source_file),
            "source_sheet": sheet_name,
            "header_range": args.header_range,
            "data_range": args.data_range,
            "metric_label_row": args.metric_label_row,
            "section_row": args.section_row,
            "raw_rows": len(raw_rows),
            "final_rows": len(final_rows),
            "file_pattern": args.file_pattern,
        }

        write_excel_output(excel_output, raw_rows, final_rows, metadata)

        if args.write_csv:
            write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
            write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

        print(f"source_file={source_file}")
        print(f"source_sheet={sheet_name}")
        print(f"header_range={args.header_range}")
        print(f"data_range={args.data_range}")
        print(f"raw_rows={len(raw_rows)}")
        print(f"final_rows={len(final_rows)}")
        print(f"excel_output={excel_output}")
        if args.write_csv:
            print(f"raw_csv_output={args.output_dir / DEFAULT_RAW_CSV_NAME}")
            print(f"final_csv_output={args.output_dir / DEFAULT_LONG_CSV_NAME}")
    finally:
        values_workbook.close()
        formulas_workbook.close()


if __name__ == "__main__":
    main()
