<<<<<<< ours
<<<<<<< ours
from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_FILE_PATTERN = "*Bavaria*.xlsx"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_RANGE = "B13:CV26"
DEFAULT_RAW_CSV_NAME = "bavaria_formula_raw.csv"
DEFAULT_LONG_CSV_NAME = "bavaria_formula_long.csv"

MONTH_ALIASES = {
    "enero": {"ene", "enero", "jan", "january"},
    "febrero": {"feb", "febrero", "february"},
    "marzo": {"mar", "marzo", "marco", "march"},
    "abril": {"abr", "abril", "apr", "april"},
    "mayo": {"may", "mayo"},
    "junio": {"jun", "junio", "june"},
    "julio": {"jul", "julio", "july"},
    "agosto": {"ago", "agosto", "aug", "august"},
    "septiembre": {"sep", "sept", "septiembre", "september"},
    "octubre": {"oct", "octubre", "october"},
    "noviembre": {"nov", "noviembre", "november"},
    "diciembre": {"dic", "diciembre", "dec", "december"},
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize(value: str) -> str:
    value = strip_accents(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def canonical_month_from_text(value: str) -> str | None:
    tokens = set(normalize(value).split())
    for month, aliases in MONTH_ALIASES.items():
        if tokens.intersection(aliases):
            return month
    return None


def select_formula_sheet(workbook: Any, source_file: Path, explicit_sheet: str | None) -> str:
    if explicit_sheet:
        if explicit_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {explicit_sheet}")
        return explicit_sheet

    formula_sheets = [
        sheet_name
        for sheet_name in workbook.sheetnames
        if "formula" in normalize(sheet_name)
    ]
    if not formula_sheets:
        raise ValueError("No sheet containing 'Formula' was found.")

    source_month = canonical_month_from_text(source_file.stem)
    if source_month:
        matching_sheets = [
            sheet_name
            for sheet_name in formula_sheets
            if canonical_month_from_text(sheet_name) == source_month
        ]
        if len(matching_sheets) == 1:
            return matching_sheets[0]
        if len(matching_sheets) > 1:
            raise ValueError(
                "Multiple formula sheets matched the source month "
                f"{source_month}: {matching_sheets}"
            )

    if len(formula_sheets) == 1:
        return formula_sheets[0]

    raise ValueError(
        "Multiple formula sheets found and no unique month match was possible: "
        f"{formula_sheets}"
    )


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def extract_raw_rows(worksheet: Any, range_ref: str) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    columns = [get_column_letter(col_idx) for col_idx in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(min_row, max_row + 1):
        row = {"source_row": row_idx}
        for col_idx, column_letter in zip(range(min_col, max_col + 1), columns):
            row[column_letter] = clean_value(worksheet.cell(row_idx, col_idx).value)
        rows.append(row)

    return rows


def extract_long_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    min_col, _, max_col, _ = range_boundaries(range_ref)
    data_start_col = max(min_col, 5)
    rows: list[dict[str, Any]] = []

    for row_idx in range(16, 27):
        if row_idx == 16:
            metric_en = value_worksheet["B15"].value
            metric_local = value_worksheet["D15"].value
            metric_label_source_row = 15
        else:
            metric_en = value_worksheet.cell(row_idx, 2).value
            metric_local = value_worksheet.cell(row_idx, 4).value
            metric_label_source_row = row_idx

        for col_idx in range(data_start_col, max_col + 1):
            value = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            if value is None:
                continue

            column_letter = get_column_letter(col_idx)
            rows.append(
                {
                    "source_file": str(source_file),
                    "source_sheet": sheet_name,
                    "source_range": range_ref,
                    "source_cell": f"{column_letter}{row_idx}",
                    "metric_row": row_idx,
                    "metric_label_source_row": metric_label_source_row,
                    "metric_en": metric_en,
                    "metric_local": metric_local,
                    "price_period": value_worksheet.cell(14, col_idx).value,
                    "index_period": value_worksheet.cell(15, col_idx).value,
                    "value": value,
                    "formula": formula_worksheet.cell(row_idx, col_idx).value,
                }
            )

    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_formula_final.xlsx"


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(row.get(header) for row in rows[:100])
        width = min(
            max(len(str(value)) for value in values if value is not None) + 2,
            45,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for key, value in metadata.items():
        worksheet.append([key, value])

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 70


def write_excel_output(
    path: Path,
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
    raw_sheet_title: str = "raw_B13_CV26",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, raw_sheet_title, raw_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the monthly Bavaria formula extraction pipeline."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Specific workbook to process. If omitted, the latest Bavaria workbook in data is used.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--range", default=DEFAULT_RANGE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=None,
        help="Optional explicit Excel output path.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files into the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)
    excel_output = args.excel_output or default_excel_output_path(
        source_file, args.output_dir
    )

    values_workbook = load_workbook(source_file, read_only=True, data_only=True)
    formulas_workbook = load_workbook(source_file, read_only=True, data_only=False)

    try:
        sheet_name = select_formula_sheet(values_workbook, source_file, args.sheet)
        value_worksheet = values_workbook[sheet_name]
        formula_worksheet = formulas_workbook[sheet_name]

        raw_rows = extract_raw_rows(value_worksheet, args.range)
        final_rows = extract_long_rows(
            value_worksheet,
            formula_worksheet,
            source_file,
            sheet_name,
            args.range,
        )

        metadata = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(source_file),
            "source_sheet": sheet_name,
            "source_range": args.range,
            "anchor_B13": repr(value_worksheet["B13"].value),
            "raw_rows": len(raw_rows),
            "final_rows": len(final_rows),
        }

        write_excel_output(excel_output, raw_rows, final_rows, metadata)

        if args.write_csv:
            write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
            write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

        print(f"source_file={source_file}")
        print(f"source_sheet={sheet_name}")
        print(f"source_range={args.range}")
        print(f"anchor_B13={value_worksheet['B13'].value!r}")
        print(f"raw_rows={len(raw_rows)}")
        print(f"final_rows={len(final_rows)}")
        print(f"excel_output={excel_output}")
        if args.write_csv:
            print(f"raw_csv_output={args.output_dir / DEFAULT_RAW_CSV_NAME}")
            print(f"final_csv_output={args.output_dir / DEFAULT_LONG_CSV_NAME}")
    finally:
        values_workbook.close()
        formulas_workbook.close()
=======
=======
>>>>>>> theirs
#!/usr/bin/env python3
from __future__ import annotations
import argparse
import csv
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path

NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def excel_serial_to_date(value: float) -> str:
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=float(value))).strftime("%Y-%m-%d")


def col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n


def num_to_col(n: int) -> str:
    s = ""
    while n:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def parse_xlsx_cells(xlsx_path: Path, sheet_regex: str) -> tuple[str, dict[str, str]]:
    with zipfile.ZipFile(xlsx_path) as zf:
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = []
        for sh in wb.find("a:sheets", NS):
            sheets.append((sh.attrib["name"], sh.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]))

        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}

        shared = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", NS):
                texts = [t.text or "" for t in si.findall(".//a:t", NS)]
                shared.append("".join(texts))

        match_name, match_target = None, None
        pattern = re.compile(sheet_regex)
        for name, rid in sheets:
            if pattern.search(name):
                match_name = name
                match_target = rid_to_target[rid]
                break

        if not match_name:
            raise ValueError(f"No worksheet matched regex: {sheet_regex}")

        xml_path = "xl/" + match_target.lstrip("/")
        ws_xml = ET.fromstring(zf.read(xml_path))

        data: dict[str, str] = {}
        for c in ws_xml.findall(".//a:sheetData/a:row/a:c", NS):
            ref = c.attrib.get("r")
            if not ref:
                continue
            t = c.attrib.get("t")
            v = c.find("a:v", NS)
            if v is None or v.text is None:
                data[ref] = ""
            elif t == "s":
                data[ref] = shared[int(v.text)]
            else:
                data[ref] = v.text

        return match_name, data


def extract_to_long(data: dict[str, str], sheet_name: str, row_start: int, row_end: int, col_start: str, col_end: str):
    c_start, c_end = col_to_num(col_start), col_to_num(col_end)
    month_labels = []
    for c in range(c_start, c_end + 1):
        col_ref = f"{num_to_col(c)}{row_start}"
        raw = data.get(col_ref, "")
        label = raw
        if raw:
            try:
                f = float(raw)
                if f > 30000:
                    label = excel_serial_to_date(f)
            except ValueError:
                pass
        month_labels.append((num_to_col(c), raw, label))

    rows = []
    for r in range(row_start + 1, row_end + 1):
        metric = data.get(f"B{r}", "") or data.get(f"C{r}", "")
        for col_letter, raw_month, norm_month in month_labels:
            val = data.get(f"{col_letter}{r}", "")
            rows.append({
                "source_file": "1 Bavaria - Precio Marzo.xlsx",
                "sheet_name": sheet_name,
                "metric_row": r,
                "metric_name": metric,
                "column_letter": col_letter,
                "period_raw": raw_month,
                "period_normalized": norm_month,
                "value_raw": val,
            })
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract Bavaria formula range and normalize to long format.")
    parser.add_argument("--input", default="data/1 Bavaria - Precio Marzo.xlsx")
    parser.add_argument("--output", default="output/bavaria_formula_mar26_long.csv")
    parser.add_argument("--sheet-regex", default=r"^Fórmula.*26")
    args = parser.parse_args()

    sheet_name, data = parse_xlsx_cells(Path(args.input), args.sheet_regex)
    rows = extract_to_long(data, sheet_name, 15, 26, "C", "CV")

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    print(f"Sheet selected: {sheet_name}")
    print(f"Rows written: {len(rows)}")
    print(f"Output: {out}")
<<<<<<< ours
>>>>>>> theirs
=======
>>>>>>> theirs


if __name__ == "__main__":
    main()
