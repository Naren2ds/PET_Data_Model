from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
BAVARIA_FILE_CANDIDATES = [
    BASE_DIR / "output" / "1 Bavaria - Precio Marzo_formula_final.xlsx",
    BASE_DIR / "output" / "1 Bavaria - Precio Marzo_formula_final_with_mapping_columns.xlsx",
]
MAPPING_FILE = BASE_DIR / "mapping files" / "Mapping_Columns.xlsx"
MAPPING_SHEET = "1 Bavaria - Precio Marzo_formul"
OUTPUT_FILE = (
    BASE_DIR / "output" / "1 Bavaria - Precio Marzo_formula_final_with_mapping_columns.xlsx"
)


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or ""))
    text = text.replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def clean_resin_index_type(value: Any) -> Any:
    value = clean_text(value)
    if isinstance(value, str):
        return value.rstrip(" .")
    return value


def derive_resin_index_type(rows: list[dict[str, Any]]) -> Any:
    for row in rows:
        if row.get("metric_en") == "Resin Price Index":
            resin_index_type = clean_resin_index_type(row.get("metric_local"))
            if resin_index_type:
                return resin_index_type
    return None


def load_mapping() -> dict[str, Any]:
    workbook = load_workbook(MAPPING_FILE, data_only=True, read_only=True)
    worksheet = workbook[MAPPING_SHEET]

    headers = [clean_header(cell.value) for cell in worksheet[1]]
    metric_idx = headers.index("metric_local")
    mapping_idx = headers.index("Mapping Columns")

    mapping: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        metric_local = row[metric_idx]
        mapping_value = row[mapping_idx]
        if metric_local is None:
            continue
        mapping[normalize_key(metric_local)] = mapping_value
    return mapping


def resolve_bavaria_file() -> Path:
    for path in BAVARIA_FILE_CANDIDATES:
        if path.exists():
            return path
    expected = ", ".join(str(path) for path in BAVARIA_FILE_CANDIDATES)
    raise FileNotFoundError(f"Could not find a Bavaria final workbook. Checked: {expected}")


def read_sheet_rows(workbook: Any, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in worksheet[1]]
    rows = [
        dict(zip(headers, values))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    return headers, rows


def ordered_headers(original_headers: list[str]) -> list[str]:
    headers: list[str] = []
    for header in original_headers:
        if header in {"resin_index_type", "Mapping Columns"}:
            continue
        headers.append(header)
        if header == "metric_local":
            headers.append("resin_index_type")
            headers.append("Mapping Columns")
    return headers


def write_rows(workbook: Workbook, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = worksheet.cell(row_idx, col_idx, value=value)
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            55,
        )


def main() -> None:
    mapping = load_mapping()
    bavaria_file = resolve_bavaria_file()
    bavaria_workbook = load_workbook(bavaria_file, data_only=False, read_only=True)
    original_headers, rows = read_sheet_rows(bavaria_workbook, "final_data")
    resin_index_type = derive_resin_index_type(rows)

    unmatched: set[Any] = set()
    for row in rows:
        metric_local = row.get("metric_local")
        mapping_value = mapping.get(normalize_key(metric_local))
        row["resin_index_type"] = resin_index_type
        row["Mapping Columns"] = mapping_value
        if mapping_value is None:
            unmatched.add(metric_local)

    output_headers = ordered_headers(original_headers)
    metadata_rows = [
        {"field": "source_file", "value": str(bavaria_file)},
        {"field": "mapping_file", "value": str(MAPPING_FILE)},
        {"field": "mapping_sheet", "value": MAPPING_SHEET},
        {"field": "lookup_key", "value": "final_data.metric_local"},
        {"field": "resin_index_type", "value": resin_index_type},
        {"field": "mapping_column", "value": "Mapping Columns"},
        {"field": "row_count", "value": len(rows)},
        {"field": "unmatched_metric_local_count", "value": len(unmatched)},
        {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
    ]
    unmatched_rows = [
        {"metric_local": value, "normalized_key": normalize_key(value)}
        for value in sorted(unmatched, key=lambda item: str(item))
    ]
    mapping_rows = [
        {"metric_local_key": key, "Mapping Columns": value}
        for key, value in sorted(mapping.items())
    ]

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    write_rows(output_workbook, "final_data", output_headers, rows)
    write_rows(output_workbook, "mapping_source", ["metric_local_key", "Mapping Columns"], mapping_rows)
    write_rows(output_workbook, "unmatched", ["metric_local", "normalized_key"], unmatched_rows)
    write_rows(output_workbook, "run_metadata", ["field", "value"], metadata_rows)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(OUTPUT_FILE)

    print(f"Output: {OUTPUT_FILE}")
    print(f"Rows: {len(rows)}")
    print(f"Unmatched metric_local values: {len(unmatched)}")
    if unmatched:
        for value in sorted(unmatched, key=lambda item: str(item)):
            print(f"  - {value}")


if __name__ == "__main__":
    main()
