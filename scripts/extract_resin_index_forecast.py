from __future__ import annotations

import argparse
import json
import re
import subprocess
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_FORECAST_DIR = Path("Forecast")
DEFAULT_FILE_PATTERN = "*Forecast*.xls*"
DEFAULT_OUTPUT_DIR = Path("Estimation")
DEFAULT_SHEET = "ICIS Price Forecast"
DEFAULT_RANGE = "B13:D82"
DEFAULT_EXCEL_OUTPUT = DEFAULT_OUTPUT_DIR / "resin_index_forecast_table.xlsx"
DEFAULT_CSV_OUTPUT = DEFAULT_OUTPUT_DIR / "resin_index_forecast_table.csv"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def resolve_source_file(
    explicit_file: Path | None,
    forecast_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Forecast file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in forecast_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No forecast files found in {forecast_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def excel_serial_to_date(value: float) -> date:
    return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()


def parse_date(text: Any, value2: Any) -> date | None:
    if isinstance(value2, (int, float)) and value2 > 20000:
        return excel_serial_to_date(float(value2))

    if not text:
        return None

    cleaned = str(text).strip()
    for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(text: Any, value2: Any) -> float | None:
    if isinstance(value2, bool) or value2 is None:
        return None
    if isinstance(value2, (int, float)):
        return float(value2)

    if text is None:
        return None

    cleaned = str(text).strip().replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def clean_text(value: Any) -> Any:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def infer_resin_index_type(forecast_series: Any) -> Any:
    text = clean_text(forecast_series)
    if not isinstance(text, str):
        return None
    text = re.sub(r"\s+Forecast\b.*$", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\s+4-12 Weeks\b.*$", "", text, flags=re.IGNORECASE).strip()
    return text or clean_text(forecast_series)


def normalize_formula(value: Any) -> Any:
    text = clean_text(value)
    if text in (None, ""):
        return None
    return text


def read_xls_range_with_excel_com(source_file: Path, sheet_name: str, range_ref: str) -> list[dict[str, Any]]:
    powershell = r"""
param(
  [Parameter(Mandatory=$true)][string]$Path,
  [Parameter(Mandatory=$true)][string]$SheetName,
  [Parameter(Mandatory=$true)][string]$RangeRef
)

$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false

try {
  $workbook = $excel.Workbooks.Open($Path, 0, $true)
  $worksheet = $workbook.Worksheets.Item($SheetName)
  $range = $worksheet.Range($RangeRef)
  $rows = @()

  foreach ($cell in $range.Cells) {
    $rows += [pscustomobject]@{
      Row = [int]$cell.Row
      Column = [int]$cell.Column
      Address = [string]$cell.Address($false, $false)
      Text = [string]$cell.Text
      Value2 = $cell.Value2
      Formula = [string]$cell.Formula
    }
  }

  $rows | ConvertTo-Json -Depth 4 -Compress
  $workbook.Close($false)
}
finally {
  if ($workbook) {
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
  }
  if ($worksheet) {
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($worksheet) | Out-Null
  }
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}
"""
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as script:
        script.write(powershell)
        script_path = Path(script.name)

    try:
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-Path",
                str(source_file.resolve()),
                "-SheetName",
                sheet_name,
                "-RangeRef",
                range_ref,
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    finally:
        script_path.unlink(missing_ok=True)

    if not result.stdout.strip():
        return []

    payload = json.loads(result.stdout)
    if isinstance(payload, dict):
        return [payload]
    return payload


def read_xlsx_range_with_openpyxl(source_file: Path, sheet_name: str, range_ref: str) -> list[dict[str, Any]]:
    workbook_values = load_workbook(source_file, data_only=True, read_only=True)
    workbook_formulas = load_workbook(source_file, data_only=False, read_only=True)
    try:
        worksheet_values = workbook_values[sheet_name]
        worksheet_formulas = workbook_formulas[sheet_name]
        cells: list[dict[str, Any]] = []
        for row in worksheet_values[range_ref]:
            for cell in row:
                formula_cell = worksheet_formulas[cell.coordinate]
                cells.append(
                    {
                        "Row": cell.row,
                        "Column": cell.column,
                        "Address": cell.coordinate,
                        "Text": "" if cell.value is None else str(cell.value),
                        "Value2": cell.value,
                        "Formula": formula_cell.value,
                    }
                )
        return cells
    finally:
        workbook_values.close()
        workbook_formulas.close()


def read_range(source_file: Path, sheet_name: str, range_ref: str) -> list[dict[str, Any]]:
    if source_file.suffix.lower() == ".xls":
        return read_xls_range_with_excel_com(source_file, sheet_name, range_ref)
    return read_xlsx_range_with_openpyxl(source_file, sheet_name, range_ref)


def build_raw_dataframe(cells: list[dict[str, Any]], source_file: Path, sheet_name: str, range_ref: str) -> pd.DataFrame:
    rows = []
    for cell in cells:
        rows.append(
            {
                "source_file": str(source_file),
                "source_sheet": sheet_name,
                "source_range": range_ref,
                "source_row": cell["Row"],
                "source_column": get_column_letter(cell["Column"]),
                "source_cell": cell["Address"],
                "text": clean_text(cell.get("Text")),
                "value": cell.get("Value2"),
                "formula": normalize_formula(cell.get("Formula")),
            }
        )
    return pd.DataFrame(rows)


def build_long_dataframe(raw_df: pd.DataFrame, source_file: Path, sheet_name: str, range_ref: str) -> pd.DataFrame:
    header_df = raw_df[raw_df["source_row"] == 13].copy()
    data_df = raw_df[raw_df["source_row"] > 13].copy()

    header_by_column = {
        row["source_column"]: row["text"]
        for _, row in header_df.iterrows()
    }
    date_by_row: dict[int, dict[str, Any]] = {}
    for _, row in data_df[data_df["source_column"] == "B"].iterrows():
        forecast_date = parse_date(row["text"], row["value"])
        date_by_row[int(row["source_row"])] = {
            "forecast_date": forecast_date,
            "date_text": row["text"],
            "date_source_cell": row["source_cell"],
        }

    value_rows = []
    for _, row in data_df[data_df["source_column"].isin(["C", "D"])].iterrows():
        value = parse_number(row["text"], row["value"])
        if value is None:
            continue

        row_date = date_by_row.get(int(row["source_row"]), {})
        forecast_date = row_date.get("forecast_date")
        value_rows.append(
            {
                "source_file": str(source_file),
                "source_sheet": sheet_name,
                "source_range": range_ref,
                "source_row": int(row["source_row"]),
                "source_cell": row["source_cell"],
                "date_source_cell": row_date.get("date_source_cell"),
                "date_range": row_date.get("date_text"),
                "forecast_date": forecast_date.isoformat() if forecast_date else None,
                "time_period": (
                    f"{MONTH_LABELS[forecast_date.month]} {forecast_date.year}"
                    if forecast_date
                    else None
                ),
                "time_period_year": forecast_date.year if forecast_date else None,
                "time_period_month": forecast_date.month if forecast_date else None,
                "forecast_series": header_by_column.get(row["source_column"]),
                "resin_index_type": infer_resin_index_type(
                    header_by_column.get(row["source_column"])
                ),
                "value_source_column": row["source_column"],
                "value_source_header_cell": f"{row['source_column']}13",
                "value": value,
                "raw_value_text": row["text"],
                "formula": row["formula"],
            }
        )

    return pd.DataFrame(value_rows)


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            if worksheet.max_row == 0:
                continue
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for col_idx in range(1, worksheet.max_column + 1):
                header = worksheet.cell(1, col_idx).value
                width = len(str(header or ""))
                for row_idx in range(2, min(worksheet.max_row, 250) + 1):
                    value = worksheet.cell(row_idx, col_idx).value
                    if value is not None:
                        width = max(width, len(str(value)))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(width + 2, 12),
                    70,
                )
        workbook.save(path)
    finally:
        workbook.close()


def write_outputs(
    excel_output: Path,
    csv_output: Path,
    long_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    metadata: dict[str, Any],
) -> None:
    excel_output.parent.mkdir(parents=True, exist_ok=True)
    csv_output.parent.mkdir(parents=True, exist_ok=True)

    long_df.to_csv(csv_output, index=False, encoding="utf-8-sig")
    metadata_df = pd.DataFrame(
        [{"field": key, "value": value} for key, value in metadata.items()]
    )
    with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
        long_df.to_excel(writer, sheet_name="resin_index_forecast", index=False)
        raw_df.to_excel(writer, sheet_name="raw_B13_D82", index=False)
        metadata_df.to_excel(writer, sheet_name="run_metadata", index=False)
    style_workbook(excel_output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract and flatten the ICIS Resin Index forecast table."
    )
    parser.add_argument("--file", type=Path, default=None)
    parser.add_argument("--forecast-dir", type=Path, default=DEFAULT_FORECAST_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--range", default=DEFAULT_RANGE)
    parser.add_argument("--excel-output", type=Path, default=DEFAULT_EXCEL_OUTPUT)
    parser.add_argument("--csv-output", type=Path, default=DEFAULT_CSV_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.forecast_dir, args.file_pattern)
    cells = read_range(source_file, args.sheet, args.range)
    raw_df = build_raw_dataframe(cells, source_file, args.sheet, args.range)
    long_df = build_long_dataframe(raw_df, source_file, args.sheet, args.range)

    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "source_file": str(source_file),
        "source_sheet": args.sheet,
        "source_range": args.range,
        "raw_cells": len(raw_df),
        "forecast_rows": len(long_df),
        "output_excel": str(args.excel_output),
        "output_csv": str(args.csv_output),
        "notes": "Columns C and D from the source are flattened into one value column with forecast_series preserving the original header.",
    }
    write_outputs(args.excel_output, args.csv_output, long_df, raw_df, metadata)

    print(f"source_file={source_file}")
    print(f"source_sheet={args.sheet}")
    print(f"source_range={args.range}")
    print(f"raw_cells={len(raw_df)}")
    print(f"forecast_rows={len(long_df)}")
    print(f"excel_output={args.excel_output}")
    print(f"csv_output={args.csv_output}")


if __name__ == "__main__":
    main()
