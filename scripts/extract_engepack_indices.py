from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_FILE_PATTERN = "*Engepack*.xlsx"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_INDICES_SHEET = "\u00cdndices"
DEFAULT_TLC_SHEET = "Pre\u00e7os Engepack"
DEFAULT_INDICES_RANGE = "B2:AT5"
DEFAULT_FREIGHT_RANGE = "B23:AT23"
DEFAULT_TLC_RANGE = "B8:AP8"
DEFAULT_RAW_CSV_NAME = "engepack_indices_raw.csv"
DEFAULT_LONG_CSV_NAME = "engepack_indices_long.csv"
FX_METRIC_NAME = "Fx (M-1)"
TLC_METRIC_NAME = "Total Landing Cost"
TLC_USD_METRIC_NAME = "Total Landing Cost $"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_sheet_name(value: str) -> str:
    value = strip_accents(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def find_sheet(workbook: Any, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name

    normalized_target = normalize_sheet_name(sheet_name)
    matches = [
        candidate
        for candidate in workbook.sheetnames
        if normalize_sheet_name(candidate) == normalized_target
    ]
    if len(matches) == 1:
        return matches[0]

    raise ValueError(f"Sheet not found: {sheet_name}")


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def period_fields(period_value: Any) -> dict[str, Any]:
    if isinstance(period_value, datetime):
        period_date = period_value.date()
    elif isinstance(period_value, date):
        period_date = period_value
    else:
        return {
            "time_period": None,
            "time_period_year": None,
            "time_period_month": None,
        }

    return {
        "time_period": f"{MONTH_LABELS[period_date.month]} {period_date.year}",
        "time_period_year": period_date.year,
        "time_period_month": period_date.month,
    }


def raw_rows_for_range(
    worksheet: Any,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    columns = [get_column_letter(col_idx) for col_idx in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(min_row, max_row + 1):
        row = {
            "source_sheet": sheet_name,
            "source_range": range_ref,
            "source_row": row_idx,
        }
        for col_idx, column_letter in zip(range(min_col, max_col + 1), columns):
            row[column_letter] = clean_value(worksheet.cell(row_idx, col_idx).value)
        rows.append(row)

    return rows


def extract_raw_rows(
    indices_worksheet: Any,
    tlc_worksheet: Any,
    indices_sheet_name: str,
    tlc_sheet_name: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rows.extend(raw_rows_for_range(indices_worksheet, indices_sheet_name, DEFAULT_INDICES_RANGE))
    rows.extend(raw_rows_for_range(indices_worksheet, indices_sheet_name, DEFAULT_FREIGHT_RANGE))
    rows.extend(raw_rows_for_range(tlc_worksheet, tlc_sheet_name, DEFAULT_TLC_RANGE))
    return rows


def extract_indices_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(DEFAULT_INDICES_RANGE)
    header_row_idx = min_row
    data_start_col = 4
    rows: list[dict[str, Any]] = []
    section_name: Any = None

    for row_idx in range(min_row + 1, max_row + 1):
        section_name = clean_value(value_worksheet.cell(row_idx, 2).value) or section_name
        metric_name = clean_value(value_worksheet.cell(row_idx, 3).value)
        if metric_name is None:
            continue

        for col_idx in range(data_start_col, max_col + 1):
            value = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            if value is None:
                continue

            column_letter = get_column_letter(col_idx)
            price_period = clean_value(value_worksheet.cell(header_row_idx, col_idx).value)
            rows.append(
                {
                    "source_file": str(source_file),
                    "source_sheet": sheet_name,
                    "source_range": DEFAULT_INDICES_RANGE,
                    "source_cell": f"{column_letter}{row_idx}",
                    "metric_row": row_idx,
                    "metric_label_source_row": row_idx,
                    "period_header_source_sheet": sheet_name,
                    "period_header_source_row": header_row_idx,
                    "section_name": section_name,
                    "supplier": None,
                    "metric_name": metric_name,
                    "metric_detail": None,
                    "price_period": price_period,
                    **period_fields(price_period),
                    "value": value,
                    "formula": formula_worksheet.cell(row_idx, col_idx).value,
                }
            )

    return rows


def extract_freight_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(DEFAULT_FREIGHT_RANGE)
    if min_row != max_row:
        raise ValueError(f"Expected a single-row freight range, got {DEFAULT_FREIGHT_RANGE}.")

    header_row_idx = 20
    data_start_col = max(min_col + 1, 3)
    rows: list[dict[str, Any]] = []

    for col_idx in range(data_start_col, max_col + 1):
        value = clean_value(value_worksheet.cell(min_row, col_idx).value)
        if value is None:
            continue

        column_letter = get_column_letter(col_idx)
        price_period = clean_value(value_worksheet.cell(header_row_idx, col_idx).value)
        rows.append(
            {
                "source_file": str(source_file),
                "source_sheet": sheet_name,
                "source_range": DEFAULT_FREIGHT_RANGE,
                "source_cell": f"{column_letter}{min_row}",
                "metric_row": min_row,
                "metric_label_source_row": min_row,
                "period_header_source_sheet": sheet_name,
                "period_header_source_row": header_row_idx,
                "section_name": "Freight",
                "supplier": "Engepack",
                "metric_name": "Freight",
                "metric_detail": clean_value(value_worksheet.cell(min_row, 2).value),
                "price_period": price_period,
                **period_fields(price_period),
                "value": value,
                "formula": formula_worksheet.cell(min_row, col_idx).value,
            }
        )

    return rows


def tlc_period_col(tlc_col_idx: int) -> int:
    return tlc_col_idx + 4


def extract_tlc_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    indices_worksheet: Any,
    source_file: Path,
    tlc_sheet_name: str,
    indices_sheet_name: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(DEFAULT_TLC_RANGE)
    if min_row != max_row:
        raise ValueError(f"Expected a single-row TLC range, got {DEFAULT_TLC_RANGE}.")

    value_row_idx = min_row
    data_start_col = min_col + 1
    rows: list[dict[str, Any]] = []
    source_label = clean_value(value_worksheet.cell(value_row_idx, min_col).value)

    for col_idx in range(data_start_col, max_col + 1):
        value = clean_value(value_worksheet.cell(value_row_idx, col_idx).value)
        if value is None:
            continue

        period_col_idx = tlc_period_col(col_idx)
        column_letter = get_column_letter(col_idx)
        price_period = clean_value(indices_worksheet.cell(2, period_col_idx).value)
        rows.append(
            {
                "source_file": str(source_file),
                "source_sheet": tlc_sheet_name,
                "source_range": DEFAULT_TLC_RANGE,
                "source_cell": f"{column_letter}{value_row_idx}",
                "metric_row": value_row_idx,
                "metric_label_source_row": value_row_idx,
                "period_header_source_sheet": indices_sheet_name,
                "period_header_source_row": 2,
                "section_name": "Total Landing Cost",
                "supplier": "Engepack",
                "metric_name": "Total Landing Cost",
                "metric_detail": source_label,
                "price_period": price_period,
                **period_fields(price_period),
                "value": value,
                "formula": formula_worksheet.cell(value_row_idx, col_idx).value,
            }
        )

    return rows


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def period_key(row: dict[str, Any]) -> tuple[Any, Any]:
    return row.get("time_period_year"), row.get("time_period_month")


def build_total_landing_cost_usd_rows(
    final_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    fx_by_period: dict[tuple[Any, Any], dict[str, Any]] = {}
    for row in final_rows:
        if row.get("metric_name") == FX_METRIC_NAME:
            key = period_key(row)
            fx_value = numeric_value(row.get("value"))
            if key != (None, None) and fx_value not in (None, 0):
                fx_by_period[key] = row

    calculated_rows: list[dict[str, Any]] = []
    for tlc_row in final_rows:
        if tlc_row.get("metric_name") != TLC_METRIC_NAME:
            continue

        key = period_key(tlc_row)
        fx_row = fx_by_period.get(key)
        tlc_value = numeric_value(tlc_row.get("value"))
        fx_value = numeric_value(fx_row.get("value")) if fx_row else None
        if tlc_value is None or fx_value in (None, 0):
            continue

        calculated_row = dict(tlc_row)
        calculated_row.update(
            {
                "source_sheet": "Calculated KPI",
                "source_range": None,
                "source_cell": None,
                "metric_row": None,
                "metric_label_source_row": None,
                "period_header_source_sheet": None,
                "period_header_source_row": None,
                "section_name": TLC_USD_METRIC_NAME,
                "metric_name": TLC_USD_METRIC_NAME,
                "metric_detail": f"{TLC_METRIC_NAME} / {FX_METRIC_NAME}",
                "value": tlc_value / fx_value,
                "formula": (
                    f"{TLC_METRIC_NAME} ({tlc_value}) / "
                    f"{FX_METRIC_NAME} ({fx_value})"
                ),
            }
        )
        calculated_rows.append(calculated_row)

    return calculated_rows


def extract_long_rows(
    indices_value_worksheet: Any,
    indices_formula_worksheet: Any,
    tlc_value_worksheet: Any,
    tlc_formula_worksheet: Any,
    source_file: Path,
    indices_sheet_name: str,
    tlc_sheet_name: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rows.extend(
        extract_indices_rows(
            indices_value_worksheet,
            indices_formula_worksheet,
            source_file,
            indices_sheet_name,
        )
    )
    rows.extend(
        extract_freight_rows(
            indices_value_worksheet,
            indices_formula_worksheet,
            source_file,
            indices_sheet_name,
        )
    )
    rows.extend(
        extract_tlc_rows(
            tlc_value_worksheet,
            tlc_formula_worksheet,
            indices_value_worksheet,
            source_file,
            tlc_sheet_name,
            indices_sheet_name,
        )
    )
    return rows


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=ordered_headers(rows))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_indices_final.xlsx"


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        return

    headers = ordered_headers(rows)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(row.get(header) for row in rows[:100])
        width = min(
            max(len(str(value)) for value in values if value is not None) + 2,
            45,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for key, value in metadata.items():
        worksheet.append([key, value])

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 80


def write_excel_output(
    path: Path,
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, "raw_indices_freight_tlc", raw_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Engepack indices, freight, and TLC extraction pipeline."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Specific workbook to process. If omitted, the latest Engepack workbook in data is used.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--indices-sheet", default=DEFAULT_INDICES_SHEET)
    parser.add_argument("--tlc-sheet", default=DEFAULT_TLC_SHEET)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=None,
        help="Optional explicit Excel output path.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files into the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)
    excel_output = args.excel_output or default_excel_output_path(
        source_file,
        args.output_dir,
    )

    values_workbook = load_workbook(source_file, read_only=True, data_only=True)
    formulas_workbook = load_workbook(source_file, read_only=True, data_only=False)

    try:
        indices_sheet_name = find_sheet(values_workbook, args.indices_sheet)
        tlc_sheet_name = find_sheet(values_workbook, args.tlc_sheet)
        indices_value_worksheet = values_workbook[indices_sheet_name]
        indices_formula_worksheet = formulas_workbook[indices_sheet_name]
        tlc_value_worksheet = values_workbook[tlc_sheet_name]
        tlc_formula_worksheet = formulas_workbook[tlc_sheet_name]

        raw_rows = extract_raw_rows(
            indices_value_worksheet,
            tlc_value_worksheet,
            indices_sheet_name,
            tlc_sheet_name,
        )
        final_rows = extract_long_rows(
            indices_value_worksheet,
            indices_formula_worksheet,
            tlc_value_worksheet,
            tlc_formula_worksheet,
            source_file,
            indices_sheet_name,
            tlc_sheet_name,
        )
        tlc_usd_rows = build_total_landing_cost_usd_rows(final_rows)
        final_rows.extend(tlc_usd_rows)
        freight_rows = [
            row for row in final_rows if row.get("metric_name") == "Freight"
        ]
        tlc_rows = [
            row for row in final_rows if row.get("metric_name") == TLC_METRIC_NAME
        ]

        metadata = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(source_file),
            "indices_source_sheet": indices_sheet_name,
            "tlc_source_sheet": tlc_sheet_name,
            "indices_source_range": DEFAULT_INDICES_RANGE,
            "freight_source_range": DEFAULT_FREIGHT_RANGE,
            "tlc_source_range": DEFAULT_TLC_RANGE,
            "freight_field_name": "Freight",
            "tlc_field_name": "Total Landing Cost",
            "raw_rows": len(raw_rows),
            "final_rows": len(final_rows),
            "freight_rows": len(freight_rows),
            "tlc_rows": len(tlc_rows),
            "tlc_usd_rows": len(tlc_usd_rows),
            "tlc_period_mapping": "Precos Engepack C8:AP8 maps to Indices G2:AT2",
        }

        write_excel_output(excel_output, raw_rows, final_rows, metadata)

        if args.write_csv:
            write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
            write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

        print(f"source_file={source_file}")
        print(f"indices_source_sheet={indices_sheet_name}")
        print(f"tlc_source_sheet={tlc_sheet_name}")
        print(f"indices_source_range={DEFAULT_INDICES_RANGE}")
        print(f"freight_source_range={DEFAULT_FREIGHT_RANGE}")
        print(f"tlc_source_range={DEFAULT_TLC_RANGE}")
        print(f"raw_rows={len(raw_rows)}")
        print(f"final_rows={len(final_rows)}")
        print(f"freight_rows={len(freight_rows)}")
        print(f"tlc_rows={len(tlc_rows)}")
        print(f"tlc_usd_rows={len(tlc_usd_rows)}")
        print(f"excel_output={excel_output}")
        if args.write_csv:
            print(f"raw_csv_output={args.output_dir / DEFAULT_RAW_CSV_NAME}")
            print(f"final_csv_output={args.output_dir / DEFAULT_LONG_CSV_NAME}")
    finally:
        values_workbook.close()
        formulas_workbook.close()


if __name__ == "__main__":
    main()
