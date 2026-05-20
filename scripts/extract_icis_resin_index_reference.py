from __future__ import annotations

import argparse
import json
import re
import subprocess
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INDEX_DIR = Path("Index Fprecast")
DEFAULT_FILE_PATTERN = "*2026-05-04 152801_v1*.xls"
DEFAULT_SHEET = "ICIS Price History"
DEFAULT_RANGE = "B13:D46"
DEFAULT_EXCEL_OUTPUT = DEFAULT_INDEX_DIR / "icis_resin_index_reference_table.xlsx"
DEFAULT_CSV_OUTPUT = DEFAULT_INDEX_DIR / "icis_resin_index_reference_table.csv"
DEFAULT_WIDE_CSV_OUTPUT = DEFAULT_INDEX_DIR / "icis_resin_index_reference_wide.csv"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def resolve_source_file(
    explicit_file: Path | None,
    index_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Index forecast file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in index_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No index forecast files found in {index_dir} matching {file_pattern}"
        )
    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def excel_serial_to_date(value: float) -> date:
    return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()


def clean_text(value: Any) -> Any:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def parse_period(text: Any, value2: Any) -> date | None:
    if isinstance(value2, (int, float)) and value2 > 20000:
        return excel_serial_to_date(float(value2)).replace(day=1)

    cleaned = clean_text(text)
    if cleaned is None:
        return None

    for fmt in ("%b-%Y", "%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(cleaned, fmt).date().replace(day=1)
        except ValueError:
            continue
    return None


def parse_number(text: Any, value2: Any) -> float | None:
    if isinstance(value2, bool) or value2 is None:
        return None
    if isinstance(value2, (int, float)):
        return float(value2)

    cleaned = clean_text(text)
    if cleaned is None:
        return None
    try:
        return float(str(cleaned).replace(",", ""))
    except ValueError:
        return None


def normalize_formula(value: Any) -> Any:
    text = clean_text(value)
    if text in (None, ""):
        return None
    return text


def infer_resin_index_type(series: Any) -> str | None:
    text = clean_text(series)
    if not isinstance(text, str):
        return None
    normalized = text.lower()
    if "weekly (low)" in normalized or "(low)" in normalized:
        return "ICIS Asia SE Low"
    if "weekly (mid)" in normalized or "(mid)" in normalized:
        return "ICIS China Mid"
    return re.sub(r"\s*:\s*usd/tonne\s*$", "", text, flags=re.IGNORECASE)


def infer_resin_index_level(series: Any) -> str | None:
    text = clean_text(series)
    if not isinstance(text, str):
        return None
    normalized = text.lower()
    if "(low)" in normalized:
        return "Low"
    if "(mid)" in normalized:
        return "Mid"
    return None


def read_xls_range_with_excel_com(
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    powershell = r"""
param(
  [Parameter(Mandatory=$true)][string]$Path,
  [Parameter(Mandatory=$true)][string]$SheetName,
  [Parameter(Mandatory=$true)][string]$RangeRef
)

$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false

try {
  $workbook = $excel.Workbooks.Open($Path, 0, $true)
  $worksheet = $workbook.Worksheets.Item($SheetName)
  $range = $worksheet.Range($RangeRef)
  $rows = @()

  foreach ($cell in $range.Cells) {
    $rows += [pscustomobject]@{
      Row = [int]$cell.Row
      Column = [int]$cell.Column
      Address = [string]$cell.Address($false, $false)
      Text = [string]$cell.Text
      Value2 = $cell.Value2
      Formula = [string]$cell.Formula
    }
  }

  $rows | ConvertTo-Json -Depth 4 -Compress
  $workbook.Close($false)
}
finally {
  if ($workbook) {
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
  }
  if ($worksheet) {
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($worksheet) | Out-Null
  }
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}
"""
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as script:
        script.write(powershell)
        script_path = Path(script.name)

    try:
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-Path",
                str(source_file.resolve()),
                "-SheetName",
                sheet_name,
                "-RangeRef",
                range_ref,
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    finally:
        script_path.unlink(missing_ok=True)

    if not result.stdout.strip():
        return []

    payload = json.loads(result.stdout)
    if isinstance(payload, dict):
        return [payload]
    return payload


def build_raw_dataframe(
    cells: list[dict[str, Any]],
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> pd.DataFrame:
    rows = []
    for cell in cells:
        rows.append(
            {
                "source_file": str(source_file),
                "source_sheet": sheet_name,
                "source_range": range_ref,
                "source_row": cell["Row"],
                "source_column": get_column_letter(cell["Column"]),
                "source_cell": cell["Address"],
                "text": clean_text(cell.get("Text")),
                "value": cell.get("Value2"),
                "formula": normalize_formula(cell.get("Formula")),
            }
        )
    return pd.DataFrame(rows)


def build_long_dataframe(
    raw_df: pd.DataFrame,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> pd.DataFrame:
    header_df = raw_df[raw_df["source_row"] == 13].copy()
    data_df = raw_df[raw_df["source_row"] > 13].copy()

    header_by_column = {
        row["source_column"]: row["text"]
        for _, row in header_df.iterrows()
    }
    date_by_row: dict[int, dict[str, Any]] = {}
    for _, row in data_df[data_df["source_column"] == "B"].iterrows():
        period = parse_period(row["text"], row["value"])
        date_by_row[int(row["source_row"])] = {
            "period": period,
            "date_range": row["text"],
            "date_source_cell": row["source_cell"],
        }

    value_rows = []
    for _, row in data_df[data_df["source_column"].isin(["C", "D"])].iterrows():
        value = parse_number(row["text"], row["value"])
        if value is None:
            continue

        row_date = date_by_row.get(int(row["source_row"]), {})
        period = row_date.get("period")
        series = header_by_column.get(row["source_column"])
        value_rows.append(
            {
                "source_file": str(source_file),
                "source_sheet": sheet_name,
                "source_range": range_ref,
                "source_row": int(row["source_row"]),
                "source_cell": row["source_cell"],
                "date_source_cell": row_date.get("date_source_cell"),
                "date_range": row_date.get("date_range"),
                "forecast_date": period.isoformat() if period else None,
                "time_period": f"{MONTH_LABELS[period.month]} {period.year}" if period else None,
                "time_period_year": period.year if period else None,
                "time_period_month": period.month if period else None,
                "forecast_series": series,
                "resin_index_type": infer_resin_index_type(series),
                "resin_index_level": infer_resin_index_level(series),
                "value_source_column": row["source_column"],
                "value_source_header_cell": f"{row['source_column']}13",
                "value": value,
                "raw_value_text": row["text"],
                "formula": row["formula"],
            }
        )

    return pd.DataFrame(value_rows)


def build_wide_dataframe(long_df: pd.DataFrame) -> pd.DataFrame:
    if long_df.empty:
        return pd.DataFrame()

    key_columns = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "forecast_date",
        "date_range",
    ]
    wide = (
        long_df.pivot_table(
            index=key_columns,
            columns="resin_index_type",
            values="value",
            aggfunc="first",
        )
        .reset_index()
        .sort_values(["time_period_year", "time_period_month"])
    )
    wide.columns.name = None
    return wide


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            if worksheet.max_row == 0:
                continue
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for col_idx in range(1, worksheet.max_column + 1):
                header = worksheet.cell(1, col_idx).value
                width = len(str(header or ""))
                for row_idx in range(2, min(worksheet.max_row, 250) + 1):
                    value = worksheet.cell(row_idx, col_idx).value
                    if value is not None:
                        width = max(width, len(str(value)))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(width + 2, 12),
                    70,
                )
        workbook.save(path)
    finally:
        workbook.close()


def write_outputs(
    excel_output: Path,
    csv_output: Path,
    wide_csv_output: Path,
    long_df: pd.DataFrame,
    wide_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    metadata: dict[str, Any],
) -> None:
    excel_output.parent.mkdir(parents=True, exist_ok=True)
    csv_output.parent.mkdir(parents=True, exist_ok=True)
    wide_csv_output.parent.mkdir(parents=True, exist_ok=True)

    long_df.to_csv(csv_output, index=False, encoding="utf-8-sig")
    wide_df.to_csv(wide_csv_output, index=False, encoding="utf-8-sig")
    metadata_df = pd.DataFrame(
        [{"field": key, "value": value} for key, value in metadata.items()]
    )
    with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
        long_df.to_excel(writer, sheet_name="icis_resin_reference", index=False)
        wide_df.to_excel(writer, sheet_name="monthly_wide_reference", index=False)
        raw_df.to_excel(writer, sheet_name="raw_B13_D46", index=False)
        metadata_df.to_excel(writer, sheet_name="run_metadata", index=False)
    style_workbook(excel_output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract the combined ICIS Low and ICIS Mid resin index reference table."
    )
    parser.add_argument("--file", type=Path, default=None)
    parser.add_argument("--index-dir", type=Path, default=DEFAULT_INDEX_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--range", default=DEFAULT_RANGE)
    parser.add_argument("--excel-output", type=Path, default=DEFAULT_EXCEL_OUTPUT)
    parser.add_argument("--csv-output", type=Path, default=DEFAULT_CSV_OUTPUT)
    parser.add_argument("--wide-csv-output", type=Path, default=DEFAULT_WIDE_CSV_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.index_dir, args.file_pattern)
    cells = read_xls_range_with_excel_com(source_file, args.sheet, args.range)
    raw_df = build_raw_dataframe(cells, source_file, args.sheet, args.range)
    long_df = build_long_dataframe(raw_df, source_file, args.sheet, args.range)
    wide_df = build_wide_dataframe(long_df)

    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "source_file": str(source_file),
        "source_sheet": args.sheet,
        "source_range": args.range,
        "requested_mid_range_note": (
            "The requested B13:C46 range contains ICIS Low in column C. "
            "ICIS Mid is in column D, so B13:D46 is extracted."
        ),
        "raw_cells": len(raw_df),
        "reference_rows": len(long_df),
        "wide_rows": len(wide_df),
        "resin_index_types": ", ".join(
            sorted(str(value) for value in long_df["resin_index_type"].dropna().unique())
        ),
        "output_excel": str(args.excel_output),
        "output_csv": str(args.csv_output),
        "output_wide_csv": str(args.wide_csv_output),
    }
    write_outputs(
        args.excel_output,
        args.csv_output,
        args.wide_csv_output,
        long_df,
        wide_df,
        raw_df,
        metadata,
    )

    print(f"source_file={source_file}")
    print(f"source_sheet={args.sheet}")
    print(f"source_range={args.range}")
    print(f"raw_cells={len(raw_df)}")
    print(f"reference_rows={len(long_df)}")
    print(f"wide_rows={len(wide_df)}")
    print(f"resin_index_types={metadata['resin_index_types']}")
    print(f"excel_output={args.excel_output}")
    print(f"csv_output={args.csv_output}")
    print(f"wide_csv_output={args.wide_csv_output}")


if __name__ == "__main__":
    main()
