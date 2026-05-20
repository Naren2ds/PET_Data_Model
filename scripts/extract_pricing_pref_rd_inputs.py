from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_FILE_PATTERN = "Pricing_Pref_ABI_RD*Cliente*.xlsx"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_SHEET = "Inputs"
DEFAULT_RANGE = "G8:I12"
DEFAULT_RAW_CSV_NAME = "pricing_pref_rd_inputs_raw.csv"
DEFAULT_LONG_CSV_NAME = "pricing_pref_rd_inputs_long.csv"
DEFAULT_FALLBACK_PERIOD = date(2026, 4, 1)

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

MONTH_LOOKUP = {
    "JAN": 1,
    "JANUARY": 1,
    "ENERO": 1,
    "ENE": 1,
    "FEV": 2,
    "FEB": 2,
    "FEBRUARY": 2,
    "FEBRERO": 2,
    "MAR": 3,
    "MARCH": 3,
    "MARZO": 3,
    "ABR": 4,
    "APR": 4,
    "APRIL": 4,
    "ABRIL": 4,
    "MAY": 5,
    "MAYO": 5,
    "JUN": 6,
    "JUNE": 6,
    "JUNIO": 6,
    "JUL": 7,
    "JULY": 7,
    "JULIO": 7,
    "AGO": 8,
    "AUG": 8,
    "AUGUST": 8,
    "AGOSTO": 8,
    "SET": 9,
    "SEP": 9,
    "SEPT": 9,
    "SEPTEMBER": 9,
    "SEPTIEMBRE": 9,
    "OUT": 10,
    "OCT": 10,
    "OCTOBER": 10,
    "OCTUBRE": 10,
    "NOV": 11,
    "NOVEMBER": 11,
    "NOVIEMBRE": 11,
    "DEZ": 12,
    "DEC": 12,
    "DECEMBER": 12,
    "DICIEMBRE": 12,
}

TRANSLATIONS = {
    "PRECIO FOB": "FOB Price",
    "FLETE MARITIMO": "Ocean Freight",
    "NACIONALIZACION": "Import Clearance",
    "NACIONALIZACION %": "Import Clearance (%)",
    "PRECIO DDP": "DDP Price",
}

COUNTRY_TRANSLATIONS = {
    "EL SALVADOR": "El Salvador",
    "PERU": "Peru",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_text(value: str) -> str:
    value = strip_accents(value).upper()
    return re.sub(r"[^A-Z0-9]+", " ", value).strip()


def translate_label(value: Any) -> Any:
    if not isinstance(value, str):
        return None

    normalized = normalize_text(value)
    if normalized == "NACIONALIZACION" and "%" in value:
        return "Import Clearance (%)"

    return TRANSLATIONS.get(normalized, value.strip())


def translate_country(value: Any) -> Any:
    if not isinstance(value, str):
        return value

    normalized = normalize_text(value)
    return COUNTRY_TRANSLATIONS.get(normalized, value.strip())


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def strip_sheet_name(value: str) -> str:
    return re.sub(r"\s+", " ", strip_accents(value).lower()).strip()


def find_sheet(workbook: Any, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name

    normalized_target = strip_sheet_name(sheet_name)
    matches = [
        candidate
        for candidate in workbook.sheetnames
        if strip_sheet_name(candidate) == normalized_target
    ]
    if len(matches) == 1:
        return matches[0]

    raise ValueError(f"Sheet not found: {sheet_name}")


def parse_month(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.month
    if isinstance(value, date):
        return value.month
    if isinstance(value, (int, float)) and int(value) == value and 1 <= int(value) <= 12:
        return int(value)
    if isinstance(value, str):
        cleaned = normalize_text(value)
        if cleaned.isdigit() and 1 <= int(cleaned) <= 12:
            return int(cleaned)
        return MONTH_LOOKUP.get(cleaned[:3], MONTH_LOOKUP.get(cleaned))
    return None


def parse_year(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    if isinstance(value, (int, float)) and int(value) == value:
        year = int(value)
        return year + 2000 if year < 100 else year
    if isinstance(value, str):
        match = re.search(r"\d{2,4}", value)
        if match:
            year = int(match.group(0))
            return year + 2000 if year < 100 else year
    return None


def parse_period_argument(value: str) -> date:
    value = value.strip()
    match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", value)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)

    match = re.fullmatch(r"([A-Za-z]+)\s+(\d{2,4})", value)
    if match:
        month = parse_month(match.group(1))
        year = parse_year(match.group(2))
        if month and year:
            return date(year, month, 1)

    raise ValueError(
        "Period must be formatted as YYYY-MM or Month YYYY, for example 2026-04."
    )


def period_from_sheet(worksheet: Any) -> date | None:
    year = parse_year(worksheet["C6"].value)
    month = parse_month(worksheet["D6"].value)
    if year and month:
        return date(year, month, 1)
    return None


def period_fields(period_date: date) -> dict[str, Any]:
    return {
        "time_period": f"{MONTH_LABELS[period_date.month]} {period_date.year}",
        "time_period_year": period_date.year,
        "time_period_month": period_date.month,
    }


def extract_raw_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    range_ref: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    rows: list[dict[str, Any]] = []

    for row_idx in range(min_row, max_row + 1):
        row = {
            "source_range": range_ref,
            "source_row": row_idx,
        }
        for col_idx in range(min_col, max_col + 1):
            column_letter = get_column_letter(col_idx)
            row[column_letter] = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            row[f"{column_letter}_formula"] = clean_value(
                formula_worksheet.cell(row_idx, col_idx).value
            )
        rows.append(row)

    return rows


def extract_final_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
    period_date: date,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    label_col = min_col
    header_row = min_row
    first_metric_row = min_row + 1
    label_letter = get_column_letter(label_col)
    rows: list[dict[str, Any]] = []

    for value_col in range(label_col + 1, max_col + 1):
        value_letter = get_column_letter(value_col)
        country_original = clean_value(value_worksheet.cell(header_row, value_col).value)
        if country_original is None:
            continue

        for row_idx in range(first_metric_row, max_row + 1):
            label_original = clean_value(value_worksheet.cell(row_idx, label_col).value)
            value = clean_value(value_worksheet.cell(row_idx, value_col).value)
            if label_original is None and value is None:
                continue

            rows.append(
                {
                    "source_file": str(source_file),
                    "source_sheet": sheet_name,
                    "source_range": range_ref,
                    "source_cell": f"{value_letter}{row_idx}",
                    "metric_row": row_idx,
                    **period_fields(period_date),
                    "country_source_cell": f"{value_letter}{header_row}",
                    "country_original": country_original,
                    "country": translate_country(country_original),
                    "metric_label_source_cell": f"{label_letter}{row_idx}",
                    "metric_label_original": label_original,
                    "metric_label_english": translate_label(label_original),
                    "value": value,
                    "formula": formula_worksheet.cell(row_idx, value_col).value,
                }
            )

    return rows


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=ordered_headers(rows))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_inputs_final.xlsx"


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        worksheet.append(["No rows extracted"])
        return

    headers = ordered_headers(rows)
    worksheet.append(headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = worksheet.cell(row_idx, col_idx, value=value)
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_idx in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_length + 2, 12),
            45,
        )


def write_metadata_sheet(
    workbook: Workbook,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
    period_date: date,
    row_count: int,
) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    metadata = [
        ("source_file", str(source_file)),
        ("source_sheet", sheet_name),
        ("extraction_range", range_ref),
        ("period_source", "Inputs!C6:D6, unless --time-period is provided"),
        ("time_period", period_fields(period_date)["time_period"]),
        ("time_period_year", period_date.year),
        ("time_period_month", period_date.month),
        ("final_row_count", row_count),
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
    ]
    worksheet.append(["field", "value"])
    for item in metadata:
        worksheet.append(item)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 65


def write_excel_output(
    path: Path,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
    final_rows: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
    period_date: date,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, "raw_G8_I12", raw_rows)
    write_metadata_sheet(
        workbook=workbook,
        source_file=source_file,
        sheet_name=sheet_name,
        range_ref=range_ref,
        period_date=period_date,
        row_count=len(final_rows),
    )
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract Pricing Pref RD Inputs G8:I12 into the PET data model."
    )
    parser.add_argument("--file", type=Path, help="Specific source workbook to process.")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--range", dest="range_ref", default=DEFAULT_RANGE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--excel-output", type=Path)
    parser.add_argument(
        "--time-period",
        help="Override period as YYYY-MM or Month YYYY. Defaults to Inputs!C6:D6.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files to the output directory.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)

    value_workbook = load_workbook(source_file, data_only=True, read_only=False)
    formula_workbook = load_workbook(source_file, data_only=False, read_only=False)
    sheet_name = find_sheet(value_workbook, args.sheet)
    value_worksheet = value_workbook[sheet_name]
    formula_worksheet = formula_workbook[sheet_name]

    period_date = (
        parse_period_argument(args.time_period)
        if args.time_period
        else period_from_sheet(value_worksheet) or DEFAULT_FALLBACK_PERIOD
    )
    raw_rows = extract_raw_rows(value_worksheet, formula_worksheet, args.range_ref)
    final_rows = extract_final_rows(
        value_worksheet=value_worksheet,
        formula_worksheet=formula_worksheet,
        source_file=source_file,
        sheet_name=sheet_name,
        range_ref=args.range_ref,
        period_date=period_date,
    )

    excel_output = args.excel_output or default_excel_output_path(
        source_file, args.output_dir
    )
    write_excel_output(
        path=excel_output,
        source_file=source_file,
        sheet_name=sheet_name,
        range_ref=args.range_ref,
        final_rows=final_rows,
        raw_rows=raw_rows,
        period_date=period_date,
    )

    if args.write_csv:
        write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
        write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

    print(f"Source file: {source_file}")
    print(f"Sheet: {sheet_name}")
    print(f"Extracted rows: {len(final_rows)}")
    print(f"Time period: {period_fields(period_date)['time_period']}")
    print(f"Excel output: {excel_output}")


if __name__ == "__main__":
    main()
