from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_FILE_PATTERN = "*Amcor*.xlsx"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_SHEET = "Resina"
DEFAULT_RANGES = ("B2:O12", "B29:N36")
DEFAULT_RAW_CSV_NAME = "amcor_resina_raw.csv"
DEFAULT_LONG_CSV_NAME = "amcor_resina_long.csv"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def select_sheet(workbook: Any, explicit_sheet: str | None) -> str:
    sheet_name = explicit_sheet or DEFAULT_SHEET
    if sheet_name in workbook.sheetnames:
        return sheet_name

    normalized_target = normalize_sheet_name(sheet_name)
    matches = [
        candidate
        for candidate in workbook.sheetnames
        if normalize_sheet_name(candidate) == normalized_target
    ]
    if len(matches) == 1:
        return matches[0]

    raise ValueError(f"Sheet not found: {sheet_name}")


def normalize_sheet_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def is_period_value(value: Any) -> bool:
    return isinstance(value, (datetime, date))


def period_fields(period_value: Any) -> dict[str, Any]:
    if isinstance(period_value, datetime):
        period_date = period_value.date()
    elif isinstance(period_value, date):
        period_date = period_value
    else:
        return {
            "time_period": None,
            "time_period_year": None,
            "time_period_month": None,
        }

    return {
        "time_period": f"{MONTH_LABELS[period_date.month]} {period_date.year}",
        "time_period_year": period_date.year,
        "time_period_month": period_date.month,
    }


def first_label(*values: Any) -> Any:
    for value in values:
        cleaned = clean_value(value)
        if cleaned is not None:
            return cleaned
    return None


def find_data_start_col(worksheet: Any, min_col: int, max_col: int, row_idx: int) -> int:
    for col_idx in range(min_col, max_col + 1):
        if is_period_value(worksheet.cell(row_idx, col_idx).value):
            return col_idx

    raise ValueError(f"No period/date columns found in row {row_idx}.")


def is_header_row(worksheet: Any, min_col: int, max_col: int, row_idx: int) -> bool:
    period_cells = sum(
        1
        for col_idx in range(min_col, max_col + 1)
        if is_period_value(worksheet.cell(row_idx, col_idx).value)
    )
    return period_cells >= 2


def extract_raw_rows(worksheet: Any, range_refs: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for range_ref in range_refs:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        columns = [get_column_letter(col_idx) for col_idx in range(min_col, max_col + 1)]

        for row_idx in range(min_row, max_row + 1):
            row = {"source_range": range_ref, "source_row": row_idx}
            for col_idx, column_letter in zip(range(min_col, max_col + 1), columns):
                row[column_letter] = clean_value(worksheet.cell(row_idx, col_idx).value)
            rows.append(row)

    return rows


def extract_long_rows_for_range(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    rows: list[dict[str, Any]] = []
    header_row_idx: int | None = None
    data_start_col: int | None = None
    section_name: Any = None

    for row_idx in range(min_row, max_row + 1):
        if is_header_row(value_worksheet, min_col, max_col, row_idx):
            header_row_idx = row_idx
            data_start_col = find_data_start_col(
                value_worksheet,
                min_col,
                max_col,
                row_idx,
            )
            section_name = first_label(
                value_worksheet.cell(row_idx, 3).value,
                value_worksheet.cell(row_idx, 2).value,
                value_worksheet.cell(row_idx, 4).value,
            )
            continue

        if header_row_idx is None or data_start_col is None:
            continue

        supplier = clean_value(value_worksheet.cell(row_idx, 2).value)
        location = clean_value(value_worksheet.cell(row_idx, 3).value)
        metric_name = clean_value(value_worksheet.cell(row_idx, 4).value)
        if metric_name is None:
            continue

        for col_idx in range(data_start_col, max_col + 1):
            value = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            if value is None:
                continue

            column_letter = get_column_letter(col_idx)
            price_period = clean_value(value_worksheet.cell(header_row_idx, col_idx).value)
            rows.append(
                {
                    "source_file": str(source_file),
                    "source_sheet": sheet_name,
                    "source_range": range_ref,
                    "source_cell": f"{column_letter}{row_idx}",
                    "metric_row": row_idx,
                    "metric_label_source_row": row_idx,
                    "period_header_source_row": header_row_idx,
                    "section_name": section_name,
                    "supplier": supplier,
                    "location": location,
                    "metric_name": metric_name,
                    "price_period": price_period,
                    **period_fields(price_period),
                    "value": value,
                    "formula": formula_worksheet.cell(row_idx, col_idx).value,
                }
            )

    return rows


def extract_long_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    range_refs: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for range_ref in range_refs:
        rows.extend(
            extract_long_rows_for_range(
                value_worksheet,
                formula_worksheet,
                source_file,
                sheet_name,
                range_ref,
            )
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    headers = ordered_headers(rows)
    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_resina_final.xlsx"


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        return

    headers = ordered_headers(rows)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(row.get(header) for row in rows[:100])
        width = min(
            max(len(str(value)) for value in values if value is not None) + 2,
            45,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for key, value in metadata.items():
        worksheet.append([key, value])

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 70


def write_excel_output(
    path: Path,
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, "raw_B2_O12_B29_N36", raw_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Amcor Resina extraction pipeline."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Specific workbook to process. If omitted, the latest Amcor workbook in data is used.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--ranges", nargs="+", default=list(DEFAULT_RANGES))
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=None,
        help="Optional explicit Excel output path.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files into the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)
    range_refs = list(args.ranges)
    excel_output = args.excel_output or default_excel_output_path(
        source_file,
        args.output_dir,
    )

    values_workbook = load_workbook(source_file, read_only=True, data_only=True)
    formulas_workbook = load_workbook(source_file, read_only=True, data_only=False)

    try:
        sheet_name = select_sheet(values_workbook, args.sheet)
        value_worksheet = values_workbook[sheet_name]
        formula_worksheet = formulas_workbook[sheet_name]

        raw_rows = extract_raw_rows(value_worksheet, range_refs)
        final_rows = extract_long_rows(
            value_worksheet,
            formula_worksheet,
            source_file,
            sheet_name,
            range_refs,
        )

        metadata = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(source_file),
            "source_sheet": sheet_name,
            "source_ranges": "; ".join(range_refs),
            "raw_rows": len(raw_rows),
            "final_rows": len(final_rows),
        }
        for idx, range_ref in enumerate(range_refs, start=1):
            metadata[f"source_range_{idx}"] = range_ref

        write_excel_output(excel_output, raw_rows, final_rows, metadata)

        if args.write_csv:
            write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
            write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

        print(f"source_file={source_file}")
        print(f"source_sheet={sheet_name}")
        print(f"source_ranges={'; '.join(range_refs)}")
        print(f"raw_rows={len(raw_rows)}")
        print(f"final_rows={len(final_rows)}")
        print(f"excel_output={excel_output}")
        if args.write_csv:
            print(f"raw_csv_output={args.output_dir / DEFAULT_RAW_CSV_NAME}")
            print(f"final_csv_output={args.output_dir / DEFAULT_LONG_CSV_NAME}")
    finally:
        values_workbook.close()
        formulas_workbook.close()


if __name__ == "__main__":
    main()
