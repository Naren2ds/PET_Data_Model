from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_FILE_PATTERN = "*Valgroup*.xlsx"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_RANGE = "U2:V12"
DEFAULT_RAW_CSV_NAME = "valgroup_months_raw.csv"
DEFAULT_LONG_CSV_NAME = "valgroup_months_long.csv"

MONTHS = {
    "JAN": (1, "January"),
    "FEV": (2, "February"),
    "FEB": (2, "February"),
    "MAR": (3, "March"),
    "ABR": (4, "April"),
    "APR": (4, "April"),
    "MAI": (5, "May"),
    "MAY": (5, "May"),
    "JUN": (6, "June"),
    "JUL": (7, "July"),
    "AGO": (8, "August"),
    "AUG": (8, "August"),
    "SET": (9, "September"),
    "SEP": (9, "September"),
    "OUT": (10, "October"),
    "OCT": (10, "October"),
    "NOV": (11, "November"),
    "DEZ": (12, "December"),
    "DEC": (12, "December"),
}

TRANSLATIONS = {
    "RESINA VIRGEM": "Virgin Resin",
    "Icis Asia SE Low (n-1)": "ICIS Asia SE Low (n-1)",
    "Icis Asia 5R MID (n-1)": "ICIS Asia 5R MID (n-1)",
    "Resina c/ premissas": "Resin with assumptions",
    "Desconto": "Discount",
    "Drewry (t-1) com desconto": "Drewry (t-1) with discount",
    "Internação": "Importation",
    "Imposto Internação": "Import Tax",
    "Sobretaxa": "Surcharge",
    "Desconto Indorama": "Indorama Discount",
    "Total V-PET USD/ton ": "Total V-PET USD/ton",
    "Total V-PET R$/ton ": "Total V-PET BRL/ton",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_text(value: str) -> str:
    value = strip_accents(value).upper()
    return re.sub(r"[^A-Z0-9]+", " ", value).strip()


def translate(value: Any) -> Any:
    if not isinstance(value, str):
        return None

    normalized = normalize_text(value)
    for original, translated in TRANSLATIONS.items():
        if normalize_text(original) == normalized:
            return translated

    return value.strip()


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def parse_month_sheet(sheet_name: str) -> dict[str, Any] | None:
    match = re.fullmatch(r"\s*([A-Za-z]{3})\.?(\d{2}|\d{4})\s*", sheet_name)
    if not match:
        return None

    month_key = match.group(1).upper()
    if month_key not in MONTHS:
        return None

    year = int(match.group(2))
    if year < 100:
        year += 2000

    month_number, month_english = MONTHS[month_key]
    return {
        "month_original": match.group(1).upper(),
        "month_english": month_english,
        "time_period": f"{month_english} {year}",
        "time_period_year": year,
        "time_period_month": month_number,
    }


def select_month_sheets(workbook: Any, explicit_sheets: list[str] | None) -> list[str]:
    if explicit_sheets:
        missing = [sheet for sheet in explicit_sheets if sheet not in workbook.sheetnames]
        if missing:
            raise ValueError(f"Sheet(s) not found: {missing}")
        return explicit_sheets

    month_sheets = [
        sheet_name for sheet_name in workbook.sheetnames if parse_month_sheet(sheet_name)
    ]
    if not month_sheets:
        raise ValueError("No month-named sheets were found.")

    return month_sheets


def extract_raw_rows(
    worksheet: Any,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    columns = [get_column_letter(col_idx) for col_idx in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(min_row, max_row + 1):
        row = {
            "source_sheet": sheet_name,
            "source_range": range_ref,
            "source_row": row_idx,
        }
        for col_idx, column_letter in zip(range(min_col, max_col + 1), columns):
            row[column_letter] = clean_value(worksheet.cell(row_idx, col_idx).value)
        rows.append(row)

    return rows


def is_label_value_layout(worksheet: Any) -> bool:
    return normalize_text(str(worksheet["U2"].value or "")) == "RESINA VIRGEM"


def build_row(
    source_file: Path,
    sheet_name: str,
    range_ref: str,
    source_cell: str,
    metric_row: int,
    period: dict[str, Any],
    layout_type: str,
    section_original: Any,
    product: Any,
    metric_label_source_cell: str,
    metric_label_original: Any,
    column_u_original: Any,
    value: Any,
    formula: Any,
) -> dict[str, Any]:
    return {
        "source_file": str(source_file),
        "source_sheet": sheet_name,
        "source_range": range_ref,
        "source_cell": source_cell,
        "metric_row": metric_row,
        **period,
        "layout_type": layout_type,
        "section_original": section_original,
        "section_english": translate(section_original),
        "product": product,
        "metric_label_source_cell": metric_label_source_cell,
        "metric_label_original": metric_label_original,
        "metric_label_english": translate(metric_label_original),
        "column_u_original": column_u_original,
        "column_u_english": translate(column_u_original),
        "value": value,
        "formula": formula,
    }


def extract_long_rows_for_sheet(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
) -> list[dict[str, Any]]:
    period = parse_month_sheet(sheet_name)
    if not period:
        raise ValueError(f"Cannot parse month from sheet name: {sheet_name}")

    rows: list[dict[str, Any]] = []
    if is_label_value_layout(value_worksheet):
        section_original = clean_value(value_worksheet["U2"].value)
        product = clean_value(value_worksheet["V2"].value)
        for row_idx in range(3, 13):
            value = clean_value(value_worksheet.cell(row_idx, 22).value)
            if value is None:
                continue

            metric_label_original = clean_value(value_worksheet.cell(row_idx, 21).value)
            rows.append(
                build_row(
                    source_file=source_file,
                    sheet_name=sheet_name,
                    range_ref=range_ref,
                    source_cell=f"V{row_idx}",
                    metric_row=row_idx,
                    period=period,
                    layout_type="labels_in_u_values_in_v",
                    section_original=section_original,
                    product=product,
                    metric_label_source_cell=f"U{row_idx}",
                    metric_label_original=metric_label_original,
                    column_u_original=metric_label_original,
                    value=value,
                    formula=formula_worksheet.cell(row_idx, 22).value,
                )
            )
        return rows

    section_original = clean_value(value_worksheet["T2"].value)
    products = [
        (21, clean_value(value_worksheet["U2"].value)),
        (22, clean_value(value_worksheet["V2"].value)),
    ]
    for col_idx, product in products:
        for row_idx in range(3, 13):
            value = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            if value is None:
                continue

            column_letter = get_column_letter(col_idx)
            metric_label_original = clean_value(value_worksheet.cell(row_idx, 20).value)
            column_u_original = clean_value(value_worksheet.cell(row_idx, 21).value)
            rows.append(
                build_row(
                    source_file=source_file,
                    sheet_name=sheet_name,
                    range_ref=range_ref,
                    source_cell=f"{column_letter}{row_idx}",
                    metric_row=row_idx,
                    period=period,
                    layout_type="products_in_u_v_labels_in_t",
                    section_original=section_original,
                    product=product,
                    metric_label_source_cell=f"T{row_idx}",
                    metric_label_original=metric_label_original,
                    column_u_original=column_u_original,
                    value=value,
                    formula=formula_worksheet.cell(row_idx, col_idx).value,
                )
            )

    return rows


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=ordered_headers(rows))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_months_final.xlsx"


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        return

    headers = ordered_headers(rows)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(row.get(header) for row in rows[:100])
        width = min(
            max(len(str(value)) for value in values if value is not None) + 2,
            45,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for key, value in metadata.items():
        worksheet.append([key, value])

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 80


def write_excel_output(
    path: Path,
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, "raw_U2_V12", raw_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Valgroup month-sheet extraction pipeline."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Specific workbook to process. If omitted, the latest Valgroup workbook in data is used.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--range", default=DEFAULT_RANGE)
    parser.add_argument("--sheets", nargs="+", default=None)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=None,
        help="Optional explicit Excel output path.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files into the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)
    excel_output = args.excel_output or default_excel_output_path(
        source_file,
        args.output_dir,
    )

    values_workbook = load_workbook(source_file, read_only=True, data_only=True)
    formulas_workbook = load_workbook(source_file, read_only=True, data_only=False)

    try:
        sheet_names = select_month_sheets(values_workbook, args.sheets)
        raw_rows: list[dict[str, Any]] = []
        final_rows: list[dict[str, Any]] = []

        for sheet_name in sheet_names:
            value_worksheet = values_workbook[sheet_name]
            formula_worksheet = formulas_workbook[sheet_name]
            raw_rows.extend(extract_raw_rows(value_worksheet, sheet_name, args.range))
            final_rows.extend(
                extract_long_rows_for_sheet(
                    value_worksheet,
                    formula_worksheet,
                    source_file,
                    sheet_name,
                    args.range,
                )
            )

        metadata = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(source_file),
            "source_range": args.range,
            "source_sheets": "; ".join(sheet_names),
            "raw_rows": len(raw_rows),
            "final_rows": len(final_rows),
            "translation_source": "Column U when it contains labels; Column T for JAN layout where U:V are product values",
        }

        write_excel_output(excel_output, raw_rows, final_rows, metadata)

        if args.write_csv:
            write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
            write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

        print(f"source_file={source_file}")
        print(f"source_sheets={'; '.join(sheet_names)}")
        print(f"source_range={args.range}")
        print(f"raw_rows={len(raw_rows)}")
        print(f"final_rows={len(final_rows)}")
        print(f"excel_output={excel_output}")
        if args.write_csv:
            print(f"raw_csv_output={args.output_dir / DEFAULT_RAW_CSV_NAME}")
            print(f"final_csv_output={args.output_dir / DEFAULT_LONG_CSV_NAME}")
    finally:
        values_workbook.close()
        formulas_workbook.close()


if __name__ == "__main__":
    main()
