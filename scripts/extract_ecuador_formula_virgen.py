from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_DATA_DIR = Path("data")
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_FILE_PATTERN = "*ECUADOR*.xlsx"
DEFAULT_SHEET = "Formula Virgen"
DEFAULT_RANGE = "A2:AT18"
DEFAULT_RAW_CSV_NAME = "ecuador_formula_virgen_raw.csv"
DEFAULT_LONG_CSV_NAME = "ecuador_formula_virgen_long.csv"
DEFAULT_ANCHOR_YEAR = 2026

MONTH_NUMBERS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

MONTH_ALIASES = {
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "juni": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
}

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize(value: str) -> str:
    value = strip_accents(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def canonical_month_from_text(value: str) -> str | None:
    tokens = set(normalize(value).split())
    for token in tokens:
        if token in MONTH_ALIASES:
            month_number = MONTH_ALIASES[token]
            for month_name, number in MONTH_NUMBERS.items():
                if number == month_number:
                    return month_name
    return None


def clean_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def resolve_source_file(
    explicit_file: Path | None,
    data_dir: Path,
    file_pattern: str,
) -> Path:
    if explicit_file:
        if not explicit_file.exists():
            raise FileNotFoundError(f"Source file not found: {explicit_file}")
        return explicit_file

    candidates = [
        path
        for path in data_dir.glob(file_pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No source files found in {data_dir} matching {file_pattern}"
        )

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def extract_raw_rows(worksheet: Any, range_ref: str) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    columns = [get_column_letter(col_idx) for col_idx in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(min_row, max_row + 1):
        row = {
            "source_range": range_ref,
            "source_row": row_idx,
        }
        for col_idx, column_letter in zip(range(min_col, max_col + 1), columns):
            row[column_letter] = clean_value(worksheet.cell(row_idx, col_idx).value)
        rows.append(row)

    return rows


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return

    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=ordered_headers(rows))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)
    return cleaned[:31]


def write_rows_to_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    if not rows:
        worksheet.append(["No rows extracted"])
        return

    headers = ordered_headers(rows)
    worksheet.append(headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = worksheet.cell(row_idx, col_idx, value=value)
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for row_idx in range(2, min(worksheet.max_row, 200) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            45,
        )


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for key, value in metadata.items():
        worksheet.append([key, value])

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 80


def write_excel_output(
    path: Path,
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
    raw_sheet_title: str = "raw_data",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_data", final_rows)
    write_rows_to_sheet(workbook, raw_sheet_title, raw_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(path)


def default_excel_output_path(source_file: Path, output_dir: Path) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", source_file.stem).strip()
    return output_dir / f"{safe_stem}_formula_virgen_final.xlsx"


def select_sheet(workbook: Any, explicit_sheet: str | None) -> str:
    sheet_name = explicit_sheet or DEFAULT_SHEET
    if sheet_name in workbook.sheetnames:
        return sheet_name

    normalized_target = normalize_sheet_name(sheet_name)
    matches = [
        candidate
        for candidate in workbook.sheetnames
        if normalize_sheet_name(candidate) == normalized_target
    ]
    if len(matches) == 1:
        return matches[0]

    raise ValueError(f"Sheet not found: {sheet_name}")


def normalize_sheet_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def parse_anchor_period(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", value.strip())
    if not match:
        raise ValueError("Anchor period must be in YYYY-MM format.")

    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        raise ValueError("Anchor period month must be between 1 and 12.")

    return year, month


def month_from_label(value: Any) -> int:
    tokens = normalize(str(value or "")).split()
    for token in tokens:
        if token in MONTH_ALIASES:
            return MONTH_ALIASES[token]
    raise ValueError(f"Could not parse month label: {value!r}")


def infer_file_period(source_file: Path) -> tuple[int, int] | None:
    month_name = canonical_month_from_text(source_file.stem)
    year_match = re.search(r"(?:^|[^0-9])(\d{4}|\d{2})(?:[^0-9]|$)", source_file.stem)
    if not month_name or not year_match:
        return None

    year = int(year_match.group(1))
    if year < 100:
        year += 2000

    return year, MONTH_NUMBERS[month_name]


def infer_anchor_period(
    source_file: Path,
    worksheet: Any,
    range_ref: str,
    explicit_anchor_period: str | None,
) -> tuple[int, int]:
    _, _, max_col, _ = range_boundaries(range_ref)
    rightmost_row_3_month = month_from_label(worksheet.cell(3, max_col).value)

    if explicit_anchor_period:
        anchor_year, anchor_month = parse_anchor_period(explicit_anchor_period)
        if anchor_month != rightmost_row_3_month:
            raise ValueError(
                "The explicit anchor period month does not match the row 3 "
                f"rightmost month. {worksheet.cell(3, max_col).coordinate} is "
                f"{worksheet.cell(3, max_col).value!r}."
            )
        return anchor_year, anchor_month

    file_period = infer_file_period(source_file)
    if file_period:
        file_year, file_month = file_period
        anchor_year = file_year
        if rightmost_row_3_month > file_month:
            anchor_year -= 1
        return anchor_year, rightmost_row_3_month

    return DEFAULT_ANCHOR_YEAR, rightmost_row_3_month


def add_months(year: int, month: int, offset: int) -> tuple[int, int]:
    month_index = year * 12 + (month - 1) + offset
    return month_index // 12, month_index % 12 + 1


def period_fields(year: int, month: int) -> dict[str, Any]:
    return {
        "time_period": f"{MONTH_LABELS[month]} {year}",
        "time_period_year": year,
        "time_period_month": month,
    }


def build_time_periods_by_column(
    worksheet: Any,
    data_start_col: int,
    max_col: int,
    anchor_period: tuple[int, int],
) -> dict[int, tuple[int, int]]:
    anchor_year, anchor_month = anchor_period
    rightmost_month = month_from_label(worksheet.cell(3, max_col).value)
    if anchor_month != rightmost_month:
        raise ValueError(
            f"Anchor month {anchor_month} does not match row 3 rightmost month "
            f"{rightmost_month}."
        )

    periods: dict[int, tuple[int, int]] = {}
    previous_month = anchor_month
    previous_year = anchor_year

    for col_idx in range(max_col, data_start_col - 1, -1):
        row_3_month = month_from_label(worksheet.cell(3, col_idx).value)
        year = previous_year
        if col_idx != max_col and row_3_month > previous_month:
            year -= 1

        periods[col_idx] = (year, row_3_month)
        previous_month = row_3_month
        previous_year = year

    return periods


def extract_long_rows(
    value_worksheet: Any,
    formula_worksheet: Any,
    source_file: Path,
    sheet_name: str,
    range_ref: str,
    anchor_period: tuple[int, int],
) -> list[dict[str, Any]]:
    min_col, _, max_col, _ = range_boundaries(range_ref)
    data_start_col = max(min_col, 15)
    time_periods_by_column = build_time_periods_by_column(
        value_worksheet,
        data_start_col,
        max_col,
        anchor_period,
    )
    rows: list[dict[str, Any]] = []

    for row_idx in range(4, 19):
        metric_en = value_worksheet.cell(row_idx, 1).value
        metric_component = value_worksheet.cell(row_idx, 3).value
        metric_detail = value_worksheet.cell(row_idx, 6).value
        if (
            (metric_en is None or str(metric_en).strip() == "")
            and str(metric_component or "").strip() == "Landed Price"
        ):
            metric_en = metric_component

        for col_idx in range(data_start_col, max_col + 1):
            value = clean_value(value_worksheet.cell(row_idx, col_idx).value)
            if value is None:
                continue

            column_letter = get_column_letter(col_idx)
            period_year, period_month = time_periods_by_column[col_idx]
            rows.append(
                {
                    "source_file": str(source_file),
                    "source_sheet": sheet_name,
                    "source_range": range_ref,
                    "source_cell": f"{column_letter}{row_idx}",
                    "metric_row": row_idx,
                    "metric_label_source_row": row_idx,
                    "metric_en": metric_en,
                    "metric_component": metric_component,
                    "metric_detail": metric_detail,
                    "price_period": value_worksheet.cell(3, col_idx).value,
                    "index_period": value_worksheet.cell(2, col_idx).value,
                    **period_fields(period_year, period_month),
                    "value": value,
                    "formula": formula_worksheet.cell(row_idx, col_idx).value,
                }
            )

    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the monthly Ecuador Formula Virgen extraction pipeline."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Specific workbook to process. If omitted, the latest Ecuador workbook in data is used.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--file-pattern", default=DEFAULT_FILE_PATTERN)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--range", default=DEFAULT_RANGE)
    parser.add_argument(
        "--anchor-period",
        default=None,
        help="Period represented by the rightmost row 3 month, in YYYY-MM format. Defaults to row 3 month plus year inferred from the source file name.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=None,
        help="Optional explicit Excel output path.",
    )
    parser.add_argument(
        "--write-csv",
        action="store_true",
        help="Also write raw and final CSV files into the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_file = resolve_source_file(args.file, args.data_dir, args.file_pattern)
    excel_output = args.excel_output or default_excel_output_path(
        source_file, args.output_dir
    )

    values_workbook = load_workbook(source_file, read_only=True, data_only=True)
    formulas_workbook = load_workbook(source_file, read_only=True, data_only=False)

    try:
        sheet_name = select_sheet(values_workbook, args.sheet)
        value_worksheet = values_workbook[sheet_name]
        formula_worksheet = formulas_workbook[sheet_name]
        anchor_period = infer_anchor_period(
            source_file,
            value_worksheet,
            args.range,
            args.anchor_period,
        )

        raw_rows = extract_raw_rows(value_worksheet, args.range)
        final_rows = extract_long_rows(
            value_worksheet,
            formula_worksheet,
            source_file,
            sheet_name,
            args.range,
            anchor_period,
        )
        anchor_period_text = f"{MONTH_LABELS[anchor_period[1]]} {anchor_period[0]}"

        metadata = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(source_file),
            "source_sheet": sheet_name,
            "source_range": args.range,
            "anchor_A2": repr(value_worksheet["A2"].value),
            "anchor_period_cell": "AT3",
            "anchor_period_source": "row_3_month_label",
            "anchor_period": anchor_period_text,
            "raw_rows": len(raw_rows),
            "final_rows": len(final_rows),
        }

        write_excel_output(
            excel_output,
            raw_rows,
            final_rows,
            metadata,
            raw_sheet_title="raw_A2_AT18",
        )

        if args.write_csv:
            write_csv(args.output_dir / DEFAULT_RAW_CSV_NAME, raw_rows)
            write_csv(args.output_dir / DEFAULT_LONG_CSV_NAME, final_rows)

        print(f"source_file={source_file}")
        print(f"source_sheet={sheet_name}")
        print(f"source_range={args.range}")
        print(f"anchor_A2={value_worksheet['A2'].value!r}")
        print(f"anchor_period_cell=AT3")
        print("anchor_period_source=row_3_month_label")
        print(f"anchor_period={anchor_period_text}")
        print(f"raw_rows={len(raw_rows)}")
        print(f"final_rows={len(final_rows)}")
        print(f"excel_output={excel_output}")
        if args.write_csv:
            print(f"raw_csv_output={args.output_dir / DEFAULT_RAW_CSV_NAME}")
            print(f"final_csv_output={args.output_dir / DEFAULT_LONG_CSV_NAME}")
    finally:
        values_workbook.close()
        formulas_workbook.close()


if __name__ == "__main__":
    main()
