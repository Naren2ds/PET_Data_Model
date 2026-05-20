from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_FILE = (
    Path("output") / "1 Bavaria - Precio Marzo_formula_final_with_mapping_columns.xlsx"
)
DEFAULT_TEMPLATE_FILE = Path("standardization") / "Data_Standardization.xlsx"
DEFAULT_FORMULA_OVERVIEW_FILE = (
    Path("mapping files") / "Formula overview by countries.xlsx"
)
DEFAULT_FORMULA_SHEETS = ("Pricing formula_Updated", "Pricing formula overview")
DEFAULT_FINAL_SHEET = "final_data"
DEFAULT_TEMPLATE_SHEET = "Sheet1"
DEFAULT_OUTPUT_DIR = Path("data standardization")
DEFAULT_OUTPUT_FILE = DEFAULT_OUTPUT_DIR / "Amcor_Bavaria_final_standardized.xlsx"
DEFAULT_SUPPLIER = "Amcor"
DEFAULT_DESTINATION_COUNTRY = "Colombia"

SOURCE_FILE_COLUMN = "source_file"
RAW_COST_BREAKDOWN_COLUMN = "metric_en"
RESIN_INDEX_TYPE_COLUMN = "resin_index_type"
TIME_PERIOD_COLUMN = "price_period"
MAPPING_COLUMN = "Mapping Columns"
VALUE_COLUMN = "value"
RESIN_INDEX_OUTPUT_COLUMN = "Resin Index Type"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

MONTH_LOOKUP = {
    "jan": 1,
    "january": 1,
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "february": 2,
    "febrero": 2,
    "mar": 3,
    "march": 3,
    "marzo": 3,
    "apr": 4,
    "april": 4,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "june": 6,
    "junio": 6,
    "jul": 7,
    "july": 7,
    "julio": 7,
    "aug": 8,
    "august": 8,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "set": 9,
    "septiembre": 9,
    "oct": 10,
    "october": 10,
    "octubre": 10,
    "nov": 11,
    "november": 11,
    "noviembre": 11,
    "dec": 12,
    "december": 12,
    "dic": 12,
    "diciembre": 12,
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or "")).lower()
    text = re.sub(r"\.[A-Za-z0-9]+$", "", text)
    text = re.sub(
        r"(_formula_final_with_mapping_columns|_formula_final|_with_mapping_columns|_final)$",
        "",
        text,
    )
    text = text.replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def load_template_headers(template_file: Path, template_sheet: str) -> list[str]:
    workbook = load_workbook(template_file, data_only=True, read_only=True)
    try:
        worksheet = workbook[template_sheet]
        headers = [cell.value for cell in worksheet[1] if cell.value is not None]
        if RESIN_INDEX_OUTPUT_COLUMN not in headers:
            insert_at = headers.index("Raw Cost Breakdown") + 1
            headers.insert(insert_at, RESIN_INDEX_OUTPUT_COLUMN)
        return headers
    finally:
        workbook.close()


def load_formula_lookup(
    formula_overview_file: Path,
    sheet_names: tuple[str, ...],
) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(formula_overview_file, data_only=True, read_only=True)
    try:
        lookup: dict[str, dict[str, Any]] = {}
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]
            headers = [clean_header(cell.value) for cell in worksheet[2]]
            required = {"Countries", "Supplier", "Pricing Sheet"}
            if not required.issubset(set(headers)):
                continue

            country_idx = headers.index("Countries")
            supplier_idx = headers.index("Supplier")
            pricing_sheet_idx = headers.index("Pricing Sheet")
            index_idx = headers.index("Index") if "Index" in headers else None
            current_country: Any = None

            for row in worksheet.iter_rows(min_row=3, values_only=True):
                if row[country_idx] is not None:
                    current_country = clean_text(row[country_idx])

                pricing_sheet = clean_text(row[pricing_sheet_idx])
                supplier = clean_text(row[supplier_idx])
                if not pricing_sheet:
                    continue

                normalized = normalize_key(pricing_sheet)
                lookup.setdefault(
                    normalized,
                    {
                        "supplier": supplier,
                        "destination_country": current_country,
                        "pricing_sheet": pricing_sheet,
                        "formula_sheet": sheet_name,
                        "formula_index": clean_text(row[index_idx]) if index_idx is not None else None,
                    },
                )
        return lookup
    finally:
        workbook.close()


def load_existing_standardization_metadata(output_file: Path) -> dict[str, Any]:
    metadata = {
        "supplier": DEFAULT_SUPPLIER,
        "destination_country": DEFAULT_DESTINATION_COUNTRY,
    }
    if not output_file.exists():
        return metadata

    try:
        workbook = load_workbook(output_file, data_only=True, read_only=True)
    except PermissionError:
        return metadata

    try:
        if "final_standardized" not in workbook.sheetnames:
            return metadata

        worksheet = workbook["final_standardized"]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        supplier_idx = headers.index("Supplier Name") if "Supplier Name" in headers else None
        country_idx = (
            headers.index("Destination Country")
            if "Destination Country" in headers
            else None
        )

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if supplier_idx is not None and row[supplier_idx]:
                metadata["supplier"] = clean_text(row[supplier_idx])
            if country_idx is not None and row[country_idx]:
                metadata["destination_country"] = clean_text(row[country_idx])
            if metadata.get("supplier") and metadata.get("destination_country"):
                break
        return metadata
    finally:
        workbook.close()


def fallback_formula_lookup(
    pricing_sheet_key: str,
    output_file: Path,
    final_rows: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    existing_metadata = load_existing_standardization_metadata(output_file)
    resin_index_type = next(
        (
            clean_text(row.get(RESIN_INDEX_TYPE_COLUMN))
            for row in final_rows
            if row.get(RESIN_INDEX_TYPE_COLUMN)
        ),
        None,
    )
    return {
        normalize_key(pricing_sheet_key): {
            "supplier": existing_metadata.get("supplier"),
            "destination_country": existing_metadata.get("destination_country"),
            "pricing_sheet": pricing_sheet_key,
            "formula_sheet": "fallback_existing_standardized_output",
            "formula_index": resin_index_type,
        }
    }


def derive_pricing_sheet_key(source_file_value: Any, input_file: Path) -> str:
    if source_file_value:
        return Path(str(source_file_value)).stem

    stem = input_file.stem
    for suffix in (
        "_formula_final_with_mapping_columns",
        "_formula_final",
        "_with_mapping_columns",
        "_final",
    ):
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def find_formula_metadata(
    lookup: dict[str, dict[str, Any]],
    pricing_sheet_key: str,
) -> dict[str, Any]:
    normalized = normalize_key(pricing_sheet_key)
    if normalized in lookup:
        return lookup[normalized]

    for candidate_key, metadata in lookup.items():
        if candidate_key and (candidate_key in normalized or normalized in candidate_key):
            return metadata

    raise ValueError(
        f"No supplier/country mapping found for pricing sheet {pricing_sheet_key!r}."
    )


def parse_period(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    if value is None:
        return None

    text = strip_accents(str(value)).strip()
    if not text:
        return None

    parts = [part for part in re.split(r"[-/\s]+", text) if part]
    if len(parts) < 2:
        return None

    month_token = parts[0].lower()
    month = MONTH_LOOKUP.get(month_token, MONTH_LOOKUP.get(month_token[:3]))
    if month is None:
        return None

    year_token = re.sub(r"\D", "", parts[-1])
    if not year_token:
        return None

    year = int(year_token)
    if year < 100:
        year += 2000

    return date(year, month, 1)


def format_period(value: Any) -> str | None:
    parsed = parse_period(value)
    if parsed is None:
        return None
    return f"{MONTH_LABELS[parsed.month]} {parsed.year}"


def require_columns(headers: list[str], required_columns: list[str]) -> None:
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValueError(f"Missing required final_data columns: {missing}")


def load_final_rows(input_file: Path, final_sheet: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_file, data_only=True, read_only=True)
    try:
        if final_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {final_sheet!r} not found in {input_file}.")

        worksheet = workbook[final_sheet]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        rows = [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
        return headers, rows
    finally:
        workbook.close()


def build_standardized_rows(
    input_file: Path,
    final_rows: list[dict[str, Any]],
    formula_lookup: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not final_rows:
        return [], {}

    pricing_sheet_key = derive_pricing_sheet_key(
        final_rows[0].get(SOURCE_FILE_COLUMN),
        input_file,
    )
    formula_metadata = find_formula_metadata(formula_lookup, pricing_sheet_key)
    file_resin_index_type = next(
        (
            clean_text(row.get(RESIN_INDEX_TYPE_COLUMN))
            for row in final_rows
            if row.get(RESIN_INDEX_TYPE_COLUMN)
        ),
        formula_metadata.get("formula_index"),
    )

    standardized_rows: list[dict[str, Any]] = []
    for row in final_rows:
        source_file = clean_text(row.get(SOURCE_FILE_COLUMN))
        source_file_name = Path(str(source_file)).name if source_file else input_file.name
        standardized_rows.append(
            {
                "Source File ": source_file_name,
                "Supplier Name": formula_metadata.get("supplier"),
                "Destination Country": formula_metadata.get("destination_country"),
                "Time_Period ": format_period(row.get(TIME_PERIOD_COLUMN)),
                "Raw Cost Breakdown": clean_text(row.get(RAW_COST_BREAKDOWN_COLUMN)),
                RESIN_INDEX_OUTPUT_COLUMN: clean_text(row.get(RESIN_INDEX_TYPE_COLUMN))
                or file_resin_index_type,
                "Mapping Columns": clean_text(row.get(MAPPING_COLUMN)),
                "Value ": row.get(VALUE_COLUMN),
            }
        )

    return standardized_rows, {
        "pricing_sheet_key": pricing_sheet_key,
        "resin_index_type": file_resin_index_type,
        **formula_metadata,
    }


def style_header(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def adjust_widths(worksheet: Any) -> None:
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            60,
        )


def write_rows_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    adjust_widths(worksheet)


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for key, value in metadata.items():
        worksheet.append([key, value])

    style_header(worksheet)
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 80


def write_output(
    output_file: Path,
    headers: list[str],
    rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "final_standardized", headers, rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Standardize the Bavaria final_data workbook into the web-app data model."
    )
    parser.add_argument("--input-file", type=Path, default=DEFAULT_INPUT_FILE)
    parser.add_argument("--template-file", type=Path, default=DEFAULT_TEMPLATE_FILE)
    parser.add_argument("--template-sheet", default=DEFAULT_TEMPLATE_SHEET)
    parser.add_argument(
        "--formula-overview-file",
        type=Path,
        default=DEFAULT_FORMULA_OVERVIEW_FILE,
    )
    parser.add_argument("--final-sheet", default=DEFAULT_FINAL_SHEET)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    template_headers = load_template_headers(args.template_file, args.template_sheet)
    final_headers, final_rows = load_final_rows(args.input_file, args.final_sheet)
    require_columns(
        final_headers,
        [
            SOURCE_FILE_COLUMN,
            RAW_COST_BREAKDOWN_COLUMN,
            TIME_PERIOD_COLUMN,
            MAPPING_COLUMN,
            VALUE_COLUMN,
        ],
    )
    pricing_sheet_key = derive_pricing_sheet_key(
        final_rows[0].get(SOURCE_FILE_COLUMN) if final_rows else None,
        args.input_file,
    )
    try:
        formula_lookup = load_formula_lookup(
            args.formula_overview_file,
            DEFAULT_FORMULA_SHEETS,
        )
        formula_lookup_source = str(args.formula_overview_file)
    except PermissionError:
        formula_lookup = fallback_formula_lookup(
            pricing_sheet_key,
            args.output_file,
            final_rows,
        )
        formula_lookup_source = "fallback_existing_standardized_output"
    standardized_rows, formula_metadata = build_standardized_rows(
        args.input_file,
        final_rows,
        formula_lookup,
    )
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "input_file": str(args.input_file),
        "template_file": str(args.template_file),
        "formula_overview_file": str(args.formula_overview_file),
        "formula_lookup_source": formula_lookup_source,
        "output_file": str(args.output_file),
        "source_rows": len(final_rows),
        "standardized_rows": len(standardized_rows),
        "pricing_sheet_key": formula_metadata.get("pricing_sheet_key"),
        "matched_pricing_sheet": formula_metadata.get("pricing_sheet"),
        "supplier": formula_metadata.get("supplier"),
        "destination_country": formula_metadata.get("destination_country"),
        "resin_index_type": formula_metadata.get("resin_index_type"),
        "formula_lookup_sheet": formula_metadata.get("formula_sheet"),
    }
    write_output(args.output_file, template_headers, standardized_rows, metadata)

    print(f"input_file={args.input_file}")
    print(f"output_file={args.output_file}")
    print(f"source_rows={len(final_rows)}")
    print(f"standardized_rows={len(standardized_rows)}")
    print(f"supplier={formula_metadata.get('supplier')}")
    print(f"destination_country={formula_metadata.get('destination_country')}")


if __name__ == "__main__":
    main()
