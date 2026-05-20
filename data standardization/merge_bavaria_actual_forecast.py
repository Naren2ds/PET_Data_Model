from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_ACTUAL_FILE = Path("data standardization") / "Amcor_Bavaria_final_standardized.xlsx"
DEFAULT_FORECAST_FILE = Path("Estimation") / "bavaria_2026_forward_estimation.xlsx"
DEFAULT_TLC_FILE = Path("data calculation") / "Bavaria_tlc_calculation_validation.xlsx"
DEFAULT_OUTPUT_FILE = (
    Path("data standardization") / "Bavaria_actual_forecast_final_standardized.xlsx"
)
DEFAULT_FRONTEND_CSV = (
    Path("Front End")
    / "PET-Backend"
    / "PET-Backend"
    / "standardized_data"
    / "bavaria_actual_forecast_final_standardized.csv"
)

ACTUAL_SHEET = "final_standardized"
FORECAST_SHEET = "standardized_rows"
TLC_VALIDATION_SHEET = "tlc_validation"
TLC_FORMULA_COLUMN = "TLC Formula"
DEFAULT_TLC_FORMULA = (
    "Total Landing Cost = Sub Total (with Incremental Freight) + "
    "Duty 5% (Change According to Regulation) + Landed Factor 8% + "
    "ZF Legislation Change + Sur Charge Alpek Br"
)

OUTPUT_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Value ",
    TLC_FORMULA_COLUMN,
]

MONTH_LOOKUP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def clean_header(value: Any) -> str:
    return str(value or "").replace("\xa0", " ")


def row_value(row: dict[str, Any], header: str) -> Any:
    if header in row:
        return row.get(header)
    return row.get(header.strip())


def read_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found in {path}.")

        worksheet = workbook[sheet_name]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


def load_tlc_formula(path: Path) -> str:
    if not path.exists():
        return DEFAULT_TLC_FORMULA

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if TLC_VALIDATION_SHEET not in workbook.sheetnames:
            return DEFAULT_TLC_FORMULA

        worksheet = workbook[TLC_VALIDATION_SHEET]
        headers = [clean_header(cell.value).strip() for cell in worksheet[1]]
        if "formula_description" not in headers:
            return DEFAULT_TLC_FORMULA

        formula_idx = headers.index("formula_description")
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            formula = row[formula_idx]
            if formula:
                formula_text = str(formula).strip()
                if formula_text.lower().startswith("total landing cost"):
                    return formula_text
                return f"Total Landing Cost = {formula_text}"
        return DEFAULT_TLC_FORMULA
    finally:
        workbook.close()


def period_parts(value: Any) -> tuple[int | None, int | None]:
    if value is None:
        return None, None

    parts = str(value).strip().split()
    if len(parts) < 2:
        return None, None

    month = MONTH_LOOKUP.get(parts[0].lower())
    try:
        year = int(parts[-1])
    except ValueError:
        year = None

    return year, month


def normalize_row(row: dict[str, Any], data_type: str, tlc_formula: str) -> dict[str, Any]:
    period = row_value(row, "Time_Period ")
    year, month = period_parts(period)
    normalized = {header: row_value(row, header) for header in OUTPUT_HEADERS}
    normalized["Data Type"] = data_type
    normalized["Time Period Year"] = year
    normalized["Time Period Month"] = month
    normalized[TLC_FORMULA_COLUMN] = tlc_formula
    return normalized


def merged_rows(
    actual_rows: list[dict[str, Any]],
    forecast_rows: list[dict[str, Any]],
    tlc_formula: str,
) -> list[dict[str, Any]]:
    rows = [normalize_row(row, "Actual", tlc_formula) for row in actual_rows]
    rows.extend(normalize_row(row, "Forecast", tlc_formula) for row in forecast_rows)
    return sorted(
        rows,
        key=lambda row: (
            row.get("Destination Country") or "",
            row.get("Supplier Name") or "",
            row.get("Time Period Year") or 0,
            row.get("Time Period Month") or 0,
            0 if row.get("Data Type") == "Actual" else 1,
            row.get("Mapping Columns") or "",
            row.get("Raw Cost Breakdown") or "",
        ),
    )


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            60,
        )


def write_excel(path: Path, rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "final_standardized"
    worksheet.append(OUTPUT_HEADERS)
    for row in rows:
        worksheet.append([row.get(header) for header in OUTPUT_HEADERS])
    style_sheet(worksheet)

    metadata_sheet = workbook.create_sheet("run_metadata")
    metadata_sheet.append(["field", "value"])
    for key, value in metadata.items():
        metadata_sheet.append([key, value])
    style_sheet(metadata_sheet)
    metadata_sheet.column_dimensions["A"].width = 32
    metadata_sheet.column_dimensions["B"].width = 80

    workbook.save(path)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge Bavaria actual and forecast standardized rows for web-app use."
    )
    parser.add_argument("--actual-file", type=Path, default=DEFAULT_ACTUAL_FILE)
    parser.add_argument("--forecast-file", type=Path, default=DEFAULT_FORECAST_FILE)
    parser.add_argument("--tlc-file", type=Path, default=DEFAULT_TLC_FILE)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--frontend-csv", type=Path, default=DEFAULT_FRONTEND_CSV)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    actual_rows = read_rows(args.actual_file, ACTUAL_SHEET)
    forecast_rows = read_rows(args.forecast_file, FORECAST_SHEET)
    tlc_formula = load_tlc_formula(args.tlc_file)
    rows = merged_rows(actual_rows, forecast_rows, tlc_formula)
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "actual_file": str(args.actual_file),
        "forecast_file": str(args.forecast_file),
        "tlc_file": str(args.tlc_file),
        "output_file": str(args.output_file),
        "frontend_csv": str(args.frontend_csv),
        "tlc_formula": tlc_formula,
        "actual_rows": len(actual_rows),
        "forecast_rows": len(forecast_rows),
        "merged_rows": len(rows),
    }

    write_excel(args.output_file, rows, metadata)
    write_csv(args.frontend_csv, rows)

    print(f"actual_rows={len(actual_rows)}")
    print(f"forecast_rows={len(forecast_rows)}")
    print(f"merged_rows={len(rows)}")
    print(f"output_file={args.output_file}")
    print(f"frontend_csv={args.frontend_csv}")


if __name__ == "__main__":
    main()
