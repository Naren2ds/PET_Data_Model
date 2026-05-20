from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PIPELINE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[3]

SUPPLIER_PIPELINE_NAME = "3 Precios AB Inbev_Abril 26 ECUADOR"
EXTRACTED_FILE = (
    PIPELINE_DIR
    / "artifacts"
    / "extraction"
    / "3 Precios AB Inbev_Abril 26 ECUADOR_formula_virgen_final.xlsx"
)
ACTUAL_FRONT_END_CSV = (
    PIPELINE_DIR
    / "artifacts"
    / "front_end"
    / "3 Precios AB Inbev_Abril 26 ECUADOR_front_end_standardized.csv"
)
RESIN_FORECAST_CSV = REPO_ROOT / "Estimation" / "resin_index_forecast_table.csv"

CALCULATION_DIR = PIPELINE_DIR / "artifacts" / "calculation"
FORECAST_DIR = PIPELINE_DIR / "artifacts" / "forecast"
FRONT_END_DIR = PIPELINE_DIR / "artifacts" / "front_end"
VALIDATION_DIR = PIPELINE_DIR / "artifacts" / "validation"

TLC_VALIDATION_FILE = CALCULATION_DIR / "ecuador_tlc_calculation_validation.xlsx"
FORECAST_INPUTS_FILE = FORECAST_DIR / "ecuador_forecast_inputs_template.xlsx"
FORWARD_ESTIMATION_FILE = FORECAST_DIR / "ecuador_2026_forward_estimation.xlsx"
FRONT_END_ACTUAL_FORECAST_XLSX = (
    FRONT_END_DIR / "3 Precios AB Inbev_Abril 26 ECUADOR_actual_forecast_front_end_standardized.xlsx"
)
FRONT_END_ACTUAL_FORECAST_CSV = (
    FRONT_END_DIR / "3 Precios AB Inbev_Abril 26 ECUADOR_actual_forecast_front_end_standardized.csv"
)
FORECAST_VALIDATION_FILE = VALIDATION_DIR / "forecast_validation_summary.json"

FINAL_DATA_SHEET = "final_data"
SUPPLIER = "Amcor"
DESTINATION_COUNTRY = "Ecuador"
SOURCE_RESIN_INDEX_TYPE = "ICIS PET China MID (N-2) USD/ton"
SOURCE_RESIN_LABEL = "ICIS China Mid (n-2)"
FORECAST_RESIN_INDEX_TYPE = "PET Bottle Grade FOB China Spot"
TLC_FORMULA = (
    "Landed Price = CIF + 0.10% of CIF Value + Foreign Currency Exit Tax 4% of CIF Value "
    "+ 0.5% of CIF Value + customs clearance/freight fixed components + Government Diesel Surcharge"
)

FORECAST_MONTHS = [
    "April 2026",
    "May 2026",
    "June 2026",
    "July 2026",
    "August 2026",
    "September 2026",
    "October 2026",
    "November 2026",
    "December 2026",
]

COMPONENT_COLUMNS = {
    "resin_price_index": "ICIS China Mid (n-2)",
    "freight": "Freight ",
    "finance_fee": "Finance Fee V (SOFR +3.8%@175)",
    "cif": None,
    "insurance_0_10": "0.10% of CIF Value",
    "foreign_currency_exit_tax_4": "Foreign Currency Exit Tax 4% of CIF Value",
    "fodinfa_0_5": "0.5% of CIF Value",
    "customs_282": "282 USD/Container Customs",
    "port_storage": "Port Storage - USD250/Container",
    "local_freight": "Freight GYE - UIO USD671/Container",
    "shipping_container_return": "Shipping Exp. & Container Return - UIO USD430/Container",
    "fixed_bl_inspection": "Fixed Value per BL & +30 if Physical Inspection",
    "customs_clearance_fixed": "Customs Clearance Fixed: 36.02/ton",
    "government_diesel_surcharge": "Government Diesel Surcharge",
    "landed_price": "Landed Price",
}

FRONT_END_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Value ",
    "TLC Formula",
]


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return " ".join(value.replace("\xa0", " ").split())
    return value


def numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def read_final_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[FINAL_DATA_SHEET]
        headers = [clean_text(cell.value) for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


def rows_by_period(final_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in final_rows:
        grouped.setdefault(row["time_period"], []).append(row)
    return grouped


def row_for_metric(rows: list[dict[str, Any]], metric_name: Any) -> dict[str, Any] | None:
    for row in rows:
        if clean_text(row.get("metric_en")) == clean_text(metric_name):
            return row
    return None


def row_for_metric_row(rows: list[dict[str, Any]], metric_row: int) -> dict[str, Any] | None:
    for row in rows:
        if row.get("metric_row") == metric_row:
            return row
    return None


def period_sort_key(row: dict[str, Any]) -> tuple[int, int]:
    return int(row.get("time_period_year") or 0), int(row.get("time_period_month") or 0)


def latest_period_rows(grouped: dict[str, list[dict[str, Any]]]) -> tuple[str, list[dict[str, Any]]]:
    period = max(grouped, key=lambda key: period_sort_key(grouped[key][0]))
    return period, grouped[period]


def value_for(rows: list[dict[str, Any]], metric_name: Any) -> float:
    row = row_for_metric(rows, metric_name)
    return numeric(row.get("value") if row else None)


def validation_rows(final_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped = rows_by_period(final_rows)
    rows: list[dict[str, Any]] = []
    for period, period_rows in sorted(grouped.items(), key=lambda item: period_sort_key(item[1][0])):
        source_landed_row = row_for_metric(period_rows, "Landed Price")
        if source_landed_row is None:
            continue

        component_rows = [
            row
            for row in period_rows
            if isinstance(row.get("metric_row"), int) and 7 <= row["metric_row"] <= 17
        ]
        calculated_landed_price = sum(numeric(row.get("value")) for row in component_rows)
        source_landed_price = numeric(source_landed_row.get("value"))
        difference = calculated_landed_price - source_landed_price

        values = {
            key: value_for(period_rows, metric)
            for key, metric in COMPONENT_COLUMNS.items()
            if metric is not None and key != "landed_price"
        }
        cif_row = row_for_metric_row(period_rows, 7)
        values["cif"] = numeric(cif_row.get("value") if cif_row else None)

        rows.append(
            {
                "time_period": period,
                "time_period_year": source_landed_row.get("time_period_year"),
                "time_period_month": source_landed_row.get("time_period_month"),
                "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
                **values,
                "calculated_landed_price": calculated_landed_price,
                "source_landed_price": source_landed_price,
                "difference": difference,
                "abs_difference": abs(difference),
                "validation_status": "match" if abs(difference) <= 0.000001 else "check",
                "source_landed_formula": source_landed_row.get("formula"),
                "tlc_formula": TLC_FORMULA,
            }
        )
    return rows


def latest_defaults(latest_rows: list[dict[str, Any]]) -> dict[str, float]:
    local_freight = value_for(latest_rows, COMPONENT_COLUMNS["local_freight"])
    diesel = value_for(latest_rows, COMPONENT_COLUMNS["government_diesel_surcharge"])
    return {
        "freight": value_for(latest_rows, COMPONENT_COLUMNS["freight"]),
        "finance_fee": value_for(latest_rows, COMPONENT_COLUMNS["finance_fee"]),
        "insurance_rate": 0.001,
        "foreign_currency_exit_tax_rate": 0.04,
        "fodinfa_rate": 0.005,
        "customs_282": value_for(latest_rows, COMPONENT_COLUMNS["customs_282"]),
        "port_storage": value_for(latest_rows, COMPONENT_COLUMNS["port_storage"]),
        "local_freight": local_freight,
        "shipping_container_return": value_for(latest_rows, COMPONENT_COLUMNS["shipping_container_return"]),
        "fixed_bl_inspection": value_for(latest_rows, COMPONENT_COLUMNS["fixed_bl_inspection"]),
        "customs_clearance_fixed": value_for(latest_rows, COMPONENT_COLUMNS["customs_clearance_fixed"]),
        "diesel_surcharge_rate": diesel / local_freight if local_freight else 0.23,
    }


def read_resin_forecast(path: Path) -> dict[str, dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        rows = list(csv.DictReader(csv_file))
    forecast: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("resin_index_type") != FORECAST_RESIN_INDEX_TYPE:
            continue
        period = row.get("time_period")
        if period in FORECAST_MONTHS and period not in forecast:
            forecast[period] = row
    return forecast


def forecast_rows(defaults: dict[str, float], resin_forecast: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for period in FORECAST_MONTHS:
        forecast_row = resin_forecast.get(period)
        if not forecast_row:
            continue

        resin_price_index = numeric(forecast_row.get("value"))
        freight = defaults["freight"]
        finance_fee = defaults["finance_fee"]
        cif = resin_price_index + freight + finance_fee
        insurance = cif * defaults["insurance_rate"]
        exit_tax = cif * defaults["foreign_currency_exit_tax_rate"]
        fodinfa = cif * defaults["fodinfa_rate"]
        local_freight = defaults["local_freight"]
        diesel = local_freight * defaults["diesel_surcharge_rate"]
        landed_price = sum(
            [
                cif,
                insurance,
                exit_tax,
                fodinfa,
                defaults["customs_282"],
                defaults["port_storage"],
                local_freight,
                defaults["shipping_container_return"],
                defaults["fixed_bl_inspection"],
                defaults["customs_clearance_fixed"],
                diesel,
            ]
        )

        rows.append(
            {
                "time_period": period,
                "time_period_year": int(forecast_row["time_period_year"]),
                "time_period_month": int(forecast_row["time_period_month"]),
                "resin_price_index": resin_price_index,
                "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
                "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
                "resin_forecast_date": forecast_row.get("forecast_date"),
                "resin_forecast_series": forecast_row.get("forecast_series"),
                "freight": freight,
                "finance_fee": finance_fee,
                "cif": cif,
                "insurance_0_10": insurance,
                "foreign_currency_exit_tax_4": exit_tax,
                "fodinfa_0_5": fodinfa,
                "customs_282": defaults["customs_282"],
                "port_storage": defaults["port_storage"],
                "local_freight": local_freight,
                "shipping_container_return": defaults["shipping_container_return"],
                "fixed_bl_inspection": defaults["fixed_bl_inspection"],
                "customs_clearance_fixed": defaults["customs_clearance_fixed"],
                "government_diesel_surcharge": diesel,
                "landed_price_forecast": landed_price,
                "tlc_formula": TLC_FORMULA,
            }
        )
    return rows


def forecast_input_rows(defaults: dict[str, float], estimates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "time_period": row["time_period"],
            "time_period_year": row["time_period_year"],
            "time_period_month": row["time_period_month"],
            "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
            "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
            "resin_price_index": row["resin_price_index"],
            **defaults,
            "tlc_formula": TLC_FORMULA,
            "notes": "Update resin/freight/tax/customs assumptions here before rerunning forecast, if needed.",
        }
        for row in estimates
    ]


def forecast_standardized_rows(estimates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    component_specs = [
        ("resin_price_index", "ICIS China Mid (n-2)", "Resin Index vPET"),
        ("freight", "Freight", "Freight"),
        ("finance_fee", "Finance Fee V (SOFR +3.8%@175)", "Insurance"),
        ("cif", None, None),
        ("insurance_0_10", "0.10% of CIF Value", "Tax"),
        ("foreign_currency_exit_tax_4", "Foreign Currency Exit Tax 4% of CIF Value", "Tax"),
        ("fodinfa_0_5", "0.5% of CIF Value", "Customs clearance"),
        ("customs_282", "282 USD/Container Customs", "Customs clearance"),
        ("port_storage", "Port Storage - USD250/Container", "Customs clearance"),
        ("local_freight", "Freight GYE - UIO USD671/Container", "Customs clearance"),
        ("shipping_container_return", "Shipping Exp. & Container Return - UIO USD430/Container", "Customs clearance"),
        ("fixed_bl_inspection", "Fixed Value per BL & +30 if Physical Inspection", "Customs clearance"),
        ("customs_clearance_fixed", "Customs Clearance Fixed: 36.02/ton", "Customs clearance"),
        ("government_diesel_surcharge", "Government Diesel Surcharge", "Customs clearance"),
        ("landed_price_forecast", "Landed Price", "Total Landing Cost"),
    ]
    rows: list[dict[str, Any]] = []
    for estimate in estimates:
        for key, metric, mapping_column in component_specs:
            rows.append(
                {
                    "Data Type": "Forecast",
                    "Source File ": RESIN_FORECAST_CSV.name,
                    "Supplier Name": SUPPLIER,
                    "Destination Country": DESTINATION_COUNTRY,
                    "Time_Period ": estimate["time_period"],
                    "Time Period Year": estimate["time_period_year"],
                    "Time Period Month": estimate["time_period_month"],
                    "Raw Cost Breakdown": metric,
                    "Resin Index Type": SOURCE_RESIN_INDEX_TYPE,
                    "Forecast Resin Index Type": FORECAST_RESIN_INDEX_TYPE,
                    "Mapping Columns": mapping_column,
                    "Value ": estimate[key],
                    "TLC Formula": TLC_FORMULA,
                }
            )
    return rows


def read_actual_front_end_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        return list(csv.DictReader(csv_file))


def sorted_front_end_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("Destination Country") or "",
            row.get("Supplier Name") or "",
            int(row.get("Time Period Year") or 0),
            int(row.get("Time Period Month") or 0),
            0 if row.get("Data Type") == "Actual" else 1,
            row.get("Mapping Columns") or "",
            row.get("Raw Cost Breakdown") or "",
        ),
    )


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    style_sheet(worksheet)


def write_workbook(path: Path, sheets: list[tuple[str, list[str], list[dict[str, Any]]]], metadata: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers, rows in sheets:
        write_sheet(workbook, sheet_name, headers, rows)
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_outputs(
    validation: list[dict[str, Any]],
    defaults: dict[str, float],
    estimates: list[dict[str, Any]],
    forecast_inputs: list[dict[str, Any]],
    forecast_components: list[dict[str, Any]],
    merged_front_end: list[dict[str, Any]],
) -> None:
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "pipeline": SUPPLIER_PIPELINE_NAME,
        "extracted_file": str(EXTRACTED_FILE),
        "resin_forecast_csv": str(RESIN_FORECAST_CSV),
        "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
        "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
        "tlc_formula": TLC_FORMULA,
        "validation_rows": len(validation),
        "forecast_months": len(estimates),
        "forecast_component_rows": len(forecast_components),
        "merged_front_end_rows": len(merged_front_end),
    }
    validation_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_resin_index_type",
        "resin_price_index",
        "freight",
        "finance_fee",
        "cif",
        "insurance_0_10",
        "foreign_currency_exit_tax_4",
        "fodinfa_0_5",
        "customs_282",
        "port_storage",
        "local_freight",
        "shipping_container_return",
        "fixed_bl_inspection",
        "customs_clearance_fixed",
        "government_diesel_surcharge",
        "calculated_landed_price",
        "source_landed_price",
        "difference",
        "abs_difference",
        "validation_status",
        "source_landed_formula",
        "tlc_formula",
    ]
    estimate_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "resin_price_index",
        "source_resin_index_type",
        "forecast_resin_index_type",
        "resin_forecast_date",
        "resin_forecast_series",
        "freight",
        "finance_fee",
        "cif",
        "insurance_0_10",
        "foreign_currency_exit_tax_4",
        "fodinfa_0_5",
        "customs_282",
        "port_storage",
        "local_freight",
        "shipping_container_return",
        "fixed_bl_inspection",
        "customs_clearance_fixed",
        "government_diesel_surcharge",
        "landed_price_forecast",
        "tlc_formula",
    ]
    forecast_input_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_resin_index_type",
        "forecast_resin_index_type",
        "resin_price_index",
        "freight",
        "finance_fee",
        "insurance_rate",
        "foreign_currency_exit_tax_rate",
        "fodinfa_rate",
        "customs_282",
        "port_storage",
        "local_freight",
        "shipping_container_return",
        "fixed_bl_inspection",
        "customs_clearance_fixed",
        "diesel_surcharge_rate",
        "tlc_formula",
        "notes",
    ]
    catalog_rows = [
        {
            "field": key,
            "value": value,
            "classification": "latest_actual_assumption",
            "notes": "Carried from latest Ecuador actual period, March 2026.",
        }
        for key, value in defaults.items()
    ]
    catalog_rows.extend(
        [
            {
                "field": "resin_price_index",
                "value": None,
                "classification": "variable_forecast_input",
                "notes": f"Loaded from {RESIN_FORECAST_CSV.name}.",
            },
            {
                "field": "landed_price_forecast",
                "value": None,
                "classification": "calculated_output",
                "notes": TLC_FORMULA,
            },
        ]
    )

    write_workbook(
        TLC_VALIDATION_FILE,
        [("tlc_validation", validation_headers, validation)],
        metadata,
    )
    write_workbook(
        FORECAST_INPUTS_FILE,
        [
            ("forecast_inputs", forecast_input_headers, forecast_inputs),
            ("formula_catalog", ["field", "value", "classification", "notes"], catalog_rows),
        ],
        metadata,
    )
    write_workbook(
        FORWARD_ESTIMATION_FILE,
        [
            ("ecuador_2026_estimate", estimate_headers, estimates),
            ("standardized_rows", FRONT_END_HEADERS, forecast_components),
        ],
        metadata,
    )
    write_workbook(
        FRONT_END_ACTUAL_FORECAST_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, merged_front_end)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_FORECAST_CSV, FRONT_END_HEADERS, merged_front_end)

    statuses = {row["validation_status"] for row in validation}
    summary = {
        **metadata,
        "validation_statuses": sorted(statuses),
        "max_abs_difference": max((row["abs_difference"] for row in validation), default=0),
        "front_end_actual_forecast_csv": str(FRONT_END_ACTUAL_FORECAST_CSV),
    }
    FORECAST_VALIDATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    FORECAST_VALIDATION_FILE.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate Ecuador TLC and build forward forecast/front-end artifacts."
    )
    parser.add_argument("--extracted-file", type=Path, default=EXTRACTED_FILE)
    parser.add_argument("--actual-front-end-csv", type=Path, default=ACTUAL_FRONT_END_CSV)
    parser.add_argument("--resin-forecast-csv", type=Path, default=RESIN_FORECAST_CSV)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    global EXTRACTED_FILE, ACTUAL_FRONT_END_CSV, RESIN_FORECAST_CSV
    EXTRACTED_FILE = args.extracted_file
    ACTUAL_FRONT_END_CSV = args.actual_front_end_csv
    RESIN_FORECAST_CSV = args.resin_forecast_csv

    final_rows = read_final_rows(EXTRACTED_FILE)
    grouped = rows_by_period(final_rows)
    latest_period, latest_rows = latest_period_rows(grouped)
    defaults = latest_defaults(latest_rows)
    validation = validation_rows(final_rows)
    resin_forecast = read_resin_forecast(RESIN_FORECAST_CSV)
    estimates = forecast_rows(defaults, resin_forecast)
    forecast_inputs = forecast_input_rows(defaults, estimates)
    forecast_components = forecast_standardized_rows(estimates)
    actual_front_end_rows = read_actual_front_end_rows(ACTUAL_FRONT_END_CSV)
    merged_front_end = sorted_front_end_rows(actual_front_end_rows + forecast_components)

    write_outputs(
        validation,
        defaults,
        estimates,
        forecast_inputs,
        forecast_components,
        merged_front_end,
    )

    print(f"latest_actual_period={latest_period}")
    print(f"validation_rows={len(validation)}")
    print(f"forecast_months={len(estimates)}")
    print(f"forecast_component_rows={len(forecast_components)}")
    print(f"merged_front_end_rows={len(merged_front_end)}")
    print(f"tlc_validation_file={TLC_VALIDATION_FILE}")
    print(f"forecast_inputs_file={FORECAST_INPUTS_FILE}")
    print(f"forward_estimation_file={FORWARD_ESTIMATION_FILE}")
    print(f"front_end_actual_forecast_csv={FRONT_END_ACTUAL_FORECAST_CSV}")


if __name__ == "__main__":
    main()
