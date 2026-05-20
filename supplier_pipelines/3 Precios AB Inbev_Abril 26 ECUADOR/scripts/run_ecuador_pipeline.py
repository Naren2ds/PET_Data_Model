from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PIPELINE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[3]

SUPPLIER_PIPELINE_NAME = "3 Precios AB Inbev_Abril 26 ECUADOR"
SOURCE_FILE = REPO_ROOT / "data" / "3 Precios AB Inbev_Abril 26 ECUADOR.xlsx"
EXTRACT_SCRIPT = REPO_ROOT / "scripts" / "extract_ecuador_formula_virgen.py"
MAPPING_SCRIPT = REPO_ROOT / "scripts" / "data_mapping_column_ecuador.py"
FORECAST_SCRIPT = PIPELINE_DIR / "scripts" / "build_ecuador_calculation_forecast.py"
FORMULA_OVERVIEW_FILE = REPO_ROOT / "mapping files" / "Formula overview by countries.xlsx"

ARTIFACTS_DIR = PIPELINE_DIR / "artifacts"
EXTRACTION_DIR = ARTIFACTS_DIR / "extraction"
STANDARDIZED_DIR = ARTIFACTS_DIR / "standardized"
FRONT_END_DIR = ARTIFACTS_DIR / "front_end"
VALIDATION_DIR = ARTIFACTS_DIR / "validation"

EXTRACTED_FILE = (
    EXTRACTION_DIR / "3 Precios AB Inbev_Abril 26 ECUADOR_formula_virgen_final.xlsx"
)
STANDARDIZED_FILE = (
    STANDARDIZED_DIR / "3 Precios AB Inbev_Abril 26 ECUADOR_actual_standardized.xlsx"
)
FRONT_END_EXCEL_FILE = (
    FRONT_END_DIR / "3 Precios AB Inbev_Abril 26 ECUADOR_front_end_standardized.xlsx"
)
FRONT_END_CSV_FILE = (
    FRONT_END_DIR / "3 Precios AB Inbev_Abril 26 ECUADOR_front_end_standardized.csv"
)
VALIDATION_FILE = VALIDATION_DIR / "validation_summary.json"

FINAL_DATA_SHEET = "final_data"
FORMULA_OVERVIEW_SHEETS = ("Pricing formula_Updated", "Pricing formula overview")

DEFAULT_METADATA = {
    "supplier": "Amcor",
    "destination_country": "Ecuador",
    "sourcing_country": "China",
    "resin_index_type": "ICIS PET China MID (N-2) USD/ton",
    "tlc_formula": (
        "PET Resin: {[(Index + Int. Freight) x (1 + % Customs Clearance Cost)] "
        "+ Customs Clearance Fixed + Local Freight + Scrap}"
    ),
}

STANDARDIZED_HEADERS = [
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Mapping Columns",
    "Value ",
    "TLC Formula",
]

FRONT_END_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Value ",
    "TLC Formula",
]


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def clean_header(value: Any) -> str:
    return str(clean_text(value) or "")


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or "")).lower()
    text = text.replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def clean_index(value: Any) -> Any:
    text = clean_text(value)
    if isinstance(text, str):
        return re.sub(r"^index\s*:\s*", "", text, flags=re.IGNORECASE).strip()
    return text


def run_command(command: list[str]) -> None:
    print(" ".join(f'"{part}"' if " " in part else part for part in command))
    subprocess.run(command, cwd=REPO_ROOT, check=True)


def run_extraction() -> None:
    run_command(
        [
            sys.executable,
            str(EXTRACT_SCRIPT),
            "--file",
            str(SOURCE_FILE),
            "--output-dir",
            str(EXTRACTION_DIR),
            "--excel-output",
            str(EXTRACTED_FILE),
            "--write-csv",
        ]
    )


def run_mapping() -> None:
    run_command(
        [
            sys.executable,
            str(MAPPING_SCRIPT),
            "--output-file",
            str(EXTRACTED_FILE),
        ]
    )


def run_calculation_forecast() -> None:
    run_command([sys.executable, str(FORECAST_SCRIPT)])


def load_final_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[FINAL_DATA_SHEET]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


def formula_overview_metadata() -> dict[str, Any]:
    metadata = DEFAULT_METADATA.copy()
    if not FORMULA_OVERVIEW_FILE.exists():
        return metadata

    try:
        workbook = load_workbook(FORMULA_OVERVIEW_FILE, data_only=True, read_only=True)
    except PermissionError:
        return metadata

    try:
        target_key = normalize_key(SUPPLIER_PIPELINE_NAME)
        current_country: Any = None
        for sheet_name in FORMULA_OVERVIEW_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]
            headers = [clean_header(cell.value) for cell in worksheet[2]]
            header_map = {header: idx for idx, header in enumerate(headers)}
            if "Pricing Sheet" not in header_map:
                continue

            for row in worksheet.iter_rows(min_row=3, values_only=True):
                country_idx = header_map.get("Countries")
                if country_idx is not None and row[country_idx]:
                    current_country = clean_text(row[country_idx])

                pricing_sheet = clean_text(row[header_map["Pricing Sheet"]])
                if normalize_key(pricing_sheet) != target_key:
                    continue

                supplier_idx = header_map.get("Supplier")
                sourcing_idx = header_map.get("Sourcing Country")
                formula_idx = header_map.get("vPET Resin pricing adjustment formula")
                index_idx = header_map.get("Index")

                metadata["supplier"] = (
                    clean_text(row[supplier_idx])
                    if supplier_idx is not None and row[supplier_idx]
                    else metadata["supplier"]
                )
                metadata["destination_country"] = current_country or metadata["destination_country"]
                metadata["sourcing_country"] = (
                    clean_text(row[sourcing_idx])
                    if sourcing_idx is not None and row[sourcing_idx]
                    else metadata["sourcing_country"]
                )
                metadata["tlc_formula"] = (
                    clean_text(row[formula_idx])
                    if formula_idx is not None and row[formula_idx]
                    else metadata["tlc_formula"]
                )
                metadata["resin_index_type"] = (
                    clean_index(row[index_idx])
                    if index_idx is not None and row[index_idx]
                    else metadata["resin_index_type"]
                )
                metadata["formula_overview_sheet"] = sheet_name
                return metadata

        return metadata
    finally:
        workbook.close()


def source_resin_index_type(final_rows: list[dict[str, Any]]) -> Any:
    for row in final_rows:
        mapping_column = str(row.get("Mapping Columns") or "").strip()
        if mapping_column == "Resin Index vPET" and row.get("metric_en"):
            return clean_text(row.get("metric_en"))
    return None


def build_standardized_rows(
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in final_rows:
        rows.append(
            {
                "Source File ": Path(str(row.get("source_file") or SOURCE_FILE)).name,
                "Supplier Name": metadata["supplier"],
                "Destination Country": metadata["destination_country"],
                "Time_Period ": row.get("time_period"),
                "Raw Cost Breakdown": clean_text(row.get("metric_en")),
                "Resin Index Type": metadata["resin_index_type"],
                "Mapping Columns": clean_text(row.get("Mapping Columns")),
                "Value ": row.get("value"),
                "TLC Formula": metadata["tlc_formula"],
            }
        )
    return rows


def build_front_end_rows(
    standardized_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for standardized_row, final_row in zip(standardized_rows, final_rows):
        rows.append(
            {
                "Data Type": "Actual",
                **standardized_row,
                "Time Period Year": final_row.get("time_period_year"),
                "Time Period Month": final_row.get("time_period_month"),
                "Forecast Resin Index Type": None,
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            row.get("Destination Country") or "",
            row.get("Supplier Name") or "",
            row.get("Time Period Year") or 0,
            row.get("Time Period Month") or 0,
            row.get("Mapping Columns") or "",
            row.get("Raw Cost Breakdown") or "",
        ),
    )


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            70,
        )


def write_excel(
    path: Path,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    style_sheet(worksheet)

    metadata_sheet = workbook.create_sheet("run_metadata")
    metadata_sheet.append(["field", "value"])
    for key, value in metadata.items():
        metadata_sheet.append([key, value])
    style_sheet(metadata_sheet)
    metadata_sheet.column_dimensions["A"].width = 32
    metadata_sheet.column_dimensions["B"].width = 90
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def unique_nonblank(rows: list[dict[str, Any]], column: str) -> list[Any]:
    values: list[Any] = []
    for row in rows:
        value = row.get(column)
        if value in (None, ""):
            continue
        if value not in values:
            values.append(value)
    return values


def write_validation_summary(
    final_rows: list[dict[str, Any]],
    standardized_rows: list[dict[str, Any]],
    front_end_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    mapped_rows = [
        row
        for row in final_rows
        if row.get("metric_en") not in (None, "") and row.get("Mapping Columns")
    ]
    unmapped_nonblank_rows = [
        row
        for row in final_rows
        if row.get("metric_en") not in (None, "") and not row.get("Mapping Columns")
    ]
    landed_price_rows = [
        row for row in final_rows if clean_text(row.get("metric_en")) == "Landed Price"
    ]
    periods = sorted(
        {
            (
                row.get("time_period_year"),
                row.get("time_period_month"),
                row.get("time_period"),
            )
            for row in final_rows
            if row.get("time_period_year") and row.get("time_period_month")
        }
    )
    summary = {
        "pipeline": SUPPLIER_PIPELINE_NAME,
        "source_file": str(SOURCE_FILE),
        "extracted_file": str(EXTRACTED_FILE),
        "standardized_file": str(STANDARDIZED_FILE),
        "front_end_excel_file": str(FRONT_END_EXCEL_FILE),
        "front_end_csv_file": str(FRONT_END_CSV_FILE),
        "supplier": metadata["supplier"],
        "destination_country": metadata["destination_country"],
        "resin_index_type": metadata["resin_index_type"],
        "source_resin_index_type": metadata.get("source_resin_index_type"),
        "tlc_formula": metadata["tlc_formula"],
        "final_rows": len(final_rows),
        "standardized_rows": len(standardized_rows),
        "front_end_rows": len(front_end_rows),
        "mapped_nonblank_metric_rows": len(mapped_rows),
        "unmapped_nonblank_metric_rows": len(unmapped_nonblank_rows),
        "blank_metric_rows": len(final_rows) - len(mapped_rows) - len(unmapped_nonblank_rows),
        "landed_price_rows": len(landed_price_rows),
        "first_period": periods[0][2] if periods else None,
        "last_period": periods[-1][2] if periods else None,
        "unique_periods": len(periods),
        "mapping_columns": unique_nonblank(final_rows, "Mapping Columns"),
        "raw_cost_breakdowns": unique_nonblank(standardized_rows, "Raw Cost Breakdown"),
    }
    VALIDATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_FILE.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def write_artifacts() -> dict[str, Any]:
    final_rows = load_final_rows(EXTRACTED_FILE)
    metadata = formula_overview_metadata()
    metadata["source_resin_index_type"] = source_resin_index_type(final_rows)
    metadata["source_file"] = str(SOURCE_FILE)
    metadata["extracted_file"] = str(EXTRACTED_FILE)
    metadata["run_timestamp"] = datetime.now().isoformat(timespec="seconds")

    standardized_rows = build_standardized_rows(final_rows, metadata)
    front_end_rows = build_front_end_rows(standardized_rows, final_rows)
    metadata["standardized_rows"] = len(standardized_rows)
    metadata["front_end_rows"] = len(front_end_rows)

    write_excel(
        STANDARDIZED_FILE,
        "final_standardized",
        STANDARDIZED_HEADERS,
        standardized_rows,
        metadata,
    )
    write_excel(
        FRONT_END_EXCEL_FILE,
        "front_end_standardized",
        FRONT_END_HEADERS,
        front_end_rows,
        metadata,
    )
    write_csv(FRONT_END_CSV_FILE, FRONT_END_HEADERS, front_end_rows)
    write_validation_summary(final_rows, standardized_rows, front_end_rows, metadata)
    return metadata


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Ecuador supplier pipeline using existing extraction and mapping scripts."
    )
    parser.add_argument(
        "--skip-extraction",
        action="store_true",
        help="Reuse the existing supplier-local extracted workbook.",
    )
    parser.add_argument(
        "--skip-mapping",
        action="store_true",
        help="Reuse the existing mapping columns in the supplier-local workbook.",
    )
    parser.add_argument(
        "--skip-calculation-forecast",
        action="store_true",
        help="Do not regenerate TLC validation and forecast/front-end actual+forecast artifacts.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for directory in (EXTRACTION_DIR, STANDARDIZED_DIR, FRONT_END_DIR, VALIDATION_DIR):
        directory.mkdir(parents=True, exist_ok=True)

    if not args.skip_extraction:
        run_extraction()
    if not args.skip_mapping:
        run_mapping()

    metadata = write_artifacts()
    if not args.skip_calculation_forecast:
        run_calculation_forecast()

    print(f"supplier={metadata['supplier']}")
    print(f"destination_country={metadata['destination_country']}")
    print(f"resin_index_type={metadata['resin_index_type']}")
    print(f"source_resin_index_type={metadata.get('source_resin_index_type')}")
    print(f"standardized_output={STANDARDIZED_FILE}")
    print(f"front_end_excel={FRONT_END_EXCEL_FILE}")
    print(f"front_end_csv={FRONT_END_CSV_FILE}")
    print(f"validation_summary={VALIDATION_FILE}")
    print(f"front_end_rows={metadata['front_end_rows']}")


if __name__ == "__main__":
    main()
