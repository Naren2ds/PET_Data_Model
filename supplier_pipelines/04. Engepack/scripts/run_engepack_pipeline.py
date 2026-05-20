from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PIPELINE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[3]

PIPELINE_NAME = "04. Engepack"
SOURCE_FILE = REPO_ROOT / "data" / "04. Engepack.xlsx"
EXTRACT_SCRIPT = REPO_ROOT / "scripts" / "extract_engepack_indices.py"
MAPPING_SCRIPT = REPO_ROOT / "scripts" / "data_mapping_column_engepack.py"
RESIN_FORECAST_CSV = REPO_ROOT / "Estimation" / "resin_index_forecast_table.csv"

ARTIFACTS_DIR = PIPELINE_DIR / "artifacts"
EXTRACTION_DIR = ARTIFACTS_DIR / "extraction"
STANDARDIZED_DIR = ARTIFACTS_DIR / "standardized"
CALCULATION_DIR = ARTIFACTS_DIR / "calculation"
FORECAST_DIR = ARTIFACTS_DIR / "forecast"
FRONT_END_DIR = ARTIFACTS_DIR / "front_end"
VALIDATION_DIR = ARTIFACTS_DIR / "validation"

EXTRACTED_FILE = EXTRACTION_DIR / "04. Engepack_indices_final.xlsx"
STANDARDIZED_FILE = STANDARDIZED_DIR / "04. Engepack_actual_standardized.xlsx"
TLC_VALIDATION_FILE = CALCULATION_DIR / "engepack_tlc_calculation_validation.xlsx"
FORECAST_TEMPLATE_FILE = FORECAST_DIR / "engepack_forecast_inputs_template.xlsx"
FORECAST_ESTIMATION_FILE = FORECAST_DIR / "engepack_2026_forward_estimation.xlsx"
FRONT_END_ACTUAL_XLSX = FRONT_END_DIR / "04. Engepack_front_end_standardized.xlsx"
FRONT_END_ACTUAL_CSV = FRONT_END_DIR / "04. Engepack_front_end_standardized.csv"
FRONT_END_ACTUAL_FORECAST_XLSX = (
    FRONT_END_DIR / "04. Engepack_actual_forecast_front_end_standardized.xlsx"
)
FRONT_END_ACTUAL_FORECAST_CSV = (
    FRONT_END_DIR / "04. Engepack_actual_forecast_front_end_standardized.csv"
)
VALIDATION_SUMMARY_FILE = VALIDATION_DIR / "validation_summary.json"

FINAL_DATA_SHEET = "final_data"
SUPPLIER = "Engepack"
DESTINATION_COUNTRY = "Brazil"
SOURCE_RESIN_INDEX_TYPE = "IHS (M-1)"
FORECAST_RESIN_INDEX_TYPE = "PET Bottle Grade FOB China Spot"
MONTHLY_FORECAST_SERIES_KEYWORD = "Forecast Apr-2026 Latest"

RESIN_METRIC = "IHS (M-1)"
FX_METRIC = "Fx (M-1)"
FREIGHT_METRIC = "Freight"
TLC_LOCAL_METRIC = "Total Landing Cost"
TLC_USD_METRIC = "Total Landing Cost $"

LATEST_TLC_FORMULA = (
    "Total Landing Cost $ = ((IHS (M-1) + 100 + Adjusted Freight) * (1 + 20%)) "
    "+ 125 + 8% * Adjusted Freight"
)
OLD_TLC_FORMULA = (
    "Total Landing Cost $ = (IHS (M-1) + Adjusted Freight) * "
    "(1 + Internalization Cost - Discount) + Other Cost"
)
FREIGHT_ADJUSTMENT_FORMULA = (
    "Adjusted Freight = IF(AVG Q-1 > 108, MAX(108, (AVG Q-1 - 108) * 0.85 + 90), "
    "IF(AVG Q-1 < 60, 60, AVG Q-1))"
)

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

STANDARDIZED_HEADERS = [
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]

FRONT_END_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def clean_header(value: Any) -> str:
    return str(clean_text(value) or "")


def numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def maybe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def period_from_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    return None


def period_label(period: date | None) -> str | None:
    if period is None:
        return None
    return f"{MONTH_LABELS[period.month]} {period.year}"


def add_months(value: date, months: int) -> date:
    month_number = value.month - 1 + months
    year = value.year + month_number // 12
    month = month_number % 12 + 1
    return date(year, month, 1)


def period_sort_key(row: dict[str, Any]) -> tuple[int, int]:
    return int(row.get("time_period_year") or 0), int(row.get("time_period_month") or 0)


def run_command(command: list[str]) -> None:
    print(" ".join(f'"{part}"' if " " in part else part for part in command))
    subprocess.run(command, cwd=REPO_ROOT, check=True)


def run_extraction() -> None:
    run_command(
        [
            sys.executable,
            str(EXTRACT_SCRIPT),
            "--file",
            str(SOURCE_FILE),
            "--output-dir",
            str(EXTRACTION_DIR),
            "--excel-output",
            str(EXTRACTED_FILE),
            "--write-csv",
        ]
    )


def run_mapping() -> None:
    run_command([sys.executable, str(MAPPING_SCRIPT), "--output-file", str(EXTRACTED_FILE)])


def read_final_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[FINAL_DATA_SHEET]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


def rows_by_metric_period(final_rows: list[dict[str, Any]]) -> dict[tuple[str, int, int], dict[str, Any]]:
    rows: dict[tuple[str, int, int], dict[str, Any]] = {}
    for row in final_rows:
        metric = clean_text(row.get("metric_name"))
        year = row.get("time_period_year")
        month = row.get("time_period_month")
        if metric and year and month:
            rows[(metric, int(year), int(month))] = row
    return rows


def adjusted_freight(avg_q1: float) -> float:
    if avg_q1 > 108:
        return max(108.0, (avg_q1 - 108.0) * 0.85 + 90.0)
    if avg_q1 < 60:
        return 60.0
    return avg_q1


def source_workbook_validation(final_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    final_lookup = rows_by_metric_period(final_rows)
    values_wb = load_workbook(SOURCE_FILE, data_only=True, read_only=False)
    formulas_wb = load_workbook(SOURCE_FILE, data_only=False, read_only=False)
    try:
        indices_values = values_wb[values_wb.sheetnames[0]]
        precos_values = values_wb[values_wb.sheetnames[1]]
        precos_formulas = formulas_wb[formulas_wb.sheetnames[1]]

        rows: list[dict[str, Any]] = []
        for precos_col in range(3, 43):
            source_period = period_from_value(precos_values.cell(1, precos_col).value)
            if source_period is None:
                continue

            indices_col = precos_col + 4
            year = source_period.year
            month = source_period.month
            source_tlc_local = maybe_float(precos_values.cell(8, precos_col).value)
            fx = maybe_float(indices_values.cell(5, indices_col).value)
            if source_tlc_local is None or fx in (None, 0):
                continue

            formula_text = str(precos_formulas.cell(8, precos_col).value or "")
            formula_family = (
                "latest_2025_contract"
                if "+100+" in formula_text and "+125+8%" in formula_text
                else "legacy_contract"
            )
            resin = numeric(indices_values.cell(3, indices_col).value)
            extracted_avg_q1 = maybe_float(indices_values.cell(23, indices_col).value)
            adjusted = numeric(precos_values.cell(6, precos_col).value)
            source_tlc_usd = source_tlc_local / fx

            if formula_family == "latest_2025_contract":
                resin_fixed_adder = 100.0
                internalization_rate = 0.20
                other_cost = 125.0
                freight_surcharge_rate = 0.08
                calculated_tlc_usd = (
                    (resin + resin_fixed_adder + adjusted) * (1 + internalization_rate)
                    + other_cost
                    + freight_surcharge_rate * adjusted
                )
                formula_used = LATEST_TLC_FORMULA
                discount = None
            else:
                discount = maybe_float(precos_values.cell(2, precos_col).value) or 0.0
                internalization_rate = numeric(precos_values.cell(13, 8).value)
                other_cost = numeric(precos_values.cell(13, 9).value)
                resin_fixed_adder = 0.0
                freight_surcharge_rate = 0.0
                calculated_tlc_usd = (
                    (resin + adjusted) * (1 + internalization_rate - discount)
                    + other_cost
                )
                formula_used = OLD_TLC_FORMULA

            final_tlc_usd_row = final_lookup.get((TLC_USD_METRIC, year, month), {})
            final_tlc_usd = maybe_float(final_tlc_usd_row.get("value"))
            source_vs_final_diff = (
                source_tlc_usd - final_tlc_usd
                if final_tlc_usd is not None
                else None
            )
            difference = calculated_tlc_usd - source_tlc_usd

            rows.append(
                {
                    "time_period": period_label(source_period),
                    "time_period_year": year,
                    "time_period_month": month,
                    "formula_family": formula_family,
                    "resin_index_type": SOURCE_RESIN_INDEX_TYPE,
                    "resin_price_index": resin,
                    "fx": fx,
                    "freight_avg_q1_extracted_same_period": extracted_avg_q1,
                    "adjusted_freight_used_by_tlc": adjusted,
                    "discount": discount,
                    "resin_fixed_adder": resin_fixed_adder,
                    "internalization_rate": internalization_rate,
                    "other_cost": other_cost,
                    "freight_surcharge_rate": freight_surcharge_rate,
                    "source_tlc_local": source_tlc_local,
                    "source_tlc_usd": source_tlc_usd,
                    "final_data_tlc_usd": final_tlc_usd,
                    "source_vs_final_data_difference": source_vs_final_diff,
                    "calculated_tlc_usd": calculated_tlc_usd,
                    "difference": difference,
                    "abs_difference": abs(difference),
                    "validation_status": "match" if abs(difference) <= 0.000001 else "check",
                    "source_formula": formula_text,
                    "formula_used": formula_used,
                    "freight_adjustment_formula": FREIGHT_ADJUSTMENT_FORMULA,
                }
            )
        return rows
    finally:
        values_wb.close()
        formulas_wb.close()


def standardized_rows(
    final_rows: list[dict[str, Any]],
    validation_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    formula_by_period = {
        (row["time_period_year"], row["time_period_month"]): row["formula_used"]
        for row in validation_rows
    }
    rows: list[dict[str, Any]] = []
    for row in final_rows:
        period_key = (row.get("time_period_year"), row.get("time_period_month"))
        rows.append(
            {
                "Source File ": Path(str(row.get("source_file") or SOURCE_FILE)).name,
                "Supplier Name": SUPPLIER,
                "Destination Country": DESTINATION_COUNTRY,
                "Time_Period ": row.get("time_period"),
                "Time Period Year": row.get("time_period_year"),
                "Time Period Month": row.get("time_period_month"),
                "Location": None,
                "Raw Cost Breakdown": clean_text(row.get("metric_name")),
                "Resin Index Type": SOURCE_RESIN_INDEX_TYPE,
                "Mapping Columns": clean_text(row.get("Mapping Columns")),
                "Column Required for Calculation": clean_text(
                    row.get("Column Required for Calculation")
                ),
                "Value ": row.get("value"),
                "TLC Formula": formula_by_period.get(period_key, LATEST_TLC_FORMULA),
            }
        )
    return rows


def front_end_actual_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sort_front_end_rows(
        [{"Data Type": "Actual", **row, "Forecast Resin Index Type": None} for row in rows]
    )


def read_resin_forecast(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        return list(csv.DictReader(csv_file))


def forecast_period_value(
    forecast_rows: list[dict[str, Any]],
    source_period: date,
) -> dict[str, Any] | None:
    candidates = [
        row
        for row in forecast_rows
        if clean_text(row.get("resin_index_type")) == FORECAST_RESIN_INDEX_TYPE
        and int(float(row.get("time_period_year") or 0)) == source_period.year
        and int(float(row.get("time_period_month") or 0)) == source_period.month
    ]
    if not candidates:
        return None

    monthly = [
        row
        for row in candidates
        if MONTHLY_FORECAST_SERIES_KEYWORD.lower()
        in str(row.get("forecast_series") or "").lower()
    ]
    if monthly:
        return sorted(monthly, key=lambda row: str(row.get("forecast_date") or ""))[-1]

    values = [numeric(row.get("value")) for row in candidates if row.get("value") not in (None, "")]
    if not values:
        return None
    latest = sorted(candidates, key=lambda row: str(row.get("forecast_date") or ""))[-1].copy()
    latest["value"] = sum(values) / len(values)
    latest["forecast_series"] = f"Monthly average fallback from {len(values)} forecast rows"
    latest["value_selection_method"] = "monthly_average_fallback"
    return latest


def latest_valid_validation_row(validation: list[dict[str, Any]]) -> dict[str, Any]:
    matched = [row for row in validation if row["validation_status"] == "match"]
    if not matched:
        raise ValueError("No matching Engepack TLC validation rows found.")
    return sorted(matched, key=period_sort_key)[-1]


def forecast_months_after_latest(latest_row: dict[str, Any], end_year: int = 2026) -> list[date]:
    latest = date(int(latest_row["time_period_year"]), int(latest_row["time_period_month"]), 1)
    start = add_months(latest, 1)
    months: list[date] = []
    current = start
    while current.year <= end_year:
        months.append(current)
        current = add_months(current, 1)
    return months


def build_forecast_inputs(
    validation: list[dict[str, Any]],
    forecast_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    latest = latest_valid_validation_row(validation)
    latest_avg_q1 = numeric(latest.get("adjusted_freight_used_by_tlc"))
    rows: list[dict[str, Any]] = []
    for target_period in forecast_months_after_latest(latest):
        resin_source_period = add_months(target_period, -1)
        resin_forecast = forecast_period_value(forecast_rows, resin_source_period)
        resin_value = numeric(resin_forecast.get("value")) if resin_forecast else 0.0
        rows.append(
            {
                "time_period": period_label(target_period),
                "time_period_year": target_period.year,
                "time_period_month": target_period.month,
                "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
                "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
                "resin_forecast_source_period": period_label(resin_source_period),
                "resin_forecast_date": resin_forecast.get("forecast_date") if resin_forecast else None,
                "resin_forecast_series": resin_forecast.get("forecast_series") if resin_forecast else None,
                "resin_price_index": resin_value,
                "freight_avg_q1": latest_avg_q1,
                "adjusted_freight_override": None,
                "adjusted_freight_used": adjusted_freight(latest_avg_q1),
                "resin_fixed_adder": 100.0,
                "internalization_rate": 0.20,
                "other_cost": 125.0,
                "freight_surcharge_rate": 0.08,
                "latest_actual_period_used": latest["time_period"],
                "notes": "Update resin/freight inputs if business provides new future values.",
            }
        )
    return rows


def build_forecast_estimates(
    forecast_inputs: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    estimates: list[dict[str, Any]] = []
    front_end_rows: list[dict[str, Any]] = []
    for row in forecast_inputs:
        resin = numeric(row["resin_price_index"])
        avg_q1 = numeric(row["freight_avg_q1"])
        override = maybe_float(row.get("adjusted_freight_override"))
        adjusted = override if override is not None else adjusted_freight(avg_q1)
        resin_fixed = numeric(row["resin_fixed_adder"])
        internalization = numeric(row["internalization_rate"])
        other = numeric(row["other_cost"])
        freight_surcharge_rate = numeric(row["freight_surcharge_rate"])
        freight_surcharge = freight_surcharge_rate * adjusted
        tlc_usd = (resin + resin_fixed + adjusted) * (1 + internalization) + other + freight_surcharge
        estimate = {
            **row,
            "adjusted_freight_used": adjusted,
            "freight_surcharge": freight_surcharge,
            "total_landing_cost_usd_forecast": tlc_usd,
            "tlc_formula": LATEST_TLC_FORMULA,
        }
        estimates.append(estimate)

        component_specs = [
            (RESIN_METRIC, "Resin Index vPET", "Yes", resin),
            ("Freight AVG Q-1", "Freight", "Yes", avg_q1),
            ("Freight Adjusted for TLC", "Freight", "Yes", adjusted),
            ("Resin Fixed Adder", "Others", "Yes", resin_fixed),
            ("Internalization Cost Rate", "Others", "Yes", internalization),
            ("Other Cost", "Others", "Yes", other),
            ("Freight Surcharge 8%", "Others", "Yes", freight_surcharge),
            (TLC_USD_METRIC, "Total Landing Cost", "Yes", tlc_usd),
        ]
        for raw_cost, mapping_column, required, value in component_specs:
            front_end_rows.append(
                {
                    "Data Type": "Forecast",
                    "Source File ": RESIN_FORECAST_CSV.name,
                    "Supplier Name": SUPPLIER,
                    "Destination Country": DESTINATION_COUNTRY,
                    "Time_Period ": row["time_period"],
                    "Time Period Year": row["time_period_year"],
                    "Time Period Month": row["time_period_month"],
                    "Location": None,
                    "Raw Cost Breakdown": raw_cost,
                    "Resin Index Type": SOURCE_RESIN_INDEX_TYPE,
                    "Forecast Resin Index Type": FORECAST_RESIN_INDEX_TYPE,
                    "Mapping Columns": mapping_column,
                    "Column Required for Calculation": required,
                    "Value ": value,
                    "TLC Formula": LATEST_TLC_FORMULA,
                }
            )
    return estimates, front_end_rows


def formula_catalog_rows(latest: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "component": RESIN_METRIC,
            "classification": "variable_forecast_input",
            "default_value": None,
            "formula_or_source": "Forecast table source period is target month - 1.",
            "notes": f"Source resin index type: {SOURCE_RESIN_INDEX_TYPE}.",
        },
        {
            "component": "Freight AVG Q-1",
            "classification": "variable_forecast_input",
            "default_value": latest.get("adjusted_freight_used_by_tlc"),
            "formula_or_source": "Carried from latest actual unless updated.",
            "notes": "The extracted Freight row is AVG Q-1; TLC uses the adjusted freight row.",
        },
        {
            "component": "Freight Adjusted for TLC",
            "classification": "formula_or_override",
            "default_value": latest.get("adjusted_freight_used_by_tlc"),
            "formula_or_source": FREIGHT_ADJUSTMENT_FORMULA,
            "notes": "This is the freight component used by the Total Landing Cost formula.",
        },
        {
            "component": "Resin Fixed Adder",
            "classification": "constant",
            "default_value": 100.0,
            "formula_or_source": "Latest Engepack source formula.",
            "notes": None,
        },
        {
            "component": "Internalization Cost Rate",
            "classification": "constant",
            "default_value": 0.20,
            "formula_or_source": "Latest Engepack source formula.",
            "notes": None,
        },
        {
            "component": "Other Cost",
            "classification": "constant",
            "default_value": 125.0,
            "formula_or_source": "Latest Engepack source formula.",
            "notes": None,
        },
        {
            "component": "Freight Surcharge 8%",
            "classification": "formula",
            "default_value": 0.08,
            "formula_or_source": "Adjusted Freight * 8%.",
            "notes": None,
        },
        {
            "component": TLC_USD_METRIC,
            "classification": "calculated_output",
            "default_value": None,
            "formula_or_source": LATEST_TLC_FORMULA,
            "notes": "Final Engepack TLC output is in USD.",
        },
    ]


def sort_front_end_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("Destination Country") or "",
            row.get("Supplier Name") or "",
            int(row.get("Time Period Year") or 0),
            int(row.get("Time Period Month") or 0),
            0 if row.get("Data Type") == "Actual" else 1,
            row.get("Mapping Columns") or "",
            row.get("Raw Cost Breakdown") or "",
        ),
    )


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> Any:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"
    style_sheet(worksheet)
    return worksheet


def write_workbook(
    path: Path,
    sheets: list[tuple[str, list[str], list[dict[str, Any]]]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers, rows in sheets:
        write_sheet(workbook, sheet_name, headers, rows)
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_forecast_template(
    path: Path,
    forecast_inputs: list[dict[str, Any]],
    formula_catalog: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    input_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_resin_index_type",
        "forecast_resin_index_type",
        "resin_forecast_source_period",
        "resin_forecast_date",
        "resin_forecast_series",
        "resin_price_index",
        "freight_avg_q1",
        "adjusted_freight_override",
        "adjusted_freight_used",
        "resin_fixed_adder",
        "internalization_rate",
        "other_cost",
        "freight_surcharge_rate",
        "total_landing_cost_usd",
        "latest_actual_period_used",
        "notes",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "forecast_inputs"
    worksheet.append(input_headers)
    for row in forecast_inputs:
        worksheet.append([row.get(header) for header in input_headers])

    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    formula_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet[f"L{row_idx}"] = (
            f'=IF(NOT(ISBLANK(K{row_idx})),K{row_idx},'
            f'IF(J{row_idx}>108,MAX(108,(J{row_idx}-108)*0.85+90),'
            f'IF(J{row_idx}<60,60,J{row_idx})))'
        )
        worksheet[f"Q{row_idx}"] = (
            f'=((I{row_idx}+M{row_idx}+L{row_idx})*(1+N{row_idx}))+'
            f'O{row_idx}+(P{row_idx}*L{row_idx})'
        )
        for col_idx in [9, 10, 11, 13, 14, 15, 16]:
            worksheet.cell(row_idx, col_idx).fill = input_fill
        for col_idx in [12, 17]:
            worksheet.cell(row_idx, col_idx).fill = formula_fill
    style_sheet(worksheet)

    write_sheet(
        workbook,
        "formula_catalog",
        ["component", "classification", "default_value", "formula_or_source", "notes"],
        formula_catalog,
    )
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def unique_nonblank(rows: list[dict[str, Any]], column: str) -> list[Any]:
    values: list[Any] = []
    for row in rows:
        value = row.get(column)
        if value in (None, ""):
            continue
        if value not in values:
            values.append(value)
    return values


def write_outputs(
    final_rows: list[dict[str, Any]],
    actual_standardized: list[dict[str, Any]],
    actual_front_end: list[dict[str, Any]],
    validation: list[dict[str, Any]],
    forecast_inputs: list[dict[str, Any]],
    forecast_estimates: list[dict[str, Any]],
    forecast_front_end: list[dict[str, Any]],
) -> None:
    latest = latest_valid_validation_row(validation)
    formula_catalog = formula_catalog_rows(latest)
    actual_forecast = sort_front_end_rows(actual_front_end + forecast_front_end)
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "pipeline": PIPELINE_NAME,
        "source_file": str(SOURCE_FILE),
        "extracted_file": str(EXTRACTED_FILE),
        "supplier": SUPPLIER,
        "destination_country": DESTINATION_COUNTRY,
        "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
        "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
        "tlc_formula": LATEST_TLC_FORMULA,
        "freight_adjustment_formula": FREIGHT_ADJUSTMENT_FORMULA,
        "note": "Extracted Freight is AVG Q-1; TLC validation and forecast use adjusted freight from source row 26.",
        "final_rows": len(final_rows),
        "actual_standardized_rows": len(actual_standardized),
        "tlc_validation_rows": len(validation),
        "forecast_input_rows": len(forecast_inputs),
        "forecast_estimate_rows": len(forecast_estimates),
        "forecast_front_end_rows": len(forecast_front_end),
        "actual_forecast_front_end_rows": len(actual_forecast),
    }
    validation_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "formula_family",
        "resin_index_type",
        "resin_price_index",
        "fx",
        "freight_avg_q1_extracted_same_period",
        "adjusted_freight_used_by_tlc",
        "discount",
        "resin_fixed_adder",
        "internalization_rate",
        "other_cost",
        "freight_surcharge_rate",
        "source_tlc_local",
        "source_tlc_usd",
        "final_data_tlc_usd",
        "source_vs_final_data_difference",
        "calculated_tlc_usd",
        "difference",
        "abs_difference",
        "validation_status",
        "source_formula",
        "formula_used",
        "freight_adjustment_formula",
    ]
    forecast_estimate_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_resin_index_type",
        "forecast_resin_index_type",
        "resin_forecast_source_period",
        "resin_forecast_date",
        "resin_forecast_series",
        "resin_price_index",
        "freight_avg_q1",
        "adjusted_freight_override",
        "adjusted_freight_used",
        "resin_fixed_adder",
        "internalization_rate",
        "other_cost",
        "freight_surcharge_rate",
        "freight_surcharge",
        "total_landing_cost_usd_forecast",
        "latest_actual_period_used",
        "tlc_formula",
        "notes",
    ]

    write_workbook(
        STANDARDIZED_FILE,
        [("final_standardized", STANDARDIZED_HEADERS, actual_standardized)],
        metadata,
    )
    write_workbook(
        TLC_VALIDATION_FILE,
        [
            ("tlc_validation", validation_headers, validation),
            ("formula_catalog", ["component", "classification", "default_value", "formula_or_source", "notes"], formula_catalog),
        ],
        metadata,
    )
    write_forecast_template(FORECAST_TEMPLATE_FILE, forecast_inputs, formula_catalog, metadata)
    write_workbook(
        FORECAST_ESTIMATION_FILE,
        [
            ("engepack_2026_estimate", forecast_estimate_headers, forecast_estimates),
            ("standardized_rows", FRONT_END_HEADERS, forecast_front_end),
        ],
        metadata,
    )
    write_workbook(
        FRONT_END_ACTUAL_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, actual_front_end)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_CSV, FRONT_END_HEADERS, actual_front_end)
    write_workbook(
        FRONT_END_ACTUAL_FORECAST_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, actual_forecast)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_FORECAST_CSV, FRONT_END_HEADERS, actual_forecast)

    periods = sorted(
        {
            (
                int(row.get("Time Period Year") or 0),
                int(row.get("Time Period Month") or 0),
                row.get("Time_Period "),
            )
            for row in actual_front_end
            if row.get("Time Period Year") and row.get("Time Period Month")
        }
    )
    summary = {
        **metadata,
        "validation_statuses": sorted({row["validation_status"] for row in validation}),
        "max_abs_difference": max((row["abs_difference"] for row in validation), default=0),
        "max_abs_source_vs_final_data_difference": max(
            (
                abs(row["source_vs_final_data_difference"])
                for row in validation
                if row["source_vs_final_data_difference"] is not None
            ),
            default=0,
        ),
        "first_actual_period": periods[0][2] if periods else None,
        "last_actual_period": periods[-1][2] if periods else None,
        "mapping_columns": unique_nonblank(final_rows, "Mapping Columns"),
        "raw_cost_breakdowns": unique_nonblank(actual_standardized, "Raw Cost Breakdown"),
        "front_end_actual_forecast_csv": str(FRONT_END_ACTUAL_FORECAST_CSV),
    }
    VALIDATION_SUMMARY_FILE.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_SUMMARY_FILE.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Engepack supplier pipeline from extraction through forecast artifacts."
    )
    parser.add_argument("--skip-extraction", action="store_true")
    parser.add_argument("--skip-mapping", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for directory in (
        EXTRACTION_DIR,
        STANDARDIZED_DIR,
        CALCULATION_DIR,
        FORECAST_DIR,
        FRONT_END_DIR,
        VALIDATION_DIR,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    if not args.skip_extraction:
        run_extraction()
    if not args.skip_mapping:
        run_mapping()

    final_rows = read_final_rows(EXTRACTED_FILE)
    validation = source_workbook_validation(final_rows)
    actual_standardized = standardized_rows(final_rows, validation)
    actual_front_end = front_end_actual_rows(actual_standardized)
    resin_forecast = read_resin_forecast(RESIN_FORECAST_CSV)
    forecast_inputs = build_forecast_inputs(validation, resin_forecast)
    forecast_estimates, forecast_front_end = build_forecast_estimates(forecast_inputs)
    write_outputs(
        final_rows,
        actual_standardized,
        actual_front_end,
        validation,
        forecast_inputs,
        forecast_estimates,
        forecast_front_end,
    )

    print(f"supplier={SUPPLIER}")
    print(f"destination_country={DESTINATION_COUNTRY}")
    print(f"source_resin_index_type={SOURCE_RESIN_INDEX_TYPE}")
    print(f"forecast_resin_index_type={FORECAST_RESIN_INDEX_TYPE}")
    print(f"final_rows={len(final_rows)}")
    print(f"validation_rows={len(validation)}")
    print(f"forecast_months={len(forecast_estimates)}")
    print(f"forecast_template={FORECAST_TEMPLATE_FILE}")
    print(f"forecast_estimation={FORECAST_ESTIMATION_FILE}")
    print(f"front_end_actual_forecast_csv={FRONT_END_ACTUAL_FORECAST_CSV}")
    print(f"validation_summary={VALIDATION_SUMMARY_FILE}")


if __name__ == "__main__":
    main()
