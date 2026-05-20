from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PIPELINE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[3]

PIPELINE_NAME = "04. Valgroup"
SOURCE_FILE = REPO_ROOT / "data" / "04. Valgroup.xlsx"
LEGACY_EXTRACT_SCRIPT = REPO_ROOT / "scripts" / "extract_valgroup_months.py"
LEGACY_MAPPING_SCRIPT = REPO_ROOT / "scripts" / "data_mapping_column_valgroup.py"
MAPPING_FILE = REPO_ROOT / "mapping files" / "Mapping_Columns.xlsx"
INDEX_REFERENCE_CSV = REPO_ROOT / "Index Fprecast" / "icis_resin_index_reference_table.csv"

ARTIFACTS_DIR = PIPELINE_DIR / "artifacts"
EXTRACTION_DIR = ARTIFACTS_DIR / "extraction"
STANDARDIZED_DIR = ARTIFACTS_DIR / "standardized"
CALCULATION_DIR = ARTIFACTS_DIR / "calculation"
FORECAST_DIR = ARTIFACTS_DIR / "forecast"
FRONT_END_DIR = ARTIFACTS_DIR / "front_end"
VALIDATION_DIR = ARTIFACTS_DIR / "validation"

EXTRACTED_FILE = EXTRACTION_DIR / "04. Valgroup_months_full_final.xlsx"
LEGACY_EXTRACTED_FILE = EXTRACTION_DIR / "04. Valgroup_months_final.xlsx"
STANDARDIZED_FILE = STANDARDIZED_DIR / "04. Valgroup_actual_standardized.xlsx"
TLC_VALIDATION_FILE = CALCULATION_DIR / "valgroup_tlc_calculation_validation.xlsx"
FORECAST_TEMPLATE_FILE = FORECAST_DIR / "valgroup_forecast_inputs_template.xlsx"
FORECAST_ESTIMATION_FILE = FORECAST_DIR / "valgroup_2026_forward_estimation.xlsx"
FRONT_END_ACTUAL_XLSX = FRONT_END_DIR / "04. Valgroup_front_end_standardized.xlsx"
FRONT_END_ACTUAL_CSV = FRONT_END_DIR / "04. Valgroup_front_end_standardized.csv"
FRONT_END_ACTUAL_FORECAST_XLSX = (
    FRONT_END_DIR / "04. Valgroup_actual_forecast_front_end_standardized.xlsx"
)
FRONT_END_ACTUAL_FORECAST_CSV = (
    FRONT_END_DIR / "04. Valgroup_actual_forecast_front_end_standardized.csv"
)
VALIDATION_SUMMARY_FILE = VALIDATION_DIR / "validation_summary.json"

SUPPLIER = "Valgroup"
DESTINATION_COUNTRY = "Brazil"
PRIMARY_INDEX_TYPE = "ICIS Asia SE Low"
MID_INDEX_TYPE = "ICIS China Mid"
SOURCE_LOW_LABEL = "ICIS Asia SE Low (n-1)"
SOURCE_MID_LABEL = "ICIS Asia 5R MID (n-1)"
SOURCE_RESIN_INDEX_TYPE = "ICIS Asia SE Low (n-1) with Mid guardrail"

TLC_USD_METRIC = "Total V-PET USD/ton"
TLC_BRL_METRIC = "Total V-PET BRL/ton"
TLC_FORMULA = (
    "Total V-PET USD/ton = (((Resin with assumptions * (1 - Discount)) + Freight) "
    "* (1 + Importation + Import Tax)) + Surcharge + Indorama Discount"
)
RESIN_ASSUMPTION_FORMULA = (
    "Resin with assumptions = IF(Low > Mid, Mid, IF(Mid - Low > 75, Mid - 75, Low))"
)

METRIC_ROWS = range(3, 13)
TLC_BRL_ROW = 13
MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}
MONTHS = {
    "JAN": (1, "January"),
    "FEV": (2, "February"),
    "FEB": (2, "February"),
    "MAR": (3, "March"),
    "ABR": (4, "April"),
    "APR": (4, "April"),
    "MAI": (5, "May"),
    "MAY": (5, "May"),
    "JUN": (6, "June"),
    "JUL": (7, "July"),
    "AGO": (8, "August"),
    "AUG": (8, "August"),
    "SET": (9, "September"),
    "SEP": (9, "September"),
    "OUT": (10, "October"),
    "OCT": (10, "October"),
    "NOV": (11, "November"),
    "DEZ": (12, "December"),
    "DEC": (12, "December"),
}

TRANSLATIONS = {
    "Icis Asia SE Low (n-1)": SOURCE_LOW_LABEL,
    "Icis Asia 5R MID (n-1)": SOURCE_MID_LABEL,
    "Resina c/ premissas": "Resin with assumptions",
    "Desconto": "Discount",
    "Drewry (t-1) com desconto": "Drewry (t-1) with discount",
    "Internacao": "Importation",
    "Internação": "Importation",
    "Imposto Internacao": "Import Tax",
    "Imposto Internação": "Import Tax",
    "Sobretaxa": "Surcharge",
    "Desconto Indorama": "Indorama Discount",
    "Total V-PET USD/ton": TLC_USD_METRIC,
    "Total V-PET USD/ton ": TLC_USD_METRIC,
    "Total V-PET R$/ton": TLC_BRL_METRIC,
    "Total V-PET R$/ton ": TLC_BRL_METRIC,
}

STANDARDIZED_HEADERS = [
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]

FRONT_END_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def clean_header(value: Any) -> str:
    return str(clean_text(value) or "")


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or "")).replace("\xa0", " ").strip().lower()
    return re.sub(r"[^a-z0-9$%]+", " ", text).strip()


def translate(value: Any) -> Any:
    text = clean_text(value)
    if not isinstance(text, str):
        return text
    normalized = normalize_key(text)
    for original, translated in TRANSLATIONS.items():
        if normalize_key(original) == normalized:
            return translated
    return text


def numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def maybe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def parse_month_sheet(sheet_name: str) -> dict[str, Any] | None:
    match = re.fullmatch(r"\s*([A-Za-z]{3})\.?(\d{2}|\d{4})\s*", sheet_name)
    if not match:
        return None
    month_key = match.group(1).upper()
    if month_key not in MONTHS:
        return None
    year = int(match.group(2))
    if year < 100:
        year += 2000
    month_number, month_english = MONTHS[month_key]
    return {
        "month_original": match.group(1).upper(),
        "month_english": month_english,
        "time_period": f"{month_english} {year}",
        "time_period_year": year,
        "time_period_month": month_number,
    }


def period_date(year: Any, month: Any) -> date:
    return date(int(year), int(month), 1)


def add_months(value: date, months: int) -> date:
    month_number = value.month - 1 + months
    year = value.year + month_number // 12
    month = month_number % 12 + 1
    return date(year, month, 1)


def period_label(value: date) -> str:
    return f"{MONTH_LABELS[value.month]} {value.year}"


def period_sort_key(row: dict[str, Any]) -> tuple[int, int, str]:
    return (
        int(row.get("time_period_year") or row.get("Time Period Year") or 0),
        int(row.get("time_period_month") or row.get("Time Period Month") or 0),
        str(row.get("product") or row.get("Location") or ""),
    )


def run_command(command: list[str]) -> None:
    print(" ".join(f'"{part}"' if " " in part else part for part in command))
    subprocess.run(command, cwd=REPO_ROOT, check=True)


def run_legacy_extraction() -> None:
    run_command(
        [
            sys.executable,
            str(LEGACY_EXTRACT_SCRIPT),
            "--file",
            str(SOURCE_FILE),
            "--output-dir",
            str(EXTRACTION_DIR),
            "--excel-output",
            str(LEGACY_EXTRACTED_FILE),
            "--write-csv",
        ]
    )


def run_legacy_mapping() -> None:
    run_command([sys.executable, str(LEGACY_MAPPING_SCRIPT), "--output-file", str(LEGACY_EXTRACTED_FILE)])


def load_mapping() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(MAPPING_FILE, data_only=True, read_only=True)
    try:
        worksheet = workbook["04. Valgroup"]
        headers = [clean_header(cell.value).rstrip() for cell in worksheet[1]]
        lookup_idx = headers.index("metric_label_english")
        mapping_col_idx = next(
            idx for idx, header in enumerate(headers) if normalize_key(header) == "mapping columns"
        )
        calc_idx = next(
            idx for idx, header in enumerate(headers) if normalize_key(header) == "column required for calculation"
        )
        mapping: dict[str, dict[str, Any]] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            lookup_value = row[lookup_idx]
            if lookup_value is None:
                continue
            mapping[normalize_key(lookup_value)] = {
                "Mapping Columns": clean_text(row[mapping_col_idx]),
                "Column Required for Calculation": clean_text(row[calc_idx]),
            }
        return mapping
    finally:
        workbook.close()


def product_block(worksheet: Any) -> tuple[int, list[int], int]:
    label_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        if normalize_key(worksheet.cell(2, col_idx).value) == "resina virgem":
            label_col = col_idx
            break
    if label_col is None:
        raise ValueError(f"Cannot find RESINA VIRGEM label in sheet {worksheet.title}")

    product_cols: list[int] = []
    for col_idx in range(label_col + 1, worksheet.max_column + 1):
        product = clean_text(worksheet.cell(2, col_idx).value)
        if product in (None, ""):
            break
        product_cols.append(col_idx)
    if not product_cols:
        raise ValueError(f"Cannot find Valgroup product columns in sheet {worksheet.title}")
    return label_col, product_cols, product_cols[-1]


def extract_full_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping = load_mapping()
    values_wb = load_workbook(SOURCE_FILE, data_only=True, read_only=False)
    formulas_wb = load_workbook(SOURCE_FILE, data_only=False, read_only=False)
    try:
        final_rows: list[dict[str, Any]] = []
        raw_rows: list[dict[str, Any]] = []
        for sheet_name in values_wb.sheetnames:
            period = parse_month_sheet(sheet_name)
            if not period:
                continue
            value_ws = values_wb[sheet_name]
            formula_ws = formulas_wb[sheet_name]
            label_col, product_cols, ptax_col = product_block(value_ws)
            source_range = (
                f"{get_column_letter(label_col)}2:{get_column_letter(product_cols[-1])}{TLC_BRL_ROW}"
            )
            for row_idx in range(2, TLC_BRL_ROW + 1):
                raw = {
                    "source_sheet": sheet_name,
                    "source_range": source_range,
                    "source_row": row_idx,
                    "label": clean_text(value_ws.cell(row_idx, label_col).value),
                }
                for product_col in product_cols:
                    raw[get_column_letter(product_col)] = value_ws.cell(row_idx, product_col).value
                raw_rows.append(raw)

            for product_col in product_cols:
                product = clean_text(value_ws.cell(2, product_col).value)
                ptax = maybe_float(value_ws.cell(1, ptax_col).value)
                for row_idx in METRIC_ROWS:
                    metric_original = clean_text(value_ws.cell(row_idx, label_col).value)
                    metric_english = translate(metric_original)
                    mapping_values = mapping.get(normalize_key(metric_english), {})
                    col_letter = get_column_letter(product_col)
                    row = {
                        "source_file": str(SOURCE_FILE),
                        "source_sheet": sheet_name,
                        "source_range": source_range,
                        "source_cell": f"{col_letter}{row_idx}",
                        "metric_row": row_idx,
                        **period,
                        "product": product,
                        "ptax": ptax,
                        "metric_label_source_cell": f"{get_column_letter(label_col)}{row_idx}",
                        "metric_label_original": metric_original,
                        "metric_label_english": metric_english,
                        "Mapping Columns": mapping_values.get("Mapping Columns"),
                        "Column Required for Calculation": mapping_values.get("Column Required for Calculation"),
                        "value": value_ws.cell(row_idx, product_col).value,
                        "formula": formula_ws.cell(row_idx, product_col).value,
                    }
                    final_rows.append(row)
        return raw_rows, final_rows
    finally:
        values_wb.close()
        formulas_wb.close()


def rows_by_period_product(final_rows: list[dict[str, Any]]) -> dict[tuple[int, int, str], list[dict[str, Any]]]:
    groups: dict[tuple[int, int, str], list[dict[str, Any]]] = {}
    for row in final_rows:
        groups.setdefault(
            (
                int(row["time_period_year"]),
                int(row["time_period_month"]),
                str(row["product"]),
            ),
            [],
        ).append(row)
    return groups


def row_for_metric(rows: list[dict[str, Any]], metric_name: str) -> dict[str, Any] | None:
    target = normalize_key(metric_name)
    for row in rows:
        if normalize_key(row.get("metric_label_english")) == target:
            return row
    return None


def value_for(rows: list[dict[str, Any]], metric_name: str) -> float:
    row = row_for_metric(rows, metric_name)
    return numeric(row.get("value") if row else None)


def resin_with_assumptions(low: float, mid: float) -> float:
    if low > mid:
        return mid
    if mid - low > 75:
        return mid - 75
    return low


def calculate_tlc_usd(
    resin: float,
    discount: float,
    freight: float,
    importation: float,
    import_tax: float,
    surcharge: float,
    indorama_discount: float,
) -> float:
    return (((resin * (1 - discount)) + freight) * (1 + importation + import_tax)) + surcharge + indorama_discount


def read_index_reference() -> dict[tuple[int, int, str], dict[str, Any]]:
    if not INDEX_REFERENCE_CSV.exists():
        return {}
    with INDEX_REFERENCE_CSV.open(newline="", encoding="utf-8-sig") as csv_file:
        rows = list(csv.DictReader(csv_file))
    reference: dict[tuple[int, int, str], dict[str, Any]] = {}
    for row in rows:
        try:
            year = int(float(row.get("time_period_year") or 0))
            month = int(float(row.get("time_period_month") or 0))
        except ValueError:
            continue
        index_type = clean_text(row.get("resin_index_type"))
        if index_type:
            reference[(year, month, index_type)] = row
    return reference


def reference_value(
    reference: dict[tuple[int, int, str], dict[str, Any]],
    source_period: date,
    index_type: str,
) -> tuple[float | None, dict[str, Any] | None]:
    row = reference.get((source_period.year, source_period.month, index_type))
    if row is None:
        return None, None
    return maybe_float(row.get("value")), row


def validation_rows(final_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups = rows_by_period_product(final_rows)
    reference = read_index_reference()
    validations: list[dict[str, Any]] = []
    for (year, month, product), rows in sorted(groups.items()):
        period = date(year, month, 1)
        source_index_period = add_months(period, -1)
        low = value_for(rows, SOURCE_LOW_LABEL)
        mid = value_for(rows, SOURCE_MID_LABEL)
        resin_source = value_for(rows, "Resin with assumptions")
        discount = value_for(rows, "Discount")
        freight = value_for(rows, "Drewry (t-1) with discount")
        importation = value_for(rows, "Importation")
        import_tax = value_for(rows, "Import Tax")
        surcharge = value_for(rows, "Surcharge")
        indorama_discount = value_for(rows, "Indorama Discount")
        tlc_source = value_for(rows, TLC_USD_METRIC)
        ptax = maybe_float(rows[0].get("ptax")) or 0.0

        resin_calc = resin_with_assumptions(low, mid)
        tlc_calc = calculate_tlc_usd(
            resin_calc,
            discount,
            freight,
            importation,
            import_tax,
            surcharge,
            indorama_discount,
        )
        tlc_brl_source = None
        tlc_brl_formula = None
        values_wb = None
        formulas_wb = None
        try:
            values_wb = load_workbook(SOURCE_FILE, data_only=True, read_only=False)
            formulas_wb = load_workbook(SOURCE_FILE, data_only=False, read_only=False)
            sheet = rows[0]["source_sheet"]
            col_letter = re.match(r"([A-Z]+)", str(rows[0]["source_cell"])).group(1)
            value_ws = values_wb[sheet]
            formula_ws = formulas_wb[sheet]
            tlc_brl_source = maybe_float(value_ws[f"{col_letter}{TLC_BRL_ROW}"].value)
            tlc_brl_formula = formula_ws[f"{col_letter}{TLC_BRL_ROW}"].value
        finally:
            if values_wb:
                values_wb.close()
            if formulas_wb:
                formulas_wb.close()

        tlc_brl_calc = tlc_calc * ptax
        tlc_row = row_for_metric(rows, TLC_USD_METRIC) or {}
        resin_row = row_for_metric(rows, "Resin with assumptions") or {}
        low_ref, low_ref_row = reference_value(reference, source_index_period, PRIMARY_INDEX_TYPE)
        mid_ref, mid_ref_row = reference_value(reference, source_index_period, MID_INDEX_TYPE)
        low_ref_diff = low - low_ref if low_ref is not None else None
        mid_ref_diff = mid - mid_ref if mid_ref is not None else None
        max_formula_diff = max(abs(resin_calc - resin_source), abs(tlc_calc - tlc_source))
        if tlc_brl_source is not None:
            max_formula_diff = max(max_formula_diff, abs(tlc_brl_calc - tlc_brl_source))

        validations.append(
            {
                "time_period": period_label(period),
                "time_period_year": year,
                "time_period_month": month,
                "product": product,
                "source_index_period": period_label(source_index_period),
                "low_index_source": low,
                "low_index_reference": low_ref,
                "low_reference_difference": low_ref_diff,
                "low_reference_status": "match" if low_ref_diff is not None and abs(low_ref_diff) <= 0.000001 else "check",
                "mid_index_source": mid,
                "mid_index_reference": mid_ref,
                "mid_reference_difference": mid_ref_diff,
                "mid_reference_status": "match" if mid_ref_diff is not None and abs(mid_ref_diff) <= 0.000001 else "check",
                "resin_with_assumptions_source": resin_source,
                "resin_with_assumptions_calculated": resin_calc,
                "resin_difference": resin_calc - resin_source,
                "discount": discount,
                "freight": freight,
                "importation": importation,
                "import_tax": import_tax,
                "surcharge": surcharge,
                "indorama_discount": indorama_discount,
                "ptax": ptax,
                "source_tlc_usd": tlc_source,
                "calculated_tlc_usd": tlc_calc,
                "tlc_usd_difference": tlc_calc - tlc_source,
                "source_tlc_brl": tlc_brl_source,
                "calculated_tlc_brl": tlc_brl_calc,
                "tlc_brl_difference": None if tlc_brl_source is None else tlc_brl_calc - tlc_brl_source,
                "max_abs_formula_difference": max_formula_diff,
                "formula_validation_status": "match" if max_formula_diff <= 0.000001 else "check",
                "source_resin_formula": resin_row.get("formula"),
                "source_tlc_usd_formula": tlc_row.get("formula"),
                "source_tlc_brl_formula": tlc_brl_formula,
                "resin_formula_used": RESIN_ASSUMPTION_FORMULA,
                "tlc_formula_used": TLC_FORMULA,
                "low_reference_source_cell": low_ref_row.get("source_cell") if low_ref_row else None,
                "mid_reference_source_cell": mid_ref_row.get("source_cell") if mid_ref_row else None,
            }
        )
    return validations


def build_standardized_rows(final_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in final_rows:
        metric = clean_text(row.get("metric_label_english"))
        if metric == SOURCE_LOW_LABEL:
            resin_index_type = PRIMARY_INDEX_TYPE
        elif metric == SOURCE_MID_LABEL:
            resin_index_type = MID_INDEX_TYPE
        else:
            resin_index_type = SOURCE_RESIN_INDEX_TYPE
        rows.append(
            {
                "Source File ": SOURCE_FILE.name,
                "Supplier Name": SUPPLIER,
                "Destination Country": DESTINATION_COUNTRY,
                "Time_Period ": row.get("time_period"),
                "Time Period Year": row.get("time_period_year"),
                "Time Period Month": row.get("time_period_month"),
                "Location": row.get("product"),
                "Raw Cost Breakdown": metric,
                "Resin Index Type": resin_index_type,
                "Mapping Columns": clean_text(row.get("Mapping Columns")),
                "Column Required for Calculation": clean_text(row.get("Column Required for Calculation")),
                "Value ": row.get("value"),
                "TLC Formula": TLC_FORMULA,
            }
        )
    return sorted(rows, key=period_sort_key)


def build_front_end_actual_rows(standardized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sort_front_end_rows(
        [{"Data Type": "Actual", **row, "Forecast Resin Index Type": None} for row in standardized_rows]
    )


def latest_valid_rows(validation: list[dict[str, Any]]) -> list[dict[str, Any]]:
    matched = [row for row in validation if row["formula_validation_status"] == "match"]
    if not matched:
        raise ValueError("No matching Valgroup validation rows were found.")
    latest_period = max(
        date(int(row["time_period_year"]), int(row["time_period_month"]), 1)
        for row in matched
    )
    return [
        row
        for row in matched
        if int(row["time_period_year"]) == latest_period.year
        and int(row["time_period_month"]) == latest_period.month
    ]


def forecast_periods_after_latest(latest_rows: list[dict[str, Any]], end_year: int = 2026) -> list[date]:
    latest = date(
        int(latest_rows[0]["time_period_year"]),
        int(latest_rows[0]["time_period_month"]),
        1,
    )
    periods: list[date] = []
    current = add_months(latest, 1)
    while current.year <= end_year:
        periods.append(current)
        current = add_months(current, 1)
    return periods


def build_forecast_inputs(validation: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reference = read_index_reference()
    latest_rows = latest_valid_rows(validation)
    latest_by_product = {row["product"]: row for row in latest_rows}
    future_periods = forecast_periods_after_latest(latest_rows)
    rows: list[dict[str, Any]] = []
    for product, latest in sorted(latest_by_product.items()):
        for target_period in future_periods:
            source_index_period = add_months(target_period, -1)
            low_value, low_ref = reference_value(reference, source_index_period, PRIMARY_INDEX_TYPE)
            mid_value, mid_ref = reference_value(reference, source_index_period, MID_INDEX_TYPE)
            fallback_mid = mid_value is None
            effective_mid = mid_value if mid_value is not None else low_value
            rows.append(
                {
                    "time_period": period_label(target_period),
                    "time_period_year": target_period.year,
                    "time_period_month": target_period.month,
                    "product": product,
                    "source_index_period": period_label(source_index_period),
                    "low_index_type": PRIMARY_INDEX_TYPE,
                    "low_index": low_value,
                    "low_index_source_cell": low_ref.get("source_cell") if low_ref else None,
                    "mid_index_type": MID_INDEX_TYPE,
                    "mid_index": mid_value,
                    "mid_index_source_cell": mid_ref.get("source_cell") if mid_ref else None,
                    "mid_missing_fallback_to_low": "Yes" if fallback_mid else "No",
                    "effective_mid_index": effective_mid,
                    "discount": latest["discount"],
                    "freight": latest["freight"],
                    "importation": latest["importation"],
                    "import_tax": latest["import_tax"],
                    "surcharge": latest["surcharge"],
                    "indorama_discount": latest["indorama_discount"],
                    "ptax": latest["ptax"],
                    "latest_actual_period_used": latest["time_period"],
                    "notes": "Update yellow/input values before rerunning forecast if business provides new assumptions.",
                }
            )
    return rows


def build_forecast_estimates(forecast_inputs: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    estimates: list[dict[str, Any]] = []
    front_end_rows: list[dict[str, Any]] = []
    component_specs = [
        ("low_index", SOURCE_LOW_LABEL, "Index", "No", PRIMARY_INDEX_TYPE),
        ("effective_mid_index", SOURCE_MID_LABEL, "Index", "No", MID_INDEX_TYPE),
        ("resin_with_assumptions", "Resin with assumptions", "Resin Index vPET", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("discount", "Discount", "Discount", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("freight", "Drewry (t-1) with discount", "Freight", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("importation", "Importation", "Tax", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("import_tax", "Import Tax", "Tax", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("surcharge", "Surcharge", "Others", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("indorama_discount", "Indorama Discount", "Discount", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("total_vpet_usd_forecast", TLC_USD_METRIC, "Total Landing Cost", "Yes", SOURCE_RESIN_INDEX_TYPE),
        ("total_vpet_brl_forecast", TLC_BRL_METRIC, "Total Landing Cost", "Yes", SOURCE_RESIN_INDEX_TYPE),
    ]
    for row in forecast_inputs:
        low = numeric(row.get("low_index"))
        effective_mid = numeric(row.get("effective_mid_index"))
        resin = resin_with_assumptions(low, effective_mid)
        tlc_usd = calculate_tlc_usd(
            resin,
            numeric(row["discount"]),
            numeric(row["freight"]),
            numeric(row["importation"]),
            numeric(row["import_tax"]),
            numeric(row["surcharge"]),
            numeric(row["indorama_discount"]),
        )
        tlc_brl = tlc_usd * numeric(row["ptax"])
        estimate = {
            **row,
            "resin_with_assumptions": resin,
            "total_vpet_usd_forecast": tlc_usd,
            "total_vpet_brl_forecast": tlc_brl,
            "resin_formula": RESIN_ASSUMPTION_FORMULA,
            "tlc_formula": TLC_FORMULA,
        }
        estimates.append(estimate)
        for value_key, raw_cost, mapping_column, required, resin_index_type in component_specs:
            front_end_rows.append(
                {
                    "Data Type": "Forecast",
                    "Source File ": INDEX_REFERENCE_CSV.name,
                    "Supplier Name": SUPPLIER,
                    "Destination Country": DESTINATION_COUNTRY,
                    "Time_Period ": row["time_period"],
                    "Time Period Year": row["time_period_year"],
                    "Time Period Month": row["time_period_month"],
                    "Location": row["product"],
                    "Raw Cost Breakdown": raw_cost,
                    "Resin Index Type": resin_index_type,
                    "Forecast Resin Index Type": f"{PRIMARY_INDEX_TYPE}; {MID_INDEX_TYPE}",
                    "Mapping Columns": mapping_column,
                    "Column Required for Calculation": required,
                    "Value ": estimate[value_key],
                    "TLC Formula": TLC_FORMULA,
                }
            )
    return estimates, front_end_rows


def sort_front_end_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("Destination Country") or "",
            row.get("Supplier Name") or "",
            int(row.get("Time Period Year") or 0),
            int(row.get("Time Period Month") or 0),
            row.get("Location") or "",
            0 if row.get("Data Type") == "Actual" else 1,
            row.get("Mapping Columns") or "",
            row.get("Raw Cost Breakdown") or "",
        ),
    )


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name[:31])
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"
    style_sheet(worksheet)


def write_workbook(
    path: Path,
    sheets: list[tuple[str, list[str], list[dict[str, Any]]]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers, rows in sheets:
        write_sheet(workbook, sheet_name, headers, rows)
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def formula_catalog_rows() -> list[dict[str, Any]]:
    return [
        {
            "component": SOURCE_LOW_LABEL,
            "classification": "variable_index_input",
            "formula_or_source": f"{INDEX_REFERENCE_CSV.name} / {PRIMARY_INDEX_TYPE}, target month - 1",
        },
        {
            "component": SOURCE_MID_LABEL,
            "classification": "variable_index_input",
            "formula_or_source": f"{INDEX_REFERENCE_CSV.name} / {MID_INDEX_TYPE}, target month - 1; fallback to Low when missing",
        },
        {
            "component": "Resin with assumptions",
            "classification": "formula",
            "formula_or_source": RESIN_ASSUMPTION_FORMULA,
        },
        {
            "component": "Discount",
            "classification": "product_assumption",
            "formula_or_source": "Latest actual by product: VRJ1 2.5%, VPE1 4%, VMG11 1%",
        },
        {
            "component": "Drewry (t-1) with discount",
            "classification": "variable_assumption",
            "formula_or_source": "Latest actual carried forward unless updated",
        },
        {
            "component": "Importation + Import Tax",
            "classification": "constant_assumption",
            "formula_or_source": "6% + 20% from source workbook",
        },
        {
            "component": "Surcharge + Indorama Discount",
            "classification": "constant_assumption",
            "formula_or_source": "85 + (-25) from source workbook",
        },
        {
            "component": TLC_USD_METRIC,
            "classification": "calculated_output",
            "formula_or_source": TLC_FORMULA,
        },
    ]


def write_forecast_template(
    path: Path,
    forecast_inputs: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "product",
        "source_index_period",
        "low_index_type",
        "low_index",
        "low_index_source_cell",
        "mid_index_type",
        "mid_index",
        "mid_index_source_cell",
        "mid_missing_fallback_to_low",
        "effective_mid_index",
        "discount",
        "freight",
        "importation",
        "import_tax",
        "surcharge",
        "indorama_discount",
        "ptax",
        "resin_with_assumptions",
        "total_vpet_usd",
        "total_vpet_brl",
        "latest_actual_period_used",
        "notes",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "forecast_inputs"
    worksheet.append(headers)
    for row in forecast_inputs:
        worksheet.append([row.get(header) for header in headers])
    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    formula_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet[f"M{row_idx}"] = f'=IF(ISBLANK(J{row_idx}),G{row_idx},J{row_idx})'
        worksheet[f"U{row_idx}"] = (
            f'=IF(G{row_idx}>M{row_idx},M{row_idx},'
            f'IF((M{row_idx}-G{row_idx})>75,M{row_idx}-75,G{row_idx}))'
        )
        worksheet[f"V{row_idx}"] = (
            f'=(((U{row_idx}*(1-N{row_idx}))+O{row_idx})*'
            f'(1+(P{row_idx}+Q{row_idx})))+R{row_idx}+S{row_idx}'
        )
        worksheet[f"W{row_idx}"] = f"=V{row_idx}*T{row_idx}"
        for col_idx in [7, 10, 14, 15, 16, 17, 18, 19, 20]:
            worksheet.cell(row_idx, col_idx).fill = input_fill
        for col_idx in [13, 21, 22, 23]:
            worksheet.cell(row_idx, col_idx).fill = formula_fill
    style_sheet(worksheet)
    write_sheet(
        workbook,
        "formula_catalog",
        ["component", "classification", "formula_or_source"],
        formula_catalog_rows(),
    )
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def write_outputs(
    raw_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    standardized: list[dict[str, Any]],
    actual_front_end: list[dict[str, Any]],
    validation: list[dict[str, Any]],
    forecast_inputs: list[dict[str, Any]],
    forecast_estimates: list[dict[str, Any]],
    forecast_front_end: list[dict[str, Any]],
) -> None:
    actual_forecast = sort_front_end_rows(actual_front_end + forecast_front_end)
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "pipeline": PIPELINE_NAME,
        "source_file": str(SOURCE_FILE),
        "index_reference_csv": str(INDEX_REFERENCE_CSV),
        "legacy_extracted_file": str(LEGACY_EXTRACTED_FILE),
        "extracted_file": str(EXTRACTED_FILE),
        "supplier": SUPPLIER,
        "destination_country": DESTINATION_COUNTRY,
        "source_resin_index_type": SOURCE_RESIN_INDEX_TYPE,
        "forecast_resin_index_types": f"{PRIMARY_INDEX_TYPE}; {MID_INDEX_TYPE}",
        "tlc_formula": TLC_FORMULA,
        "resin_formula": RESIN_ASSUMPTION_FORMULA,
        "note": "Full pipeline extracts VRJ1, VPE1, and VMG11 product columns; legacy extraction remains available for the old U2:V12 scope.",
        "raw_rows": len(raw_rows),
        "final_rows": len(final_rows),
        "validation_rows": len(validation),
        "forecast_input_rows": len(forecast_inputs),
        "forecast_estimate_rows": len(forecast_estimates),
        "forecast_front_end_rows": len(forecast_front_end),
        "actual_forecast_front_end_rows": len(actual_forecast),
    }
    validation_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "product",
        "source_index_period",
        "low_index_source",
        "low_index_reference",
        "low_reference_difference",
        "low_reference_status",
        "mid_index_source",
        "mid_index_reference",
        "mid_reference_difference",
        "mid_reference_status",
        "resin_with_assumptions_source",
        "resin_with_assumptions_calculated",
        "resin_difference",
        "discount",
        "freight",
        "importation",
        "import_tax",
        "surcharge",
        "indorama_discount",
        "ptax",
        "source_tlc_usd",
        "calculated_tlc_usd",
        "tlc_usd_difference",
        "source_tlc_brl",
        "calculated_tlc_brl",
        "tlc_brl_difference",
        "max_abs_formula_difference",
        "formula_validation_status",
        "source_resin_formula",
        "source_tlc_usd_formula",
        "source_tlc_brl_formula",
        "resin_formula_used",
        "tlc_formula_used",
        "low_reference_source_cell",
        "mid_reference_source_cell",
    ]
    forecast_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "product",
        "source_index_period",
        "low_index_type",
        "low_index",
        "low_index_source_cell",
        "mid_index_type",
        "mid_index",
        "mid_index_source_cell",
        "mid_missing_fallback_to_low",
        "effective_mid_index",
        "discount",
        "freight",
        "importation",
        "import_tax",
        "surcharge",
        "indorama_discount",
        "ptax",
        "latest_actual_period_used",
        "notes",
        "resin_with_assumptions",
        "total_vpet_usd_forecast",
        "total_vpet_brl_forecast",
        "resin_formula",
        "tlc_formula",
    ]
    write_workbook(
        EXTRACTED_FILE,
        [
            ("final_data", ordered_headers(final_rows), final_rows),
            ("raw_product_blocks", ordered_headers(raw_rows), raw_rows),
        ],
        metadata,
    )
    write_workbook(
        STANDARDIZED_FILE,
        [("final_standardized", STANDARDIZED_HEADERS, standardized)],
        metadata,
    )
    write_workbook(
        TLC_VALIDATION_FILE,
        [
            ("tlc_validation", validation_headers, validation),
            ("formula_catalog", ["component", "classification", "formula_or_source"], formula_catalog_rows()),
        ],
        metadata,
    )
    write_forecast_template(FORECAST_TEMPLATE_FILE, forecast_inputs, metadata)
    write_workbook(
        FORECAST_ESTIMATION_FILE,
        [
            ("valgroup_2026_estimate", forecast_headers, forecast_estimates),
            ("standardized_rows", FRONT_END_HEADERS, forecast_front_end),
        ],
        metadata,
    )
    write_workbook(
        FRONT_END_ACTUAL_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, actual_front_end)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_CSV, FRONT_END_HEADERS, actual_front_end)
    write_workbook(
        FRONT_END_ACTUAL_FORECAST_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, actual_forecast)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_FORECAST_CSV, FRONT_END_HEADERS, actual_forecast)

    summary = {
        **metadata,
        "formula_validation_statuses": sorted({row["formula_validation_status"] for row in validation}),
        "low_reference_statuses": sorted({row["low_reference_status"] for row in validation}),
        "mid_reference_statuses": sorted({row["mid_reference_status"] for row in validation}),
        "max_abs_formula_difference": max((row["max_abs_formula_difference"] for row in validation), default=0),
        "max_abs_low_reference_difference": max(
            (
                abs(row["low_reference_difference"])
                for row in validation
                if row["low_reference_difference"] is not None
            ),
            default=0,
        ),
        "max_abs_mid_reference_difference": max(
            (
                abs(row["mid_reference_difference"])
                for row in validation
                if row["mid_reference_difference"] is not None
            ),
            default=0,
        ),
        "forecast_months": sorted({row["time_period"] for row in forecast_estimates}, key=lambda p: (
            next(row["time_period_year"] for row in forecast_estimates if row["time_period"] == p),
            next(row["time_period_month"] for row in forecast_estimates if row["time_period"] == p),
        )),
        "mid_fallback_rows": sum(1 for row in forecast_inputs if row["mid_missing_fallback_to_low"] == "Yes"),
        "front_end_actual_forecast_csv": str(FRONT_END_ACTUAL_FORECAST_CSV),
    }
    VALIDATION_SUMMARY_FILE.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_SUMMARY_FILE.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Valgroup supplier pipeline from extraction through forecast artifacts."
    )
    parser.add_argument("--skip-legacy-extraction", action="store_true")
    parser.add_argument("--skip-legacy-mapping", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for directory in (
        EXTRACTION_DIR,
        STANDARDIZED_DIR,
        CALCULATION_DIR,
        FORECAST_DIR,
        FRONT_END_DIR,
        VALIDATION_DIR,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    if not args.skip_legacy_extraction:
        run_legacy_extraction()
    if not args.skip_legacy_mapping:
        run_legacy_mapping()

    raw_rows, final_rows = extract_full_rows()
    validation = validation_rows(final_rows)
    standardized = build_standardized_rows(final_rows)
    actual_front_end = build_front_end_actual_rows(standardized)
    forecast_inputs = build_forecast_inputs(validation)
    forecast_estimates, forecast_front_end = build_forecast_estimates(forecast_inputs)
    write_outputs(
        raw_rows,
        final_rows,
        standardized,
        actual_front_end,
        validation,
        forecast_inputs,
        forecast_estimates,
        forecast_front_end,
    )

    print(f"supplier={SUPPLIER}")
    print(f"destination_country={DESTINATION_COUNTRY}")
    print(f"final_rows={len(final_rows)}")
    print(f"validation_rows={len(validation)}")
    print(f"forecast_input_rows={len(forecast_inputs)}")
    print(f"forecast_estimate_rows={len(forecast_estimates)}")
    print(f"mid_fallback_rows={sum(1 for row in forecast_inputs if row['mid_missing_fallback_to_low'] == 'Yes')}")
    print(f"extracted_file={EXTRACTED_FILE}")
    print(f"tlc_validation_file={TLC_VALIDATION_FILE}")
    print(f"forecast_template={FORECAST_TEMPLATE_FILE}")
    print(f"forecast_estimation={FORECAST_ESTIMATION_FILE}")
    print(f"front_end_actual_forecast_csv={FRONT_END_ACTUAL_FORECAST_CSV}")
    print(f"validation_summary={VALIDATION_SUMMARY_FILE}")


if __name__ == "__main__":
    main()
