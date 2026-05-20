from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PIPELINE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[3]

PIPELINE_NAME = "04. Amcor"
SOURCE_FILE = REPO_ROOT / "data" / "04. Amcor.xlsx"
EXTRACT_SCRIPT = REPO_ROOT / "scripts" / "extract_amcor_resina.py"
MAPPING_SCRIPT = REPO_ROOT / "scripts" / "data_mapping_column_amcor.py"
FORMULA_OVERVIEW_FILE = REPO_ROOT / "mapping files" / "Formula overview by countries.xlsx"
RESIN_FORECAST_CSV = REPO_ROOT / "Estimation" / "resin_index_forecast_table.csv"

ARTIFACTS_DIR = PIPELINE_DIR / "artifacts"
EXTRACTION_DIR = ARTIFACTS_DIR / "extraction"
STANDARDIZED_DIR = ARTIFACTS_DIR / "standardized"
CALCULATION_DIR = ARTIFACTS_DIR / "calculation"
FORECAST_DIR = ARTIFACTS_DIR / "forecast"
FRONT_END_DIR = ARTIFACTS_DIR / "front_end"
VALIDATION_DIR = ARTIFACTS_DIR / "validation"

EXTRACTED_FILE = EXTRACTION_DIR / "04. Amcor_resina_final.xlsx"
STANDARDIZED_FILE = STANDARDIZED_DIR / "04. Amcor_actual_standardized.xlsx"
TLC_VALIDATION_FILE = CALCULATION_DIR / "amcor_tlc_calculation_validation.xlsx"
FORECAST_TEMPLATE_FILE = FORECAST_DIR / "amcor_forecast_inputs_template.xlsx"
FORECAST_ESTIMATION_FILE = FORECAST_DIR / "amcor_2026_forward_estimation.xlsx"
FRONT_END_ACTUAL_XLSX = FRONT_END_DIR / "04. Amcor_front_end_standardized.xlsx"
FRONT_END_ACTUAL_CSV = FRONT_END_DIR / "04. Amcor_front_end_standardized.csv"
FRONT_END_ACTUAL_FORECAST_XLSX = FRONT_END_DIR / "04. Amcor_actual_forecast_front_end_standardized.xlsx"
FRONT_END_ACTUAL_FORECAST_CSV = FRONT_END_DIR / "04. Amcor_actual_forecast_front_end_standardized.csv"
VALIDATION_FILE = VALIDATION_DIR / "validation_summary.json"

FINAL_DATA_SHEET = "final_data"
FORMULA_OVERVIEW_SHEETS = ("Pricing formula_Updated", "Pricing formula overview")
FORECAST_RESIN_INDEX_TYPE = "PET Bottle Grade FOB China Spot"
FORECAST_MONTHS = [
    "May 2026",
    "June 2026",
    "July 2026",
    "August 2026",
    "September 2026",
    "October 2026",
    "November 2026",
    "December 2026",
]

DEFAULT_METADATA = {
    "supplier": "Amcor",
    "destination_country": "Brazil",
    "resin_index_type": "ICIS PET China MID (M-1) USD/ton",
    "tlc_formula": "Resin: [(Index + Int. Freight) x (1+Tax) + Others + Scrap]",
}

DUTY_RATE = 0.208
TONS_PER_CONTAINER = 22.5
FIXED_COST = 167.0
LOCATION_FREIGHT_ADDON = {
    "SUAPE": 0.0,
    "MANAUS": 20.0,
}
LOCATION_FIXED_METRIC = {
    "SUAPE": "Others",
    "MANAUS": "Seguro",
}
SOURCE_RESIN_LABEL = "ICIS China mid (M-1)"
TLC_METRIC = "Sell Side Resina USD"
FREIGHT_METRIC = "International freight"
DUTIES_METRIC = "Duties"
DREWRY_METRIC = "Drewry"

STANDARDIZED_HEADERS = [
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]

FRONT_END_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Location",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def clean_header(value: Any) -> str:
    return str(clean_text(value) or "")


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or "")).lower().replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def clean_index(value: Any) -> Any:
    text = clean_text(value)
    if isinstance(text, str):
        return re.sub(r"^index\s*:\s*", "", text, flags=re.IGNORECASE).strip()
    return text


def numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def run_command(command: list[str]) -> None:
    print(" ".join(f'"{part}"' if " " in part else part for part in command))
    subprocess.run(command, cwd=REPO_ROOT, check=True)


def run_extraction() -> None:
    run_command(
        [
            sys.executable,
            str(EXTRACT_SCRIPT),
            "--file",
            str(SOURCE_FILE),
            "--output-dir",
            str(EXTRACTION_DIR),
            "--excel-output",
            str(EXTRACTED_FILE),
            "--write-csv",
        ]
    )


def run_mapping() -> None:
    run_command([sys.executable, str(MAPPING_SCRIPT), "--output-file", str(EXTRACTED_FILE)])


def formula_overview_metadata() -> dict[str, Any]:
    metadata = DEFAULT_METADATA.copy()
    if not FORMULA_OVERVIEW_FILE.exists():
        return metadata

    try:
        workbook = load_workbook(FORMULA_OVERVIEW_FILE, data_only=True, read_only=True)
    except PermissionError:
        return metadata

    try:
        target = normalize_key(PIPELINE_NAME)
        current_country: Any = None
        for sheet_name in FORMULA_OVERVIEW_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]
            headers = [clean_header(cell.value).rstrip() for cell in worksheet[2]]
            header_map = {header: idx for idx, header in enumerate(headers)}
            if "Pricing Sheet" not in header_map:
                continue

            for row in worksheet.iter_rows(min_row=3, values_only=True):
                country_idx = header_map.get("Countries")
                if country_idx is not None and row[country_idx]:
                    current_country = clean_text(row[country_idx])

                pricing_sheet = clean_text(row[header_map["Pricing Sheet"]])
                if normalize_key(pricing_sheet) != target:
                    continue

                supplier_idx = header_map.get("Supplier")
                formula_idx = header_map.get("vPET Resin pricing adjustment formula")
                index_idx = header_map.get("Index")
                metadata["supplier"] = (
                    clean_text(row[supplier_idx])
                    if supplier_idx is not None and row[supplier_idx]
                    else metadata["supplier"]
                )
                metadata["destination_country"] = current_country or metadata["destination_country"]
                metadata["tlc_formula"] = (
                    clean_text(row[formula_idx])
                    if formula_idx is not None and row[formula_idx]
                    else metadata["tlc_formula"]
                )
                metadata["resin_index_type"] = (
                    clean_index(row[index_idx])
                    if index_idx is not None and row[index_idx]
                    else metadata["resin_index_type"]
                )
                metadata["formula_overview_sheet"] = sheet_name
                return metadata
        return metadata
    finally:
        workbook.close()


def read_final_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[FINAL_DATA_SHEET]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


def row_location(row: dict[str, Any]) -> Any:
    return clean_text(row.get("section_name") or row.get("location"))


def build_standardized_rows(final_rows: list[dict[str, Any]], metadata: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in final_rows:
        rows.append(
            {
                "Source File ": Path(str(row.get("source_file") or SOURCE_FILE)).name,
                "Supplier Name": metadata["supplier"],
                "Destination Country": metadata["destination_country"],
                "Time_Period ": row.get("time_period"),
                "Time Period Year": row.get("time_period_year"),
                "Time Period Month": row.get("time_period_month"),
                "Location": row_location(row),
                "Raw Cost Breakdown": clean_text(row.get("metric_name")),
                "Resin Index Type": metadata["resin_index_type"],
                "Mapping Columns": clean_text(row.get("Mapping Columns")),
                "Column Required for Calculation": clean_text(row.get("Column Required for Calculation")),
                "Value ": row.get("value"),
                "TLC Formula": metadata["tlc_formula"],
            }
        )
    return rows


def incomplete_actual_periods(validation_rows: list[dict[str, Any]]) -> set[Any]:
    return {
        row.get("time_period")
        for row in validation_rows
        if row.get("validation_status") != "match"
    }


def complete_actual_rows(
    final_rows: list[dict[str, Any]],
    validation_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    incomplete_periods = incomplete_actual_periods(validation_rows)
    return [
        row
        for row in final_rows
        if row.get("time_period") not in incomplete_periods
    ]


def build_front_end_actual_rows(standardized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in standardized_rows:
        rows.append(
            {
                "Data Type": "Actual",
                **row,
                "Forecast Resin Index Type": None,
            }
        )
    return sort_front_end_rows(rows)


def grouped_rows(final_rows: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in final_rows:
        location = row_location(row)
        period = row.get("time_period")
        if location in LOCATION_FREIGHT_ADDON and period:
            groups.setdefault((location, period), []).append(row)
    return groups


def metric_row(rows: list[dict[str, Any]], metric_name: str) -> dict[str, Any] | None:
    for row in rows:
        if clean_text(row.get("metric_name")) == metric_name:
            return row
    return None


def metric_value(rows: list[dict[str, Any]], metric_name: str) -> float:
    row = metric_row(rows, metric_name)
    return numeric(row.get("value") if row else None)


def period_sort_from_row(row: dict[str, Any]) -> tuple[int, int]:
    return int(row.get("time_period_year") or 0), int(row.get("time_period_month") or 0)


def latest_valid_period_rows(groups: dict[tuple[str, str], list[dict[str, Any]]]) -> dict[str, list[dict[str, Any]]]:
    latest_by_location: dict[str, tuple[tuple[int, int], list[dict[str, Any]]]] = {}
    for (location, _period), rows in groups.items():
        if metric_value(rows, SOURCE_RESIN_LABEL) <= 0 or metric_value(rows, DREWRY_METRIC) <= 0:
            continue
        key = period_sort_from_row(rows[0])
        if location not in latest_by_location or key > latest_by_location[location][0]:
            latest_by_location[location] = (key, rows)
    return {location: rows for location, (_key, rows) in latest_by_location.items()}


def build_validation_rows(groups: dict[tuple[str, str], list[dict[str, Any]]]) -> list[dict[str, Any]]:
    validation: list[dict[str, Any]] = []
    for (location, period), rows in sorted(
        groups.items(),
        key=lambda item: (period_sort_from_row(item[1][0]), item[0][0]),
    ):
        resin = metric_value(rows, SOURCE_RESIN_LABEL)
        freight = metric_value(rows, FREIGHT_METRIC)
        duties = metric_value(rows, DUTIES_METRIC)
        fixed_metric = LOCATION_FIXED_METRIC[location]
        fixed = metric_value(rows, fixed_metric)
        drewry = metric_value(rows, DREWRY_METRIC)
        source_tlc = metric_value(rows, TLC_METRIC)
        calculated_freight = drewry / TONS_PER_CONTAINER + LOCATION_FREIGHT_ADDON[location]
        calculated_duties = (resin + freight) * DUTY_RATE
        calculated_tlc = resin + freight + duties + fixed
        max_difference = max(
            abs(calculated_freight - freight),
            abs(calculated_duties - duties),
            abs(calculated_tlc - source_tlc),
        )
        incomplete_inputs = resin <= 0 or drewry <= 0
        if max_difference <= 0.000001 and not incomplete_inputs:
            status = "match"
        elif max_difference <= 0.000001:
            status = "formula_match_input_check"
        else:
            status = "check"

        source_tlc_row = metric_row(rows, TLC_METRIC) or {}
        validation.append(
            {
                "location": location,
                "time_period": period,
                "time_period_year": rows[0].get("time_period_year"),
                "time_period_month": rows[0].get("time_period_month"),
                "resin_index_type": DEFAULT_METADATA["resin_index_type"],
                "resin_price_index": resin,
                "drewry": drewry,
                "source_freight": freight,
                "calculated_freight": calculated_freight,
                "source_duties": duties,
                "calculated_duties": calculated_duties,
                "fixed_cost_metric": fixed_metric,
                "fixed_cost": fixed,
                "source_total_landing_cost": source_tlc,
                "calculated_total_landing_cost": calculated_tlc,
                "difference": calculated_tlc - source_tlc,
                "max_abs_component_difference": max_difference,
                "validation_status": status,
                "source_tlc_formula": source_tlc_row.get("formula"),
                "tlc_formula": (
                    "Sell Side Resina USD = Resin Index + International Freight + Duties + "
                    f"{fixed_metric}; Duties = (Resin Index + International Freight) * 20.8%; "
                    "International Freight = Drewry / 22.5 plus location add-on."
                ),
            }
        )
    return validation


def read_resin_forecast() -> dict[str, dict[str, Any]]:
    if not RESIN_FORECAST_CSV.exists():
        return {}
    with RESIN_FORECAST_CSV.open(newline="", encoding="utf-8-sig") as csv_file:
        rows = list(csv.DictReader(csv_file))
    forecast: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("resin_index_type") != FORECAST_RESIN_INDEX_TYPE:
            continue
        period = row.get("time_period")
        if period in FORECAST_MONTHS and period not in forecast:
            forecast[period] = row
    return forecast


def fixed_assumption_rows() -> list[dict[str, Any]]:
    return [
        {
            "location": location,
            "assumption_name": "duty_rate",
            "value": DUTY_RATE,
            "classification": "fixed",
            "notes": "According to Brazilian law in formula overview.",
        }
        for location in LOCATION_FREIGHT_ADDON
    ] + [
        {
            "location": location,
            "assumption_name": "tons_per_container",
            "value": TONS_PER_CONTAINER,
            "classification": "fixed",
            "notes": "Used to convert Drewry/container to USD/ton.",
        }
        for location in LOCATION_FREIGHT_ADDON
    ] + [
        {
            "location": location,
            "assumption_name": "location_freight_addon",
            "value": addon,
            "classification": "fixed",
            "notes": "Manaus adds 20 USD/ton; Suape has no add-on.",
        }
        for location, addon in LOCATION_FREIGHT_ADDON.items()
    ] + [
        {
            "location": location,
            "assumption_name": "fixed_cost",
            "value": FIXED_COST,
            "classification": "fixed",
            "notes": f"{LOCATION_FIXED_METRIC[location]} from latest source formula.",
        }
        for location in LOCATION_FREIGHT_ADDON
    ]


def build_forecast_inputs(
    latest_rows_by_location: dict[str, list[dict[str, Any]]],
    resin_forecast: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for location, latest_rows in latest_rows_by_location.items():
        latest_drewry = metric_value(latest_rows, DREWRY_METRIC)
        latest_fixed = metric_value(latest_rows, LOCATION_FIXED_METRIC[location])
        latest_period = latest_rows[0].get("time_period")
        for period in FORECAST_MONTHS:
            forecast_row = resin_forecast.get(period, {})
            rows.append(
                {
                    "time_period": period,
                    "time_period_year": forecast_row.get("time_period_year"),
                    "time_period_month": forecast_row.get("time_period_month"),
                    "location": location,
                    "resin_index_type": DEFAULT_METADATA["resin_index_type"],
                    "forecast_resin_index_type": FORECAST_RESIN_INDEX_TYPE,
                    "resin_price_index": numeric(forecast_row.get("value")),
                    "drewry_index": latest_drewry,
                    "duty_rate": DUTY_RATE,
                    "tons_per_container": TONS_PER_CONTAINER,
                    "location_freight_addon": LOCATION_FREIGHT_ADDON[location],
                    "fixed_cost_metric": LOCATION_FIXED_METRIC[location],
                    "fixed_cost": latest_fixed or FIXED_COST,
                    "changeable_resin_price_index": "Yes",
                    "changeable_drewry_index": "Yes",
                    "changeable_duty_rate": "No",
                    "changeable_fixed_cost": "No",
                    "latest_actual_period_used": latest_period,
                    "notes": "Update changeable values before forecast if new market inputs are available.",
                }
            )
    return rows


def build_forecast_estimates(forecast_inputs: list[dict[str, Any]], metadata: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    estimates: list[dict[str, Any]] = []
    standardized: list[dict[str, Any]] = []
    for row in forecast_inputs:
        resin = numeric(row["resin_price_index"])
        drewry = numeric(row["drewry_index"])
        duty_rate = numeric(row["duty_rate"])
        tons = numeric(row["tons_per_container"]) or TONS_PER_CONTAINER
        addon = numeric(row["location_freight_addon"])
        fixed = numeric(row["fixed_cost"])
        freight = drewry / tons + addon
        duties = (resin + freight) * duty_rate
        tlc = resin + freight + duties + fixed
        estimate = {
            **row,
            "international_freight": freight,
            "duties": duties,
            "total_landing_cost_forecast": tlc,
            "tlc_formula": "TLC = Resin Index + International Freight + Duties + Fixed Cost",
        }
        estimates.append(estimate)

        component_specs = [
            (SOURCE_RESIN_LABEL, "Resin Index vPET", "Yes", resin),
            (FREIGHT_METRIC, "Freight", "Yes", freight),
            (DUTIES_METRIC, "Tax", "Yes", duties),
            (row["fixed_cost_metric"], "Others" if row["location"] == "SUAPE" else "Seguro", "Yes", fixed),
            (DREWRY_METRIC, "Drewery Index", "No", drewry),
            (TLC_METRIC, "Total Landing Cost", "Yes", tlc),
        ]
        for metric_name, mapping_column, required, value in component_specs:
            standardized.append(
                {
                    "Data Type": "Forecast",
                    "Source File ": RESIN_FORECAST_CSV.name,
                    "Supplier Name": metadata["supplier"],
                    "Destination Country": metadata["destination_country"],
                    "Time_Period ": row["time_period"],
                    "Time Period Year": row["time_period_year"],
                    "Time Period Month": row["time_period_month"],
                    "Location": row["location"],
                    "Raw Cost Breakdown": metric_name,
                    "Resin Index Type": metadata["resin_index_type"],
                    "Forecast Resin Index Type": FORECAST_RESIN_INDEX_TYPE,
                    "Mapping Columns": mapping_column,
                    "Column Required for Calculation": required,
                    "Value ": value,
                    "TLC Formula": metadata["tlc_formula"],
                }
            )
    return estimates, standardized


def sort_front_end_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("Destination Country") or "",
            row.get("Supplier Name") or "",
            int(row.get("Time Period Year") or 0),
            int(row.get("Time Period Month") or 0),
            row.get("Location") or "",
            0 if row.get("Data Type") == "Actual" else 1,
            row.get("Mapping Columns") or "",
            row.get("Raw Cost Breakdown") or "",
        ),
    )


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    style_sheet(worksheet)


def write_workbook(path: Path, sheets: list[tuple[str, list[str], list[dict[str, Any]]]], metadata: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers, rows in sheets:
        write_sheet(workbook, sheet_name, headers, rows)
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def unique_nonblank(rows: list[dict[str, Any]], column: str) -> list[Any]:
    values: list[Any] = []
    for row in rows:
        value = row.get(column)
        if value in (None, ""):
            continue
        if value not in values:
            values.append(value)
    return values


def write_outputs(
    metadata: dict[str, Any],
    source_final_rows: list[dict[str, Any]],
    actual_final_rows: list[dict[str, Any]],
    standardized_rows: list[dict[str, Any]],
    actual_front_end_rows: list[dict[str, Any]],
    validation_rows: list[dict[str, Any]],
    fixed_rows: list[dict[str, Any]],
    forecast_inputs: list[dict[str, Any]],
    forecast_estimates: list[dict[str, Any]],
    forecast_front_end_rows: list[dict[str, Any]],
) -> None:
    forecast_all_rows = sort_front_end_rows(actual_front_end_rows + forecast_front_end_rows)
    metadata = {
        **metadata,
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "source_file": str(SOURCE_FILE),
        "extracted_file": str(EXTRACTED_FILE),
        "source_extracted_rows": len(source_final_rows),
        "actual_rows": len(actual_final_rows),
        "standardized_rows": len(standardized_rows),
        "actual_front_end_rows": len(actual_front_end_rows),
        "tlc_validation_rows": len(validation_rows),
        "forecast_input_rows": len(forecast_inputs),
        "forecast_estimate_rows": len(forecast_estimates),
        "forecast_front_end_rows": len(forecast_front_end_rows),
        "actual_forecast_front_end_rows": len(forecast_all_rows),
    }

    validation_headers = [
        "location",
        "time_period",
        "time_period_year",
        "time_period_month",
        "resin_index_type",
        "resin_price_index",
        "drewry",
        "source_freight",
        "calculated_freight",
        "source_duties",
        "calculated_duties",
        "fixed_cost_metric",
        "fixed_cost",
        "source_total_landing_cost",
        "calculated_total_landing_cost",
        "difference",
        "max_abs_component_difference",
        "validation_status",
        "source_tlc_formula",
        "tlc_formula",
    ]
    fixed_headers = ["location", "assumption_name", "value", "classification", "notes"]
    forecast_input_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "location",
        "resin_index_type",
        "forecast_resin_index_type",
        "resin_price_index",
        "drewry_index",
        "duty_rate",
        "tons_per_container",
        "location_freight_addon",
        "fixed_cost_metric",
        "fixed_cost",
        "changeable_resin_price_index",
        "changeable_drewry_index",
        "changeable_duty_rate",
        "changeable_fixed_cost",
        "latest_actual_period_used",
        "notes",
    ]
    forecast_estimate_headers = [
        *forecast_input_headers,
        "international_freight",
        "duties",
        "total_landing_cost_forecast",
        "tlc_formula",
    ]
    formula_catalog = [
        {
            "component": "Resin Index",
            "classification": "changeable",
            "formula_or_source": "Forecast resin index table / user input",
        },
        {
            "component": "Drewry",
            "classification": "changeable",
            "formula_or_source": "Latest actual carried forward unless updated",
        },
        {
            "component": "International Freight",
            "classification": "calculated",
            "formula_or_source": "Drewry / 22.5 + location freight add-on",
        },
        {
            "component": "Duties",
            "classification": "calculated",
            "formula_or_source": "(Resin Index + International Freight) * 20.8%",
        },
        {
            "component": "Others/Seguro",
            "classification": "fixed",
            "formula_or_source": "167 USD/ton from source formula = 90 + 77",
        },
        {
            "component": "Sell Side Resina USD",
            "classification": "calculated_output",
            "formula_or_source": "Resin Index + International Freight + Duties + Others/Seguro",
        },
    ]

    write_workbook(
        STANDARDIZED_FILE,
        [("final_standardized", STANDARDIZED_HEADERS, standardized_rows)],
        metadata,
    )
    write_workbook(
        FRONT_END_ACTUAL_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, actual_front_end_rows)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_CSV, FRONT_END_HEADERS, actual_front_end_rows)
    write_workbook(
        TLC_VALIDATION_FILE,
        [("tlc_validation", validation_headers, validation_rows)],
        metadata,
    )
    write_workbook(
        FORECAST_TEMPLATE_FILE,
        [
            ("changeable_inputs", forecast_input_headers, forecast_inputs),
            ("fixed_assumptions", fixed_headers, fixed_rows),
            ("formula_catalog", ["component", "classification", "formula_or_source"], formula_catalog),
        ],
        metadata,
    )
    write_workbook(
        FORECAST_ESTIMATION_FILE,
        [
            ("amcor_2026_estimate", forecast_estimate_headers, forecast_estimates),
            ("standardized_rows", FRONT_END_HEADERS, forecast_front_end_rows),
        ],
        metadata,
    )
    write_workbook(
        FRONT_END_ACTUAL_FORECAST_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, forecast_all_rows)],
        metadata,
    )
    write_csv(FRONT_END_ACTUAL_FORECAST_CSV, FRONT_END_HEADERS, forecast_all_rows)

    summary = {
        **metadata,
        "validation_statuses": sorted({row["validation_status"] for row in validation_rows}),
        "max_abs_component_difference": max(
            (row["max_abs_component_difference"] for row in validation_rows),
            default=0,
        ),
        "locations": sorted({row["location"] for row in validation_rows}),
        "first_actual_period": None,
        "last_actual_period": None,
        "mapping_columns": unique_nonblank(source_final_rows, "Mapping Columns"),
        "raw_cost_breakdowns": unique_nonblank(standardized_rows, "Raw Cost Breakdown"),
    }
    actual_periods = sorted(
        {
            (
                int(row.get("Time Period Year") or 0),
                int(row.get("Time Period Month") or 0),
                row.get("Time_Period "),
            )
            for row in actual_front_end_rows
            if row.get("Time Period Year") and row.get("Time Period Month")
        }
    )
    if actual_periods:
        summary["first_actual_period"] = actual_periods[0][2]
        summary["last_actual_period"] = actual_periods[-1][2]

    VALIDATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_FILE.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Amcor supplier pipeline from extraction to validation and forecast-ready artifacts."
    )
    parser.add_argument("--skip-extraction", action="store_true")
    parser.add_argument("--skip-mapping", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for directory in (
        EXTRACTION_DIR,
        STANDARDIZED_DIR,
        CALCULATION_DIR,
        FORECAST_DIR,
        FRONT_END_DIR,
        VALIDATION_DIR,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    if not args.skip_extraction:
        run_extraction()
    if not args.skip_mapping:
        run_mapping()

    source_final_rows = read_final_rows(EXTRACTED_FILE)
    metadata = formula_overview_metadata()
    groups = grouped_rows(source_final_rows)
    validation_rows = build_validation_rows(groups)
    actual_final_rows = complete_actual_rows(source_final_rows, validation_rows)
    standardized_rows = build_standardized_rows(actual_final_rows, metadata)
    actual_front_end_rows = build_front_end_actual_rows(standardized_rows)
    latest_rows_by_location = latest_valid_period_rows(groups)
    resin_forecast = read_resin_forecast()
    fixed_rows = fixed_assumption_rows()
    forecast_inputs = build_forecast_inputs(latest_rows_by_location, resin_forecast)
    forecast_estimates, forecast_front_end_rows = build_forecast_estimates(forecast_inputs, metadata)
    write_outputs(
        metadata,
        source_final_rows,
        actual_final_rows,
        standardized_rows,
        actual_front_end_rows,
        validation_rows,
        fixed_rows,
        forecast_inputs,
        forecast_estimates,
        forecast_front_end_rows,
    )

    print(f"supplier={metadata['supplier']}")
    print(f"destination_country={metadata['destination_country']}")
    print(f"resin_index_type={metadata['resin_index_type']}")
    print(f"source_extracted_rows={len(source_final_rows)}")
    print(f"actual_rows={len(actual_final_rows)}")
    print(f"tlc_validation_rows={len(validation_rows)}")
    print(f"forecast_input_rows={len(forecast_inputs)}")
    print(f"actual_forecast_front_end_rows={len(actual_front_end_rows) + len(forecast_front_end_rows)}")
    print(f"forecast_template={FORECAST_TEMPLATE_FILE}")
    print(f"validation_summary={VALIDATION_FILE}")


if __name__ == "__main__":
    main()
