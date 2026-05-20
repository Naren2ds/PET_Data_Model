from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PIPELINE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[3]

PIPELINE_NAME = "Pricing_Pref_ABI_PAN_Men_04_26_Cliente"
SOURCE_FILE = REPO_ROOT / "data" / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente.xlsx"
EXTRACT_SCRIPT = REPO_ROOT / "scripts" / "extract_pricing_pref_pan_inputs.py"
MAPPING_SCRIPT = REPO_ROOT / "scripts" / "data_mapping_column_pan.py"
FORMULA_OVERVIEW_FILE = REPO_ROOT / "mapping files" / "Formula overview by countries.xlsx"
INDEX_REFERENCE_CSV = REPO_ROOT / "Index Fprecast" / "icis_resin_index_reference_table.csv"

ARTIFACTS_DIR = PIPELINE_DIR / "artifacts"
EXTRACTION_DIR = ARTIFACTS_DIR / "extraction"
STANDARDIZED_DIR = ARTIFACTS_DIR / "standardized"
CALCULATION_DIR = ARTIFACTS_DIR / "calculation"
FORECAST_DIR = ARTIFACTS_DIR / "forecast"
FRONT_END_DIR = ARTIFACTS_DIR / "front_end"
VALIDATION_DIR = ARTIFACTS_DIR / "validation"

EXTRACTED_FILE = EXTRACTION_DIR / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente_inputs_final.xlsx"
STANDARDIZED_FILE = STANDARDIZED_DIR / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente_actual_standardized.xlsx"
TLC_VALIDATION_FILE = CALCULATION_DIR / "pan_tlc_calculation_validation.xlsx"
FORECAST_TEMPLATE_FILE = FORECAST_DIR / "pan_forecast_inputs_template.xlsx"
FORECAST_ESTIMATION_FILE = FORECAST_DIR / "pan_2026_forward_estimation.xlsx"
FRONT_END_ACTUAL_XLSX = FRONT_END_DIR / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente_front_end_standardized.xlsx"
FRONT_END_ACTUAL_CSV = FRONT_END_DIR / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente_front_end_standardized.csv"
FRONT_END_ACTUAL_FORECAST_XLSX = (
    FRONT_END_DIR / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente_actual_forecast_front_end_standardized.xlsx"
)
FRONT_END_ACTUAL_FORECAST_CSV = (
    FRONT_END_DIR / "Pricing_Pref_ABI_PAN_Men_04_26_Cliente_actual_forecast_front_end_standardized.csv"
)
VALIDATION_FILE = VALIDATION_DIR / "validation_summary.json"

FINAL_DATA_SHEET = "final_data"
INDEX_SERIES = "ICIS Asia SE Low"
DEFAULT_INDEX_LAG_MONTHS = 2
FORECAST_MONTH_END = date(2026, 12, 1)
SOURCE_INDEX_BREAKDOWN = "ICIS Asia SE Low (M-2)"
TLC_OUTPUT_METRIC = "DDP Price"
TLC_FORMULA = "DDP Price = (Average FOB + Ocean Freight) * (1 + Import Tax) + Import Clearance"
WORKBOOK_FORMULA = "DDP Price = Average FOB + Ocean Freight + Import Clearance"

METRIC_ROW_TO_STANDARD_LABEL = {
    7: SOURCE_INDEX_BREAKDOWN,
    8: "Average FOB",
    9: "Ocean Freight",
    10: "Import Clearance (%)",
    11: "DDP Price",
}
CALCULATION_REQUIRED_ROWS = {8, 9, 10, 11}

DEFAULT_METADATA = {
    "supplier": "Pastiglas S.A",
    "destination_country": "Panama",
    "resin_index_type": "ICIS Asia SE Low index (M-2)",
    "tlc_formula": "vPET Resin USD/Ton = (ICIS + OF) * (1 + IT) + CC",
}

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

STANDARDIZED_HEADERS = [
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]

FRONT_END_HEADERS = [
    "Data Type",
    "Source File ",
    "Supplier Name",
    "Destination Country",
    "Time_Period ",
    "Time Period Year",
    "Time Period Month",
    "Raw Cost Breakdown",
    "Resin Index Type",
    "Forecast Resin Index Type",
    "Mapping Columns",
    "Column Required for Calculation",
    "Value ",
    "TLC Formula",
]


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def clean_header(value: Any) -> str:
    return str(clean_text(value) or "")


def normalize_key(value: Any) -> str:
    text = strip_accents(str(value or "")).lower().replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def maybe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def period_label(value: date) -> str:
    return f"{MONTH_LABELS[value.month]} {value.year}"


def add_months(value: date, months: int) -> date:
    month_number = value.month - 1 + months
    year = value.year + month_number // 12
    month = month_number % 12 + 1
    return date(year, month, 1)


def parse_period(row: dict[str, Any]) -> date:
    return date(int(row["time_period_year"]), int(row["time_period_month"]), 1)


def parse_source_index_period(value: Any) -> date | None:
    text = str(value or "").strip()
    match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)
    return None


def source_index_label(value: date) -> str:
    return f"{value.year}-{value.month}"


def run_command(command: list[str]) -> None:
    print(" ".join(f'"{part}"' if " " in part else part for part in command))
    subprocess.run(command, cwd=REPO_ROOT, check=True)


def run_extraction() -> None:
    run_command(
        [
            sys.executable,
            str(EXTRACT_SCRIPT),
            "--file",
            str(SOURCE_FILE),
            "--output-dir",
            str(EXTRACTION_DIR),
            "--excel-output",
            str(EXTRACTED_FILE),
            "--write-csv",
        ]
    )


def run_mapping() -> None:
    run_command([sys.executable, str(MAPPING_SCRIPT), "--output-file", str(EXTRACTED_FILE)])


def formula_overview_metadata() -> dict[str, Any]:
    metadata = DEFAULT_METADATA.copy()
    if not FORMULA_OVERVIEW_FILE.exists():
        return metadata

    try:
        workbook = load_workbook(FORMULA_OVERVIEW_FILE, data_only=True, read_only=True)
    except (PermissionError, OSError):
        metadata["formula_overview_note"] = "Formula overview file could not be opened; defaults used."
        return metadata

    try:
        current_country: Any = None
        target = normalize_key(PIPELINE_NAME)
        for sheet_name in ("Pricing formula_Updated", "Pricing formula overview"):
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            headers = [clean_header(cell.value).rstrip() for cell in worksheet[2]]
            header_map = {normalize_key(header): idx for idx, header in enumerate(headers)}
            pricing_idx = header_map.get("pricing sheet")
            if pricing_idx is None:
                continue
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                country_idx = header_map.get("countries")
                if country_idx is not None and row[country_idx]:
                    current_country = clean_text(row[country_idx])

                pricing_sheet = clean_text(row[pricing_idx])
                if normalize_key(pricing_sheet) != target:
                    continue

                supplier_idx = header_map.get("supplier")
                formula_idx = header_map.get("vpet resin pricing adjustment formula")
                index_idx = header_map.get("index")

                if supplier_idx is not None and row[supplier_idx]:
                    metadata["supplier"] = clean_text(row[supplier_idx])
                metadata["destination_country"] = current_country or metadata["destination_country"]
                if formula_idx is not None and row[formula_idx]:
                    metadata["tlc_formula"] = clean_text(row[formula_idx])
                if index_idx is not None and row[index_idx]:
                    metadata["resin_index_type"] = clean_text(row[index_idx])
                metadata["formula_overview_sheet"] = sheet_name
                return metadata
        return metadata
    finally:
        workbook.close()


def read_final_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        worksheet = workbook[FINAL_DATA_SHEET]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if row.get("metric_row") is not None:
                row["metric_row"] = int(row["metric_row"])
            rows.append(row)
        return rows
    finally:
        workbook.close()


def rows_by_period(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row["time_period"]), []).append(row)
    return grouped


def row_for_metric_row(rows: list[dict[str, Any]], metric_row: int) -> dict[str, Any] | None:
    for row in rows:
        if row.get("metric_row") == metric_row:
            return row
    return None


def read_index_reference(path: Path) -> dict[tuple[str, int, int], dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        records = list(csv.DictReader(csv_file))

    reference: dict[tuple[str, int, int], dict[str, Any]] = {}
    for row in records:
        if row.get("resin_index_type") != INDEX_SERIES:
            continue
        try:
            key = (
                str(row["resin_index_type"]),
                int(row["time_period_year"]),
                int(row["time_period_month"]),
            )
        except (KeyError, TypeError, ValueError):
            continue
        reference[key] = row
    return reference


def lookup_index(
    reference: dict[tuple[str, int, int], dict[str, Any]],
    period: date,
) -> dict[str, Any] | None:
    return reference.get((INDEX_SERIES, period.year, period.month))


def detect_index_lag(rows: list[dict[str, Any]]) -> int:
    source_row = row_for_metric_row(rows, 7)
    if not source_row:
        return DEFAULT_INDEX_LAG_MONTHS
    source_period = parse_source_index_period(source_row.get("metric_label_original"))
    if not source_period:
        return DEFAULT_INDEX_LAG_MONTHS
    target_period = parse_period(source_row)
    return (target_period.year - source_period.year) * 12 + target_period.month - source_period.month


def value_for_row(rows: list[dict[str, Any]], metric_row: int) -> float:
    row = row_for_metric_row(rows, metric_row)
    return numeric(row.get("value") if row else None)


def validation_rows(
    final_rows: list[dict[str, Any]],
    index_reference: dict[tuple[str, int, int], dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped = rows_by_period(final_rows)
    rows: list[dict[str, Any]] = []

    for period, period_rows in sorted(grouped.items(), key=lambda item: parse_period(item[1][0])):
        source_index_row = row_for_metric_row(period_rows, 7)
        average_fob_row = row_for_metric_row(period_rows, 8)
        freight_row = row_for_metric_row(period_rows, 9)
        clearance_row = row_for_metric_row(period_rows, 10)
        ddp_row = row_for_metric_row(period_rows, 11)
        if not all([source_index_row, average_fob_row, freight_row, clearance_row, ddp_row]):
            continue

        source_index_period = parse_source_index_period(source_index_row.get("metric_label_original"))
        index_ref_row = lookup_index(index_reference, source_index_period) if source_index_period else None
        index_ref_value = maybe_float(index_ref_row.get("value")) if index_ref_row else None

        source_index_value = numeric(source_index_row.get("value"))
        average_fob = numeric(average_fob_row.get("value"))
        ocean_freight = numeric(freight_row.get("value"))
        import_clearance = numeric(clearance_row.get("value"))
        source_ddp = numeric(ddp_row.get("value"))
        calculated_ddp = average_fob + ocean_freight + import_clearance
        formula_difference = calculated_ddp - source_ddp
        index_difference = (
            None if index_ref_value is None else source_index_value - index_ref_value
        )
        fob_index_difference = (
            None if index_ref_value is None else average_fob - index_ref_value
        )

        rows.append(
            {
                "time_period": period,
                "time_period_year": ddp_row.get("time_period_year"),
                "time_period_month": ddp_row.get("time_period_month"),
                "source_index_period": None if source_index_period is None else period_label(source_index_period),
                "source_index_period_label": source_index_row.get("metric_label_original"),
                "resin_index_type": INDEX_SERIES,
                "source_index_value": source_index_value,
                "reference_index_value": index_ref_value,
                "index_reference_difference": index_difference,
                "index_reference_status": (
                    "match"
                    if index_difference is not None and abs(index_difference) <= 0.000001
                    else "check"
                ),
                "average_fob": average_fob,
                "average_fob_reference_difference": fob_index_difference,
                "ocean_freight": ocean_freight,
                "import_clearance": import_clearance,
                "import_tax_rate": 0.0,
                "calculated_ddp_price": calculated_ddp,
                "source_ddp_price": source_ddp,
                "formula_difference": formula_difference,
                "formula_validation_status": (
                    "match" if abs(formula_difference) <= 0.000001 else "check"
                ),
                "source_ddp_formula": ddp_row.get("formula"),
                "tlc_formula": WORKBOOK_FORMULA,
                "index_reference_source_cell": index_ref_row.get("source_cell") if index_ref_row else None,
                "index_reference_formula": index_ref_row.get("formula") if index_ref_row else None,
            }
        )

    return rows


def latest_period_rows(grouped: dict[str, list[dict[str, Any]]]) -> tuple[str, list[dict[str, Any]]]:
    period = max(grouped, key=lambda key: parse_period(grouped[key][0]))
    return period, grouped[period]


def forecast_months(latest_period: date) -> list[date]:
    months: list[date] = []
    current = add_months(latest_period, 1)
    while current <= FORECAST_MONTH_END:
        months.append(current)
        current = add_months(current, 1)
    return months


def forecast_estimate_rows(
    latest_rows: list[dict[str, Any]],
    index_reference: dict[tuple[str, int, int], dict[str, Any]],
    index_lag_months: int,
) -> list[dict[str, Any]]:
    latest_period = parse_period(latest_rows[0])
    ocean_freight = value_for_row(latest_rows, 9)
    import_clearance = value_for_row(latest_rows, 10)
    import_tax_rate = 0.0

    rows: list[dict[str, Any]] = []
    for period in forecast_months(latest_period):
        source_period = add_months(period, -index_lag_months)
        index_ref_row = lookup_index(index_reference, source_period)
        if not index_ref_row:
            continue

        resin_price_index = numeric(index_ref_row.get("value"))
        ddp_price = (resin_price_index + ocean_freight) * (1 + import_tax_rate) + import_clearance
        rows.append(
            {
                "time_period": period_label(period),
                "time_period_year": period.year,
                "time_period_month": period.month,
                "source_index_period": period_label(source_period),
                "source_index_period_label": source_index_label(source_period),
                "index_lag_months": index_lag_months,
                "resin_index_type": INDEX_SERIES,
                "forecast_resin_index_type": INDEX_SERIES,
                "resin_forecast_series": index_ref_row.get("forecast_series"),
                "resin_forecast_source_cell": index_ref_row.get("source_cell"),
                "resin_forecast_formula": index_ref_row.get("formula"),
                "resin_price_index": resin_price_index,
                "ocean_freight": ocean_freight,
                "import_tax_rate": import_tax_rate,
                "import_clearance": import_clearance,
                "ddp_price_forecast": ddp_price,
                "tlc_formula": TLC_FORMULA,
            }
        )
    return rows


def row_mapping(row: dict[str, Any]) -> str:
    return str(clean_text(row.get("Mapping Columns")) or "")


def raw_breakdown(row: dict[str, Any]) -> str:
    metric_row = int(row.get("metric_row") or 0)
    return METRIC_ROW_TO_STANDARD_LABEL.get(
        metric_row,
        str(clean_text(row.get("metric_label_english") or row.get("metric_label_original")) or ""),
    )


def standardized_actual_rows(
    final_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in final_rows:
        metric_row = int(row.get("metric_row") or 0)
        rows.append(
            {
                "Source File ": Path(str(row.get("source_file") or SOURCE_FILE)).name,
                "Supplier Name": metadata["supplier"],
                "Destination Country": metadata["destination_country"],
                "Time_Period ": row.get("time_period"),
                "Time Period Year": row.get("time_period_year"),
                "Time Period Month": row.get("time_period_month"),
                "Raw Cost Breakdown": raw_breakdown(row),
                "Resin Index Type": metadata["resin_index_type"],
                "Mapping Columns": row_mapping(row),
                "Column Required for Calculation": "Yes" if metric_row in CALCULATION_REQUIRED_ROWS else "No",
                "Value ": row.get("value"),
                "TLC Formula": metadata["tlc_formula"],
            }
        )
    return rows


def front_end_actual_rows(standardized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in standardized_rows:
        rows.append(
            {
                "Data Type": "Actual",
                **row,
                "Forecast Resin Index Type": None,
            }
        )
    return rows


def forecast_template_rows(estimates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for estimate in estimates:
        rows.append(
            {
                "time_period": estimate["time_period"],
                "time_period_year": estimate["time_period_year"],
                "time_period_month": estimate["time_period_month"],
                "source_index_period": estimate["source_index_period"],
                "source_index_period_label": estimate["source_index_period_label"],
                "index_lag_months": estimate["index_lag_months"],
                "resin_index_type": estimate["resin_index_type"],
                "forecast_resin_index_type": estimate["forecast_resin_index_type"],
                "resin_price_index": estimate["resin_price_index"],
                "ocean_freight": estimate["ocean_freight"],
                "import_tax_rate": estimate["import_tax_rate"],
                "import_clearance": estimate["import_clearance"],
                "ddp_price_forecast": estimate["ddp_price_forecast"],
                "tlc_formula": estimate["tlc_formula"],
                "notes": "Freight, import tax, and import clearance are carried from the April 2026 source workbook.",
            }
        )
    return rows


def forecast_front_end_rows(
    estimates: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> list[dict[str, Any]]:
    component_specs = [
        ("ICIS Asia SE Low (M-2)", "Resin Index vPET", "No", "resin_price_index"),
        ("Average FOB", "Resin Index vPET", "Yes", "resin_price_index"),
        ("Ocean Freight", "Freight", "Yes", "ocean_freight"),
        ("Import Clearance (%)", "Tax", "Yes", "import_clearance"),
        ("DDP Price", "Total Landing Cost", "Yes", "ddp_price_forecast"),
    ]

    rows: list[dict[str, Any]] = []
    for estimate in estimates:
        for breakdown, mapping, required, value_key in component_specs:
            rows.append(
                {
                    "Data Type": "Forecast",
                    "Source File ": SOURCE_FILE.name,
                    "Supplier Name": metadata["supplier"],
                    "Destination Country": metadata["destination_country"],
                    "Time_Period ": estimate["time_period"],
                    "Time Period Year": estimate["time_period_year"],
                    "Time Period Month": estimate["time_period_month"],
                    "Raw Cost Breakdown": breakdown,
                    "Resin Index Type": metadata["resin_index_type"],
                    "Forecast Resin Index Type": estimate["forecast_resin_index_type"],
                    "Mapping Columns": mapping,
                    "Column Required for Calculation": required,
                    "Value ": estimate[value_key],
                    "TLC Formula": metadata["tlc_formula"],
                }
            )
    return rows


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(sheet_name[:31])
    worksheet.append(headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = worksheet.cell(row_idx, col_idx, value=value)
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"
    style_sheet(worksheet)


def write_workbook(
    path: Path,
    sheets: list[tuple[str, list[str], list[dict[str, Any]]]],
    metadata: dict[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers, rows in sheets:
        write_sheet(workbook, sheet_name, headers, rows)
    write_sheet(
        workbook,
        "run_metadata",
        ["field", "value"],
        [{"field": key, "value": value} for key, value in metadata.items()],
    )
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_outputs(
    metadata: dict[str, Any],
    final_rows: list[dict[str, Any]],
    validation: list[dict[str, Any]],
    forecast_inputs: list[dict[str, Any]],
    estimates: list[dict[str, Any]],
    actual_standardized: list[dict[str, Any]],
    actual_front_end: list[dict[str, Any]],
    forecast_front_end: list[dict[str, Any]],
    index_lag_months: int,
) -> None:
    run_metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "pipeline": PIPELINE_NAME,
        "source_file": str(SOURCE_FILE),
        "extracted_file": str(EXTRACTED_FILE),
        "index_reference_csv": str(INDEX_REFERENCE_CSV),
        "supplier": metadata["supplier"],
        "destination_country": metadata["destination_country"],
        "source_resin_index_type": metadata["resin_index_type"],
        "formula_overview_resin_index_type": metadata.get("formula_overview_resin_index_type"),
        "forecast_resin_index_type": INDEX_SERIES,
        "index_lag_months": index_lag_months,
        "source_workbook_formula": WORKBOOK_FORMULA,
        "forecast_formula": TLC_FORMULA,
        "raw_rows": len(final_rows),
        "validation_rows": len(validation),
        "forecast_months": len(estimates),
        "forecast_front_end_rows": len(forecast_front_end),
        "actual_forecast_front_end_rows": len(actual_front_end) + len(forecast_front_end),
    }

    validation_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_index_period",
        "source_index_period_label",
        "resin_index_type",
        "source_index_value",
        "reference_index_value",
        "index_reference_difference",
        "index_reference_status",
        "average_fob",
        "average_fob_reference_difference",
        "ocean_freight",
        "import_clearance",
        "import_tax_rate",
        "calculated_ddp_price",
        "source_ddp_price",
        "formula_difference",
        "formula_validation_status",
        "source_ddp_formula",
        "tlc_formula",
        "index_reference_source_cell",
        "index_reference_formula",
    ]
    estimate_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_index_period",
        "source_index_period_label",
        "index_lag_months",
        "resin_index_type",
        "forecast_resin_index_type",
        "resin_forecast_series",
        "resin_forecast_source_cell",
        "resin_forecast_formula",
        "resin_price_index",
        "ocean_freight",
        "import_tax_rate",
        "import_clearance",
        "ddp_price_forecast",
        "tlc_formula",
    ]
    forecast_input_headers = [
        "time_period",
        "time_period_year",
        "time_period_month",
        "source_index_period",
        "source_index_period_label",
        "index_lag_months",
        "resin_index_type",
        "forecast_resin_index_type",
        "resin_price_index",
        "ocean_freight",
        "import_tax_rate",
        "import_clearance",
        "ddp_price_forecast",
        "tlc_formula",
        "notes",
    ]
    catalog_rows = [
        {
            "field": "resin_price_index",
            "value": None,
            "classification": "variable_index_input",
            "notes": f"Loaded from {INDEX_REFERENCE_CSV.name}; uses {INDEX_SERIES} with {index_lag_months}-month lag.",
        },
        {
            "field": "ocean_freight",
            "value": value_for_row(final_rows, 9),
            "classification": "fixed_latest_actual",
            "notes": "Carried from April 2026 source workbook.",
        },
        {
            "field": "import_tax_rate",
            "value": 0.0,
            "classification": "constant",
            "notes": "Formula overview says import tax is currently 0%.",
        },
        {
            "field": "import_clearance",
            "value": value_for_row(final_rows, 10),
            "classification": "fixed_latest_actual",
            "notes": "Carried from April 2026 source workbook.",
        },
        {
            "field": "ddp_price_forecast",
            "value": None,
            "classification": "calculated_output",
            "notes": TLC_FORMULA,
        },
    ]

    write_workbook(
        STANDARDIZED_FILE,
        [("final_standardized", STANDARDIZED_HEADERS, actual_standardized)],
        run_metadata,
    )
    write_workbook(
        TLC_VALIDATION_FILE,
        [
            ("tlc_validation", validation_headers, validation),
            ("formula_catalog", ["field", "value", "classification", "notes"], catalog_rows),
        ],
        run_metadata,
    )
    write_workbook(
        FORECAST_TEMPLATE_FILE,
        [
            ("forecast_inputs", forecast_input_headers, forecast_inputs),
            ("formula_catalog", ["field", "value", "classification", "notes"], catalog_rows),
        ],
        run_metadata,
    )
    write_workbook(
        FORECAST_ESTIMATION_FILE,
        [
            ("pan_2026_estimate", estimate_headers, estimates),
            ("standardized_rows", FRONT_END_HEADERS, forecast_front_end),
        ],
        run_metadata,
    )
    write_workbook(
        FRONT_END_ACTUAL_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, actual_front_end)],
        run_metadata,
    )
    write_csv(FRONT_END_ACTUAL_CSV, FRONT_END_HEADERS, actual_front_end)

    merged_front_end = actual_front_end + forecast_front_end
    write_workbook(
        FRONT_END_ACTUAL_FORECAST_XLSX,
        [("front_end_standardized", FRONT_END_HEADERS, merged_front_end)],
        run_metadata,
    )
    write_csv(FRONT_END_ACTUAL_FORECAST_CSV, FRONT_END_HEADERS, merged_front_end)

    formula_diffs = [abs(row["formula_difference"]) for row in validation]
    index_diffs = [
        abs(row["index_reference_difference"])
        for row in validation
        if row["index_reference_difference"] is not None
    ]
    summary = {
        **run_metadata,
        "formula_validation_statuses": sorted({row["formula_validation_status"] for row in validation}),
        "index_reference_statuses": sorted({row["index_reference_status"] for row in validation}),
        "max_abs_formula_difference": max(formula_diffs, default=0),
        "max_abs_index_reference_difference": max(index_diffs, default=0),
        "forecast_periods": [row["time_period"] for row in estimates],
        "front_end_actual_forecast_csv": str(FRONT_END_ACTUAL_FORECAST_CSV),
    }
    VALIDATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_FILE.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the PAN supplier extraction, validation, forecast, and front-end pipeline."
    )
    parser.add_argument("--skip-extraction", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.skip_extraction:
        run_extraction()
        run_mapping()

    metadata = formula_overview_metadata()
    final_rows = read_final_rows(EXTRACTED_FILE)
    grouped = rows_by_period(final_rows)
    latest_period, latest_rows = latest_period_rows(grouped)
    index_lag_months = detect_index_lag(latest_rows)
    metadata["formula_overview_resin_index_type"] = metadata.get("resin_index_type")
    metadata["resin_index_type"] = f"{INDEX_SERIES} (M-{index_lag_months})"
    index_reference = read_index_reference(INDEX_REFERENCE_CSV)

    validation = validation_rows(final_rows, index_reference)
    estimates = forecast_estimate_rows(latest_rows, index_reference, index_lag_months)
    forecast_inputs = forecast_template_rows(estimates)
    actual_standardized = standardized_actual_rows(final_rows, metadata)
    actual_front_end = front_end_actual_rows(actual_standardized)
    forecast_front_end = forecast_front_end_rows(estimates, metadata)

    write_outputs(
        metadata=metadata,
        final_rows=final_rows,
        validation=validation,
        forecast_inputs=forecast_inputs,
        estimates=estimates,
        actual_standardized=actual_standardized,
        actual_front_end=actual_front_end,
        forecast_front_end=forecast_front_end,
        index_lag_months=index_lag_months,
    )

    print(f"pipeline={PIPELINE_NAME}")
    print(f"supplier={metadata['supplier']}")
    print(f"destination_country={metadata['destination_country']}")
    print(f"latest_actual_period={latest_period}")
    print(f"index_lag_months={index_lag_months}")
    print(f"validation_rows={len(validation)}")
    print(f"forecast_months={len(estimates)}")
    print(f"actual_front_end_rows={len(actual_front_end)}")
    print(f"forecast_front_end_rows={len(forecast_front_end)}")
    print(f"extracted_file={EXTRACTED_FILE}")
    print(f"tlc_validation_file={TLC_VALIDATION_FILE}")
    print(f"forecast_template_file={FORECAST_TEMPLATE_FILE}")
    print(f"forecast_estimation_file={FORECAST_ESTIMATION_FILE}")
    print(f"front_end_actual_forecast_csv={FRONT_END_ACTUAL_FORECAST_CSV}")
    print(f"validation_summary={VALIDATION_FILE}")


if __name__ == "__main__":
    main()
