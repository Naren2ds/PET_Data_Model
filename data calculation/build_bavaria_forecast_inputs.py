from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_FILE = (
    Path("output") / "1 Bavaria - Precio Marzo_formula_final_with_mapping_columns.xlsx"
)
DEFAULT_OUTPUT_DIR = Path("data calculation")
DEFAULT_CATALOG_OUTPUT = DEFAULT_OUTPUT_DIR / "Bavaria_formula_catalog.xlsx"
DEFAULT_TEMPLATE_OUTPUT = DEFAULT_OUTPUT_DIR / "Bavaria_future_inputs_template.xlsx"
DEFAULT_FINAL_SHEET = "final_data"
FORECAST_MONTHS = 12

RESIN_METRIC = "Resin Price Index"
FINANCE_METRIC = "Finance"
FREIGHT_REGULAR_METRIC = "Freight China-Buenaventura (Regular)"
FREIGHT_INCREMENTAL_METRIC = "Freight China-Buenaventura (Incremental)"
SUBTOTAL_INCREMENTAL_METRIC = "Sub Total (with Incremental Freight)"
SUBTOTAL_REGULAR_METRIC = "Sub Total (with Regular Freight)"
DUTY_METRIC = "Duty 5% (Change According to Regulation)"
LANDED_FACTOR_METRIC = "Landed Factor 8%"
ZF_LEGISLATION_METRIC = "ZF Legislation Change"
SURCHARGE_METRIC = "Sur Charge Alpek Br"
SOURCE_TLC_METRIC = "Total Resin Price ABI VIRGIN Formula"
RESIN_INDEX_TYPE_COLUMN = "resin_index_type"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

MONTH_LOOKUP = {
    "jan": 1,
    "january": 1,
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "february": 2,
    "febrero": 2,
    "mar": 3,
    "march": 3,
    "marzo": 3,
    "apr": 4,
    "april": 4,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "june": 6,
    "junio": 6,
    "jul": 7,
    "july": 7,
    "julio": 7,
    "aug": 8,
    "august": 8,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septie": 9,
    "septiembre": 9,
    "set": 9,
    "oct": 10,
    "october": 10,
    "octubre": 10,
    "nov": 11,
    "november": 11,
    "noviembre": 11,
    "dec": 12,
    "december": 12,
    "dic": 12,
    "diciembre": 12,
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def parse_period(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    if value is None:
        return None

    text = strip_accents(str(value)).strip()
    if not text:
        return None

    parts = [part for part in re.split(r"[-/\s]+", text) if part]
    if len(parts) < 2:
        return None

    month_token = parts[0].lower()
    month = MONTH_LOOKUP.get(month_token, MONTH_LOOKUP.get(month_token[:3]))
    if month is None:
        return None

    year_token = re.sub(r"\D", "", parts[-1])
    if not year_token:
        return None

    year = int(year_token)
    if year < 100:
        year += 2000

    return date(year, month, 1)


def add_months(value: date, months: int) -> date:
    month_number = value.month - 1 + months
    year = value.year + month_number // 12
    month = month_number % 12 + 1
    return date(year, month, 1)


def format_period(value: date) -> str:
    return f"{MONTH_LABELS[value.month]} {value.year}"


def source_column(source_cell: Any) -> str | None:
    if not source_cell:
        return None
    match = re.match(r"([A-Z]+)", str(source_cell).upper())
    return match.group(1) if match else None


def load_final_rows(input_file: Path, final_sheet: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_file, data_only=True, read_only=True)
    try:
        if final_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {final_sheet!r} not found in {input_file}.")

        worksheet = workbook[final_sheet]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        rows = [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
        return headers, rows
    finally:
        workbook.close()


def group_by_source_column(final_rows: list[dict[str, Any]]) -> dict[str, dict[str, dict[str, Any]]]:
    grouped: dict[str, dict[str, dict[str, Any]]] = {}
    for row in final_rows:
        col = source_column(row.get("source_cell"))
        metric = row.get("metric_en")
        if not col or not metric:
            continue
        grouped.setdefault(col, {})[metric] = row
    return grouped


def metric_value(metrics: dict[str, dict[str, Any]], metric_name: str) -> float | None:
    return numeric_value(metrics.get(metric_name, {}).get("value"))


def zero_if_missing(value: Any) -> float:
    return numeric_value(value) or 0.0


def latest_period_metrics(final_rows: list[dict[str, Any]]) -> dict[str, Any]:
    grouped = group_by_source_column(final_rows)
    candidates: list[tuple[date, str, dict[str, dict[str, Any]]]] = []
    for col, metrics in grouped.items():
        period_value = metrics.get(SOURCE_TLC_METRIC, {}).get("price_period") or next(
            (
                row.get("price_period")
                for row in metrics.values()
                if row.get("price_period") is not None
            ),
            None,
        )
        parsed = parse_period(period_value)
        if parsed:
            candidates.append((parsed, col, metrics))

    if not candidates:
        raise ValueError("No dated Bavaria periods found in final_data.")

    latest_period, latest_col, metrics = sorted(candidates, key=lambda item: item[0])[-1]
    resin_index_type = next(
        (
            clean_text(row.get(RESIN_INDEX_TYPE_COLUMN))
            for row in metrics.values()
            if row.get(RESIN_INDEX_TYPE_COLUMN)
        ),
        next(
            (
                clean_text(row.get("metric_local"))
                for row in metrics.values()
                if row.get("metric_en") == RESIN_METRIC and row.get("metric_local")
            ),
            None,
        ),
    )
    if isinstance(resin_index_type, str):
        resin_index_type = resin_index_type.rstrip(" .")
    resin = zero_if_missing(metric_value(metrics, RESIN_METRIC))
    finance = zero_if_missing(metric_value(metrics, FINANCE_METRIC))
    freight_regular = zero_if_missing(metric_value(metrics, FREIGHT_REGULAR_METRIC))
    freight_incremental = zero_if_missing(metric_value(metrics, FREIGHT_INCREMENTAL_METRIC))
    subtotal_incremental = zero_if_missing(metric_value(metrics, SUBTOTAL_INCREMENTAL_METRIC))
    subtotal_regular = zero_if_missing(metric_value(metrics, SUBTOTAL_REGULAR_METRIC))
    duty = zero_if_missing(metric_value(metrics, DUTY_METRIC))
    landed_factor = zero_if_missing(metric_value(metrics, LANDED_FACTOR_METRIC))
    zf_legislation = zero_if_missing(metric_value(metrics, ZF_LEGISLATION_METRIC))
    surcharge = zero_if_missing(metric_value(metrics, SURCHARGE_METRIC))
    total_landing_cost = zero_if_missing(metric_value(metrics, SOURCE_TLC_METRIC))

    finance_base = resin + freight_regular + freight_incremental
    finance_factor = finance / finance_base if finance_base else None
    duty_rate = duty / subtotal_incremental if subtotal_incremental else 0.05
    landed_factor_rate = landed_factor / subtotal_regular if subtotal_regular else 0.08

    return {
        "latest_period": latest_period,
        "latest_source_column": latest_col,
        "resin_index_type": resin_index_type,
        "resin_price_index": resin,
        "finance": finance,
        "finance_factor": finance_factor,
        "freight_regular": freight_regular,
        "freight_incremental": freight_incremental,
        "sub_total_incremental": subtotal_incremental,
        "sub_total_regular": subtotal_regular,
        "duty": duty,
        "duty_rate": duty_rate,
        "landed_factor": landed_factor,
        "landed_factor_rate": landed_factor_rate,
        "zf_legislation_change": zf_legislation,
        "surcharge_alpek_br": surcharge,
        "total_landing_cost": total_landing_cost,
    }


def catalog_rows(defaults: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "component_order": 1,
            "component_name": RESIN_METRIC,
            "component_type": "Variable Input",
            "mapping_column": "Resin Index vPET",
            "source_metric_en": RESIN_METRIC,
            "future_input_column": "resin_price_index",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Yes",
            "formula_logic": "User-provided monthly resin index.",
            "notes": "Required each future month.",
        },
        {
            "component_order": 2,
            "component_name": FREIGHT_REGULAR_METRIC,
            "component_type": "Variable Input",
            "mapping_column": "Freight",
            "source_metric_en": FREIGHT_REGULAR_METRIC,
            "future_input_column": "freight_regular",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": defaults["freight_regular"],
            "required_for_forecast": "Yes",
            "formula_logic": "User-provided monthly regular freight, defaulted to latest historical value.",
            "notes": "Update if regular freight changes.",
        },
        {
            "component_order": 3,
            "component_name": "Drewry Freight",
            "component_type": "Variable Input",
            "mapping_column": "Freight",
            "source_metric_en": None,
            "future_input_column": "drewry_freight",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Conditional",
            "formula_logic": "Used to calculate incremental freight as MAX(Drewry Freight - Freight Regular, 0).",
            "notes": "Fill this or fill freight_incremental_override directly.",
        },
        {
            "component_order": 4,
            "component_name": FREIGHT_INCREMENTAL_METRIC,
            "component_type": "Formula / Override",
            "mapping_column": "Freight",
            "source_metric_en": FREIGHT_INCREMENTAL_METRIC,
            "future_input_column": "freight_incremental_override",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Conditional",
            "formula_logic": "If override is blank, MAX(drewry_freight - freight_regular, 0).",
            "notes": "Use override when the business provides the incremental amount directly.",
        },
        {
            "component_order": 5,
            "component_name": "Finance Factor",
            "component_type": "Variable Assumption",
            "mapping_column": "Insurance",
            "source_metric_en": FINANCE_METRIC,
            "future_input_column": "finance_factor",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": defaults["finance_factor"],
            "required_for_forecast": "Yes",
            "formula_logic": "Finance = (Resin Price Index + Freight Regular + Freight Incremental) x Finance Factor, unless finance_value_override is provided.",
            "notes": "Latest factor is inferred from latest historical Finance row.",
        },
        {
            "component_order": 6,
            "component_name": FINANCE_METRIC,
            "component_type": "Formula / Override",
            "mapping_column": "Insurance",
            "source_metric_en": FINANCE_METRIC,
            "future_input_column": "finance_value_override",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Conditional",
            "formula_logic": "If override is blank, (resin_price_index + freight_regular + freight_incremental_used) x finance_factor.",
            "notes": "Fill override if Finance is provided as a value instead of factor.",
        },
        {
            "component_order": 7,
            "component_name": SUBTOTAL_INCREMENTAL_METRIC,
            "component_type": "Formula",
            "mapping_column": "Sub Total (CIF)",
            "source_metric_en": SUBTOTAL_INCREMENTAL_METRIC,
            "future_input_column": None,
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Calculated",
            "formula_logic": "resin_price_index + finance_used + freight_regular + freight_incremental_used",
            "notes": "Used as the basis for Duty and final TLC.",
        },
        {
            "component_order": 8,
            "component_name": SUBTOTAL_REGULAR_METRIC,
            "component_type": "Formula",
            "mapping_column": "Sub Total (CIF)",
            "source_metric_en": SUBTOTAL_REGULAR_METRIC,
            "future_input_column": None,
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Calculated",
            "formula_logic": "resin_price_index + finance_used + freight_regular",
            "notes": "Used as the basis for Landed Factor.",
        },
        {
            "component_order": 9,
            "component_name": "Duty Rate",
            "component_type": "Constant / Assumption",
            "mapping_column": "Tax",
            "source_metric_en": DUTY_METRIC,
            "future_input_column": "duty_rate",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": defaults["duty_rate"],
            "required_for_forecast": "Yes",
            "formula_logic": "Duty = Sub Total (with Incremental Freight) x Duty Rate.",
            "notes": "Currently 5% for future Bavaria forecast.",
        },
        {
            "component_order": 10,
            "component_name": DUTY_METRIC,
            "component_type": "Formula",
            "mapping_column": "Tax",
            "source_metric_en": DUTY_METRIC,
            "future_input_column": None,
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Calculated",
            "formula_logic": "sub_total_incremental x duty_rate",
            "notes": None,
        },
        {
            "component_order": 11,
            "component_name": "Landed Factor Rate",
            "component_type": "Constant / Assumption",
            "mapping_column": "Tax",
            "source_metric_en": LANDED_FACTOR_METRIC,
            "future_input_column": "landed_factor_rate",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": defaults["landed_factor_rate"],
            "required_for_forecast": "Yes",
            "formula_logic": "Landed Factor = Sub Total (with Regular Freight) x Landed Factor Rate.",
            "notes": "Currently 8%.",
        },
        {
            "component_order": 12,
            "component_name": LANDED_FACTOR_METRIC,
            "component_type": "Formula",
            "mapping_column": "Tax",
            "source_metric_en": LANDED_FACTOR_METRIC,
            "future_input_column": None,
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Calculated",
            "formula_logic": "sub_total_regular x landed_factor_rate",
            "notes": None,
        },
        {
            "component_order": 13,
            "component_name": ZF_LEGISLATION_METRIC,
            "component_type": "Variable Input",
            "mapping_column": "Tax",
            "source_metric_en": ZF_LEGISLATION_METRIC,
            "future_input_column": "zf_legislation_change",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": defaults["zf_legislation_change"],
            "required_for_forecast": "Yes",
            "formula_logic": "User-provided or latest known fixed monthly value.",
            "notes": "Update if ZF legislation value changes.",
        },
        {
            "component_order": 14,
            "component_name": SURCHARGE_METRIC,
            "component_type": "Variable Input",
            "mapping_column": "Tax",
            "source_metric_en": SURCHARGE_METRIC,
            "future_input_column": "surcharge_alpek_br",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": defaults["surcharge_alpek_br"],
            "required_for_forecast": "Yes",
            "formula_logic": "User-provided surcharge; use 0 when not applicable.",
            "notes": "Latest months are 0.",
        },
        {
            "component_order": 15,
            "component_name": SOURCE_TLC_METRIC,
            "component_type": "Output Formula",
            "mapping_column": "Total Landing Cost",
            "source_metric_en": SOURCE_TLC_METRIC,
            "future_input_column": "total_landing_cost_forecast",
            "resin_index_type": defaults["resin_index_type"],
            "default_value": None,
            "required_for_forecast": "Calculated",
            "formula_logic": "sub_total_incremental + duty + landed_factor + zf_legislation_change + surcharge_alpek_br",
            "notes": "This is the forecasted TLC for future months.",
        },
    ]


def field_definition_rows() -> list[dict[str, Any]]:
    return [
        {"field": "time_period", "required": "Yes", "description": "Future month label."},
        {"field": "resin_price_index", "required": "Yes", "description": "Monthly resin index value."},
        {"field": "freight_regular", "required": "Yes", "description": "Regular China-Buenaventura freight. Prefilled from latest historical value."},
        {"field": "drewry_freight", "required": "Conditional", "description": "Drewry freight value used to calculate incremental freight."},
        {"field": "freight_incremental_override", "required": "Conditional", "description": "Direct incremental freight value. If filled, this overrides Drewry-based calculation."},
        {"field": "finance_factor", "required": "Yes", "description": "Finance factor. Prefilled from latest historical implied factor."},
        {"field": "finance_value_override", "required": "No", "description": "Direct finance value. If filled, this overrides finance factor calculation."},
        {"field": "duty_rate", "required": "Yes", "description": "Duty rate, currently 5%."},
        {"field": "landed_factor_rate", "required": "Yes", "description": "Landed factor rate, currently 8%."},
        {"field": "zf_legislation_change", "required": "Yes", "description": "ZF legislation value; prefilled from latest historical month."},
        {"field": "surcharge_alpek_br", "required": "Yes", "description": "Surcharge value; use 0 when not applicable."},
    ]


def latest_default_rows(defaults: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"field": key, "value": value}
        for key, value in defaults.items()
    ]


def style_header(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def adjust_widths(worksheet: Any) -> None:
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            70,
        )


def write_rows(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return

    headers = list(rows[0])
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    adjust_widths(worksheet)


def build_catalog_workbook(output_file: Path, defaults: dict[str, Any]) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "formula_catalog"
    write_rows(worksheet, catalog_rows(defaults))

    defaults_sheet = workbook.create_sheet("latest_historical_defaults")
    write_rows(defaults_sheet, latest_default_rows(defaults))

    metadata_sheet = workbook.create_sheet("run_metadata")
    write_rows(
        metadata_sheet,
        [
            {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
            {"field": "latest_period", "value": format_period(defaults["latest_period"])},
            {"field": "latest_source_column", "value": defaults["latest_source_column"]},
        ],
    )
    workbook.save(output_file)


def add_future_input_formulas(worksheet: Any, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        worksheet[f"N{row_idx}"] = (
            f'=IF(NOT(ISBLANK(G{row_idx})),G{row_idx},'
            f'IF(ISBLANK(F{row_idx}),"",MAX(F{row_idx}-E{row_idx},0)))'
        )
        worksheet[f"O{row_idx}"] = (
            f'=IF(NOT(ISBLANK(I{row_idx})),I{row_idx},'
            f'IF(OR(ISBLANK(D{row_idx}),ISBLANK(E{row_idx}),ISBLANK(N{row_idx}),ISBLANK(H{row_idx})),"",'
            f'(D{row_idx}+E{row_idx}+N{row_idx})*H{row_idx}))'
        )
        worksheet[f"P{row_idx}"] = (
            f'=IF(OR(ISBLANK(D{row_idx}),ISBLANK(E{row_idx}),ISBLANK(N{row_idx}),ISBLANK(O{row_idx})),"",'
            f'D{row_idx}+O{row_idx}+E{row_idx}+N{row_idx})'
        )
        worksheet[f"Q{row_idx}"] = (
            f'=IF(OR(ISBLANK(D{row_idx}),ISBLANK(E{row_idx}),ISBLANK(O{row_idx})),"",'
            f'D{row_idx}+O{row_idx}+E{row_idx})'
        )
        worksheet[f"R{row_idx}"] = (
            f'=IF(OR(ISBLANK(P{row_idx}),ISBLANK(J{row_idx})),"",P{row_idx}*J{row_idx})'
        )
        worksheet[f"S{row_idx}"] = (
            f'=IF(OR(ISBLANK(Q{row_idx}),ISBLANK(K{row_idx})),"",Q{row_idx}*K{row_idx})'
        )
        worksheet[f"T{row_idx}"] = (
            f'=IF(OR(ISBLANK(P{row_idx}),ISBLANK(R{row_idx}),ISBLANK(S{row_idx}),'
            f'ISBLANK(L{row_idx}),ISBLANK(M{row_idx})),"",P{row_idx}+R{row_idx}+S{row_idx}+L{row_idx}+M{row_idx})'
        )


def build_future_rows(defaults: dict[str, Any], months: int) -> list[dict[str, Any]]:
    start_period = add_months(defaults["latest_period"], 1)
    rows: list[dict[str, Any]] = []
    for offset in range(months):
        period = add_months(start_period, offset)
        rows.append(
            {
                "time_period": format_period(period),
                "time_period_year": period.year,
                "time_period_month": period.month,
                "resin_index_type": defaults["resin_index_type"],
                "resin_price_index": None,
                "freight_regular": defaults["freight_regular"],
                "drewry_freight": None,
                "freight_incremental_override": None,
                "finance_factor": defaults["finance_factor"],
                "finance_value_override": None,
                "duty_rate": defaults["duty_rate"],
                "landed_factor_rate": defaults["landed_factor_rate"],
                "zf_legislation_change": defaults["zf_legislation_change"],
                "surcharge_alpek_br": defaults["surcharge_alpek_br"],
                "freight_incremental_used": None,
                "finance_used": None,
                "sub_total_incremental": None,
                "sub_total_regular": None,
                "duty": None,
                "landed_factor": None,
                "total_landing_cost_forecast": None,
                "notes": None,
            }
        )
    return rows


def color_future_input_columns(worksheet: Any) -> None:
    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    formula_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    input_columns = range(4, 14)
    formula_columns = range(14, 21)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for col_idx in input_columns:
            row[col_idx - 1].fill = input_fill
        for col_idx in formula_columns:
            row[col_idx - 1].fill = formula_fill


def build_template_workbook(output_file: Path, defaults: dict[str, Any], months: int) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "future_inputs"
    future_rows = build_future_rows(defaults, months)
    write_rows(worksheet, future_rows)
    add_future_input_formulas(worksheet, 2, len(future_rows) + 1)
    color_future_input_columns(worksheet)
    worksheet.sheet_view.showGridLines = True

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    definitions_sheet = workbook.create_sheet("field_definitions")
    write_rows(definitions_sheet, field_definition_rows())

    defaults_sheet = workbook.create_sheet("latest_historical_defaults")
    write_rows(defaults_sheet, latest_default_rows(defaults))

    catalog_sheet = workbook.create_sheet("formula_catalog")
    write_rows(catalog_sheet, catalog_rows(defaults))

    metadata_sheet = workbook.create_sheet("run_metadata")
    write_rows(
        metadata_sheet,
        [
            {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
            {"field": "latest_period", "value": format_period(defaults["latest_period"])},
            {"field": "resin_index_type", "value": defaults["resin_index_type"]},
            {"field": "future_months", "value": months},
            {"field": "input_columns_fill", "value": "Yellow cells are user-editable inputs/default assumptions."},
            {"field": "formula_columns_fill", "value": "Green cells are Excel formula outputs."},
        ],
    )
    workbook.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build Bavaria formula catalog and future forecast input template."
    )
    parser.add_argument("--input-file", type=Path, default=DEFAULT_INPUT_FILE)
    parser.add_argument("--final-sheet", default=DEFAULT_FINAL_SHEET)
    parser.add_argument("--catalog-output", type=Path, default=DEFAULT_CATALOG_OUTPUT)
    parser.add_argument("--template-output", type=Path, default=DEFAULT_TEMPLATE_OUTPUT)
    parser.add_argument("--forecast-months", type=int, default=FORECAST_MONTHS)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    _, final_rows = load_final_rows(args.input_file, args.final_sheet)
    defaults = latest_period_metrics(final_rows)
    build_catalog_workbook(args.catalog_output, defaults)
    build_template_workbook(args.template_output, defaults, args.forecast_months)
    print(f"input_file={args.input_file}")
    print(f"catalog_output={args.catalog_output}")
    print(f"template_output={args.template_output}")
    print(f"latest_period={format_period(defaults['latest_period'])}")
    print(f"resin_index_type={defaults['resin_index_type']}")
    print(f"future_months={args.forecast_months}")
    print(f"default_finance_factor={defaults['finance_factor']}")
    print(f"default_duty_rate={defaults['duty_rate']}")
    print(f"default_landed_factor_rate={defaults['landed_factor_rate']}")


if __name__ == "__main__":
    main()
