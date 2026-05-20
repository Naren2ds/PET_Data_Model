from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


DEFAULT_INPUT_FILE = (
    Path("output") / "1 Bavaria - Precio Marzo_formula_final_with_mapping_columns.xlsx"
)
DEFAULT_OUTPUT_FILE = Path("data calculation") / "Bavaria_tlc_calculation_validation.xlsx"
DEFAULT_FINAL_SHEET = "final_data"

RESIN_METRIC = "Resin Price Index"
FINANCE_METRIC = "Finance"
FREIGHT_REGULAR_METRIC = "Freight China-Buenaventura (Regular)"
FREIGHT_INCREMENTAL_METRIC = "Freight China-Buenaventura (Incremental)"
SUBTOTAL_INCREMENTAL_METRIC = "Sub Total (with Incremental Freight)"
SUBTOTAL_REGULAR_METRIC = "Sub Total (with Regular Freight)"
DUTY_METRIC = "Duty 5% (Change According to Regulation)"
LANDED_FACTOR_METRIC = "Landed Factor 8%"
ZF_LEGISLATION_METRIC = "ZF Legislation Change"
SURCHARGE_METRIC = "Sur Charge Alpek Br"
SOURCE_TLC_METRIC = "Total Resin Price ABI VIRGIN Formula"
CALCULATED_TLC_METRIC = "Total Landing Cost Calculated"
CALCULATED_MAPPING_COLUMN = "Total Landing Cost"
RESIN_INDEX_TYPE_COLUMN = "resin_index_type"

MONTH_LABELS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

MONTH_LOOKUP = {
    "jan": 1,
    "january": 1,
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "february": 2,
    "febrero": 2,
    "mar": 3,
    "march": 3,
    "marzo": 3,
    "apr": 4,
    "april": 4,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "june": 6,
    "junio": 6,
    "jul": 7,
    "july": 7,
    "julio": 7,
    "aug": 8,
    "august": 8,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septie": 9,
    "septiembre": 9,
    "set": 9,
    "oct": 10,
    "october": 10,
    "octubre": 10,
    "nov": 11,
    "november": 11,
    "noviembre": 11,
    "dec": 12,
    "december": 12,
    "dic": 12,
    "diciembre": 12,
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
    return value


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def zero_if_missing(value: Any) -> float:
    return numeric_value(value) or 0.0


def parse_period(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    if value is None:
        return None

    text = strip_accents(str(value)).strip()
    if not text:
        return None

    parts = [part for part in re.split(r"[-/\s]+", text) if part]
    if len(parts) < 2:
        return None

    month_token = parts[0].lower()
    month = MONTH_LOOKUP.get(month_token, MONTH_LOOKUP.get(month_token[:3]))
    if month is None:
        return None

    year_token = re.sub(r"\D", "", parts[-1])
    if not year_token:
        return None

    year = int(year_token)
    if year < 100:
        year += 2000

    return date(year, month, 1)


def format_period(value: Any) -> str | None:
    parsed = parse_period(value)
    if parsed is None:
        return None
    return f"{MONTH_LABELS[parsed.month]} {parsed.year}"


def load_final_rows(input_file: Path, final_sheet: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_file, data_only=True, read_only=True)
    try:
        if final_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {final_sheet!r} not found in {input_file}.")

        worksheet = workbook[final_sheet]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        rows = [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
        return headers, rows
    finally:
        workbook.close()


def require_columns(headers: list[str], required_columns: list[str]) -> None:
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValueError(f"Missing required final_data columns: {missing}")


def source_column(source_cell: Any) -> str | None:
    if not source_cell:
        return None
    match = re.match(r"([A-Z]+)", str(source_cell).upper())
    return match.group(1) if match else None


def metric_row_by_source_column(
    final_rows: list[dict[str, Any]],
) -> dict[Any, dict[str, dict[str, Any]]]:
    grouped: dict[Any, dict[str, dict[str, Any]]] = {}
    for row in final_rows:
        group_key = source_column(row.get("source_cell")) or row.get("price_period")
        metric = row.get("metric_en")
        if group_key is None or metric is None:
            continue
        grouped.setdefault(group_key, {})[metric] = row
    return grouped


def metric_value(metrics: dict[str, dict[str, Any]], metric_name: str) -> float:
    return zero_if_missing(metrics.get(metric_name, {}).get("value"))


def metric_formula(metrics: dict[str, dict[str, Any]], metric_name: str) -> Any:
    return metrics.get(metric_name, {}).get("formula")


def metric_source_cell(metrics: dict[str, dict[str, Any]], metric_name: str) -> Any:
    return metrics.get(metric_name, {}).get("source_cell")


def resin_index_type(metrics: dict[str, dict[str, Any]]) -> Any:
    for metric_name in (RESIN_METRIC, SOURCE_TLC_METRIC):
        value = metrics.get(metric_name, {}).get(RESIN_INDEX_TYPE_COLUMN)
        if value:
            return clean_text(value)
    for row in metrics.values():
        value = row.get(RESIN_INDEX_TYPE_COLUMN)
        if value:
            return clean_text(value)
    return None


def build_calculation_rows(
    input_file: Path,
    final_rows: list[dict[str, Any]],
    tolerance: float,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    grouped = metric_row_by_source_column(final_rows)
    validation_rows: list[dict[str, Any]] = []
    calculated_tlc_rows: list[dict[str, Any]] = []
    matched_count = 0
    max_abs_difference = 0.0

    def sort_key(item: tuple[Any, dict[str, dict[str, Any]]]) -> tuple[date, int]:
        group_key, metrics = item
        source_tlc_row = metrics.get(SOURCE_TLC_METRIC, {})
        period = source_tlc_row.get("price_period") or next(
            (
                row.get("price_period")
                for row in metrics.values()
                if row.get("price_period") is not None
            ),
            None,
        )
        parsed_period = parse_period(period) or date.max
        col_idx = 99999
        if isinstance(group_key, str) and re.fullmatch(r"[A-Z]+", group_key):
            col_idx = column_index_from_string(group_key)
        return parsed_period, col_idx

    for group_key, metrics in sorted(
        grouped.items(),
        key=sort_key,
    ):
        source_tlc_row = metrics.get(SOURCE_TLC_METRIC, {})
        period = source_tlc_row.get("price_period") or next(
            (
                row.get("price_period")
                for row in metrics.values()
                if row.get("price_period") is not None
            ),
            None,
        )
        resin = metric_value(metrics, RESIN_METRIC)
        current_resin_index_type = resin_index_type(metrics)
        finance = metric_value(metrics, FINANCE_METRIC)
        freight_regular = metric_value(metrics, FREIGHT_REGULAR_METRIC)
        freight_incremental = metric_value(metrics, FREIGHT_INCREMENTAL_METRIC)
        duty = metric_value(metrics, DUTY_METRIC)
        landed_factor = metric_value(metrics, LANDED_FACTOR_METRIC)
        zf_legislation = metric_value(metrics, ZF_LEGISLATION_METRIC)
        surcharge = metric_value(metrics, SURCHARGE_METRIC)

        calculated_subtotal_incremental = (
            resin + finance + freight_regular + freight_incremental
        )
        source_subtotal_incremental_row = metrics.get(SUBTOTAL_INCREMENTAL_METRIC)
        source_subtotal_incremental = metric_value(metrics, SUBTOTAL_INCREMENTAL_METRIC)
        source_subtotal_regular = metric_value(metrics, SUBTOTAL_REGULAR_METRIC)
        subtotal_incremental_used = (
            source_subtotal_incremental
            if source_subtotal_incremental_row is not None
            else calculated_subtotal_incremental
        )
        calculated_tlc = (
            subtotal_incremental_used
            + duty
            + landed_factor
            + zf_legislation
            + surcharge
        )
        source_tlc = metric_value(metrics, SOURCE_TLC_METRIC)
        difference = calculated_tlc - source_tlc
        abs_difference = abs(difference)
        max_abs_difference = max(max_abs_difference, abs_difference)
        source_tlc_formula = metric_formula(metrics, SOURCE_TLC_METRIC)
        source_tlc_is_formula = (
            isinstance(source_tlc_formula, str)
            and source_tlc_formula.strip().startswith("=")
        )
        if abs_difference <= tolerance:
            status = "MATCH"
        elif not source_tlc_is_formula:
            status = "SOURCE_OVERRIDE"
        else:
            status = "CHECK"
        if status == "MATCH":
            matched_count += 1

        formula_description = (
            f"{SUBTOTAL_INCREMENTAL_METRIC} + {DUTY_METRIC} + "
            f"{LANDED_FACTOR_METRIC} + {ZF_LEGISLATION_METRIC} + {SURCHARGE_METRIC}"
        )
        raw_component_subtotal_formula = (
            f"{RESIN_METRIC} + {FINANCE_METRIC} + {FREIGHT_REGULAR_METRIC} + "
            f"{FREIGHT_INCREMENTAL_METRIC}"
        )
        validation_rows.append(
            {
                "source_column": group_key,
                "price_period": period,
                "time_period": format_period(period),
                "resin_index_type": current_resin_index_type,
                "resin_price_index": resin,
                "finance": finance,
                "freight_regular": freight_regular,
                "freight_incremental": freight_incremental,
                "calculated_subtotal_incremental": calculated_subtotal_incremental,
                "source_subtotal_incremental": source_subtotal_incremental,
                "subtotal_incremental_used": subtotal_incremental_used,
                "source_subtotal_regular": source_subtotal_regular,
                "duty": duty,
                "landed_factor": landed_factor,
                "zf_legislation_change": zf_legislation,
                "surcharge_alpek_br": surcharge,
                "calculated_total_landing_cost": calculated_tlc,
                "source_total_landing_cost": source_tlc,
                "difference": difference,
                "abs_difference": abs_difference,
                "validation_status": status,
                "formula_description": formula_description,
                "raw_component_subtotal_formula": raw_component_subtotal_formula,
                "subtotal_basis": (
                    SUBTOTAL_INCREMENTAL_METRIC
                    if source_subtotal_incremental_row is not None
                    else "Calculated from raw components"
                ),
                "source_tlc_formula": source_tlc_formula,
                "source_tlc_is_formula": source_tlc_is_formula,
                "source_tlc_cell": metric_source_cell(metrics, SOURCE_TLC_METRIC),
            }
        )

        calculated_tlc_rows.append(
            {
                "source_file": source_tlc_row.get("source_file") or str(input_file),
                "source_sheet": source_tlc_row.get("source_sheet"),
                "source_range": source_tlc_row.get("source_range"),
                "source_cell": None,
                "source_column": group_key,
                "metric_row": None,
                "metric_label_source_row": None,
                "metric_en": CALCULATED_TLC_METRIC,
                "metric_local": CALCULATED_TLC_METRIC,
                "resin_index_type": current_resin_index_type,
                "Mapping Columns": CALCULATED_MAPPING_COLUMN,
                "price_period": period,
                "time_period": format_period(period),
                "index_period": None,
                "value": calculated_tlc,
                "formula": (
                    f"{SUBTOTAL_INCREMENTAL_METRIC} + {DUTY_METRIC} + "
                    f"{LANDED_FACTOR_METRIC} + {ZF_LEGISLATION_METRIC} + "
                    f"{SURCHARGE_METRIC}"
                ),
                "source_total_landing_cost": source_tlc,
                "difference": difference,
                "validation_status": status,
            }
        )

    metadata = {
        "periods_calculated": len(validation_rows),
        "matched_periods": matched_count,
        "source_override_periods": sum(
            1 for row in validation_rows if row["validation_status"] == "SOURCE_OVERRIDE"
        ),
        "check_periods": sum(
            1 for row in validation_rows if row["validation_status"] == "CHECK"
        ),
        "max_abs_difference": max_abs_difference,
    }
    return validation_rows, calculated_tlc_rows, metadata


def ordered_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def style_header(worksheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def adjust_widths(worksheet: Any) -> None:
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        width = len(str(header or ""))
        for row_idx in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 12),
            70,
        )


def write_rows_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    headers: list[str] | None = None,
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    if headers is None:
        headers = ordered_headers(rows)
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])
        for cell in worksheet[worksheet.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    adjust_widths(worksheet)


def write_metadata_sheet(workbook: Workbook, metadata: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("run_metadata")
    worksheet.append(["field", "value"])
    for key, value in metadata.items():
        worksheet.append([key, value])

    style_header(worksheet)
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 90


def write_output(
    output_file: Path,
    validation_rows: list[dict[str, Any]],
    calculated_tlc_rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_to_sheet(workbook, "tlc_validation", validation_rows)
    write_rows_to_sheet(workbook, "calculated_tlc_rows", calculated_tlc_rows)
    write_metadata_sheet(workbook, metadata)
    workbook.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate and validate Bavaria Total Landing Cost from extracted components."
    )
    parser.add_argument("--input-file", type=Path, default=DEFAULT_INPUT_FILE)
    parser.add_argument("--final-sheet", default=DEFAULT_FINAL_SHEET)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--tolerance", type=float, default=0.000001)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    headers, final_rows = load_final_rows(args.input_file, args.final_sheet)
    require_columns(
        headers,
        ["source_file", "source_sheet", "source_range", "source_cell", "metric_en", "price_period", "value", "formula"],
    )
    validation_rows, calculated_tlc_rows, calc_metadata = build_calculation_rows(
        args.input_file,
        final_rows,
        args.tolerance,
    )
    metadata = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "input_file": str(args.input_file),
        "final_sheet": args.final_sheet,
        "output_file": str(args.output_file),
        "formula": (
            "Total Landing Cost = Sub Total (with Incremental Freight) + Duty + "
            "Landed Factor + ZF Legislation Change + Sur Charge Alpek Br"
        ),
        "subtotal_formula": (
            "Sub Total (with Incremental Freight) = Resin Price Index + Finance + "
            "Freight China-Buenaventura (Regular) + Freight China-Buenaventura (Incremental). "
            "The extracted subtotal row is used when present because some historical workbook "
            "cells are rounded or hardcoded."
        ),
        "missing_components": "Missing optional components are treated as 0.",
        "tolerance": args.tolerance,
        **calc_metadata,
    }
    write_output(args.output_file, validation_rows, calculated_tlc_rows, metadata)

    print(f"input_file={args.input_file}")
    print(f"output_file={args.output_file}")
    print(f"periods_calculated={calc_metadata['periods_calculated']}")
    print(f"matched_periods={calc_metadata['matched_periods']}")
    print(f"check_periods={calc_metadata['check_periods']}")
    print(f"max_abs_difference={calc_metadata['max_abs_difference']}")


if __name__ == "__main__":
    main()
